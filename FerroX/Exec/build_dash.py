#!/usr/bin/env python3
"""Build a reusable HTML dashboard from one Excel worksheet.

The defaults follow the FerroX folder layout, but the structure name,
workbook, worksheet, case column, displayed columns, image paths, title, and
output file can all be changed from the command line.
"""

from __future__ import annotations

import argparse
import fnmatch
import html
import os
from datetime import date, datetime
from numbers import Number
from pathlib import Path

from openpyxl import load_workbook


# ============================================================
# Default settings
# ============================================================

DEFAULT_STRUCTURE = "MFIS"
DEFAULT_SHEET_NAME = "experiments"
DEFAULT_OUTPUT_NAME = "index.html"
DEFAULT_PZ_IMAGE = "figs/Pz_FE_layer_stack.png"

DEFAULT_COLUMNS = [
    "alpha",
    "beta",
    "gamma",
    "BigGamma",
    "g11",
    "g44",
    "Ec",
    "Pr",
    "Pc0",
    "rp",
    "landau_consistency_pass",
    "valid_loop",
    "has_multi",
]

DEFAULT_CASE_COLUMN_CANDIDATES = [
    "folder",
    "file_name",
    "folder_name",
    "case_key",
    "run_id",
    "case",
]

# Compatibility names for older code that imported these constants.
DEFAULT_EXCEL_NAME = f"{DEFAULT_STRUCTURE}_dataset.xlsx"
WANTED = list(DEFAULT_COLUMNS)


HELP_EPILOG = r"""
預設行為:
  --structure MFIS
  Excel          : ./MFIS_dataset.xlsx
  worksheet      : experiments
  PV image       : figs/MFIS_PV_curve.png
  Pz image       : figs/Pz_FE_layer_stack.png
  output         : ./index.html

範例 1：MFIS 預設設定
  python build_dash_generic_20260814.py

範例 2：改成 MFM
  python build_dash_generic_20260814.py --structure MFM

範例 3：自訂 workbook、worksheet、case 欄位與顯示欄位
  python build_dash_generic_20260814.py \
      --root /home/bowei/FAM/FerroX/Exec \
      --excel custom.xlsx \
      --sheet runs \
      --case-column case_id \
      --columns alpha beta gamma valid_loop has_multi \
      --pv-image figs/custom_loop.png \
      --output custom_dashboard.html

範例 4：只顯示名稱符合 CUSTOM_* 的 case
  python build_dash_generic_20260814.py \
      --excel custom.xlsx \
      --case-glob 'CUSTOM_*'

搜尋欄範例:
  valid_loop=1 has_multi=0
  landau_consistency_pass=True
  has_multi=no_data

其他 Python 程式仍可直接呼叫:
  import build_dash_generic_20260814 as build_dash

  build_dash.build_dashboard(
      "temporary_dataset.xlsx",
      data_root=".",
      output_path="preview.html",
      sheet_name="experiments",
      structure="MFIS",
  )
"""


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="Build a generic HTML dashboard from an Excel worksheet.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=HELP_EPILOG,
    )
    parser.add_argument(
        "--structure",
        default=DEFAULT_STRUCTURE,
        help=(
            "Structure name used by defaults, such as MFIS or MFM "
            f"(default: {DEFAULT_STRUCTURE})."
        ),
    )
    parser.add_argument(
        "--root",
        "--data-root",
        dest="root",
        type=Path,
        default=Path.cwd(),
        help=(
            "Directory containing the workbook and case folders "
            "(default: current directory)."
        ),
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=None,
        help=(
            "Workbook path. Default: <structure>_dataset.xlsx under --root."
        ),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(DEFAULT_OUTPUT_NAME),
        help=f"Output HTML path (default: {DEFAULT_OUTPUT_NAME}).",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET_NAME,
        help=f"Worksheet name (default: {DEFAULT_SHEET_NAME}).",
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=1,
        help="One-based header row number (default: 1).",
    )
    parser.add_argument(
        "--case-column",
        default=None,
        help=(
            "Column containing case-folder names. If omitted, common names "
            "such as folder, file_name, case_key, and run_id are detected."
        ),
    )
    parser.add_argument(
        "--case-glob",
        default="*",
        help="Only include matching case names (default: *).",
    )
    parser.add_argument(
        "--columns",
        nargs="+",
        default=None,
        help=(
            "Excel columns displayed between Folder and T_FE. "
            "Default: the standard FerroX parameter/status columns."
        ),
    )
    parser.add_argument(
        "--tfe-column",
        default="T_FE",
        help="Ferroelectric-thickness column (default: T_FE).",
    )
    parser.add_argument(
        "--start-time-column",
        default="Start Time",
        help="Start-time column (default: Start Time).",
    )
    parser.add_argument(
        "--end-time-column",
        default="End Time",
        help="End-time column (default: End Time).",
    )
    parser.add_argument(
        "--elapsed-column",
        default="Elapsed",
        help="Elapsed-time column (default: Elapsed).",
    )
    parser.add_argument(
        "--pv-image",
        default=None,
        help=(
            "PV image path relative to each case folder. Globs are allowed. "
            "Default: figs/<structure>_PV_curve.png."
        ),
    )
    parser.add_argument(
        "--pz-image",
        default=DEFAULT_PZ_IMAGE,
        help=(
            "Pz image path relative to each case folder. Globs are allowed "
            f"(default: {DEFAULT_PZ_IMAGE})."
        ),
    )
    parser.add_argument(
        "--title",
        default=None,
        help="Dashboard title (default: <structure> Dashboard).",
    )
    return parser.parse_args(argv)


def normalized_name(value):
    if value is None:
        return ""
    return str(value).strip().casefold()


def resolve_under_root(path, root):
    path = Path(path).expanduser()
    if path.is_absolute():
        return path
    return root / path


def infer_structure_from_excel(excel_path):
    """Infer MFIS from names such as MFIS_dataset.xlsx or preview copies."""

    filename = Path(excel_path).name
    marker = "_dataset"
    marker_index = filename.casefold().find(marker)
    if marker_index <= 0:
        return None

    structure = filename[:marker_index].lstrip(".")
    if not structure:
        return None
    return structure


def cleaned_columns(columns):
    result = []
    seen = set()

    for column in columns:
        name = str(column).strip()
        if not name:
            continue

        key = name.casefold()
        if key in seen:
            raise ValueError(f"Duplicate requested column: {name!r}")

        seen.add(key)
        result.append(name)

    if not result:
        raise ValueError("At least one dashboard column is required")

    return result


def format_parameter(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, Number):
        return f"{float(value):.3e}"
    return str(value)


def format_plain(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def find_image(case_folder, relative_pattern):
    """Return the first matching image, or None when it is unavailable."""

    if relative_pattern is None:
        return None

    pattern_text = str(relative_pattern).strip()
    if not pattern_text:
        return None

    pattern_path = Path(pattern_text).expanduser()
    if pattern_path.is_absolute():
        return pattern_path if pattern_path.is_file() else None

    has_glob = any(character in pattern_text for character in "*?[")
    if has_glob:
        matches = sorted(
            path
            for path in case_folder.glob(pattern_text)
            if path.is_file()
        )
        return matches[0] if matches else None

    image_path = case_folder / pattern_path
    return image_path if image_path.is_file() else None


def find_images(case_folder, image_specs=None):
    if image_specs is None:
        image_specs = {
            "PV Curve": f"figs/{DEFAULT_STRUCTURE}_PV_curve.png",
            "Pz Stack": DEFAULT_PZ_IMAGE,
        }

    images = {}
    for title, relative_pattern in image_specs.items():
        images[title] = find_image(case_folder, relative_pattern)
    return images


def read_dashboard_rows(
    excel_path,
    data_root,
    sheet_name=DEFAULT_SHEET_NAME,
    *,
    header_row=1,
    case_column=None,
    case_glob="*",
    columns=None,
    tfe_column="T_FE",
    start_time_column="Start Time",
    end_time_column="End Time",
    elapsed_column="Elapsed",
    image_specs=None,
):
    if header_row < 1:
        raise ValueError("header_row must be at least 1")

    if columns is None:
        columns = DEFAULT_COLUMNS
    columns = cleaned_columns(columns)

    if image_specs is None:
        image_specs = {
            "PV Curve": f"figs/{DEFAULT_STRUCTURE}_PV_curve.png",
            "Pz Stack": DEFAULT_PZ_IMAGE,
        }

    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(
                f"Worksheet {sheet_name!r} was not found. "
                f"Available worksheets: {workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        header_values = next(
            worksheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                values_only=True,
            ),
            (),
        )

        header_lookup = {}
        original_headers = {}
        for index, value in enumerate(header_values):
            key = normalized_name(value)
            if not key:
                continue
            if key in header_lookup:
                raise ValueError(f"Duplicate worksheet header: {value!r}")
            header_lookup[key] = index
            original_headers[key] = str(value).strip()

        if case_column is not None:
            case_key = normalized_name(case_column)
            if case_key not in header_lookup:
                available = ", ".join(original_headers.values())
                raise KeyError(
                    f"Case column {case_column!r} was not found. "
                    f"Available columns: {available}"
                )
        else:
            case_key = next(
                (
                    normalized_name(candidate)
                    for candidate in DEFAULT_CASE_COLUMN_CANDIDATES
                    if normalized_name(candidate) in header_lookup
                ),
                None,
            )
            if case_key is None:
                expected = ", ".join(DEFAULT_CASE_COLUMN_CANDIDATES)
                raise KeyError(
                    "No case-name column was found. Expected one of: "
                    f"{expected}."
                )

        case_column_index = header_lookup[case_key]

        def row_value(values, column_name):
            index = header_lookup.get(normalized_name(column_name))
            if index is None or index >= len(values):
                return None
            return values[index]

        rows = []
        for values in worksheet.iter_rows(
            min_row=header_row + 1,
            values_only=True,
        ):
            if case_column_index >= len(values):
                continue

            raw_case_name = values[case_column_index]
            if raw_case_name is None:
                continue

            case_name = str(raw_case_name).strip()
            if not case_name:
                continue
            if not fnmatch.fnmatchcase(case_name, case_glob):
                continue

            rows.append(
                {
                    "folder": case_name,
                    "values": {
                        column: row_value(values, column)
                        for column in columns
                    },
                    "tfe": row_value(values, tfe_column),
                    "start_time": row_value(values, start_time_column),
                    "end_time": row_value(values, end_time_column),
                    "elapsed": row_value(values, elapsed_column),
                    "images": find_images(
                        Path(data_root) / case_name,
                        image_specs,
                    ),
                }
            )

        return sorted(rows, key=lambda row: row["folder"].casefold())
    finally:
        workbook.close()


def image_href(image_path, output_path):
    relative_path = os.path.relpath(
        Path(image_path).resolve(),
        start=Path(output_path).parent.resolve(),
    )
    return Path(relative_path).as_posix()


def render_dashboard(
    rows,
    output_path,
    *,
    title=None,
    columns=None,
    tfe_label="T_FE",
    start_time_label="Start Time",
    end_time_label="End Time",
    elapsed_label="Elapsed",
    image_specs=None,
):
    if title is None:
        title = f"{DEFAULT_STRUCTURE} Dashboard"
    if columns is None:
        columns = DEFAULT_COLUMNS
    columns = cleaned_columns(columns)
    if image_specs is None:
        image_specs = {
            "PV Curve": f"figs/{DEFAULT_STRUCTURE}_PV_curve.png",
            "Pz Stack": DEFAULT_PZ_IMAGE,
        }

    output_path = Path(output_path)
    last_update = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    page_start = """
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>__TITLE__</title>
<style>
body {
    font-family: Arial, sans-serif;
    margin: 25px;
}
h1 {
    margin-bottom: 10px;
}
.summary {
    color: #000;
    font-size: 18px;
    font-weight: bold;
    margin-bottom: 15px;
}
.note {
    color: #444;
    font-size: 14px;
    margin: 6px 0;
}
input {
    width: 420px;
    padding: 8px;
    font-size: 15px;
    margin: 15px 0;
}
table {
    border-collapse: collapse;
    width: 100%;
}
th {
    position: sticky;
    top: 0;
    background: #f2f2f2;
}
th, td {
    border: 1px solid #bbb;
    padding: 8px;
    text-align: center;
}
tr:nth-child(even) {
    background: #fafafa;
}
tr:hover {
    background: #ffffdd;
}
img {
    max-width: 180px;
    max-height: 120px;
    transition: 0.2s;
}
img:hover {
    transform: scale(1.8);
    position: relative;
    z-index: 100;
}
</style>
</head>
<body>
<h1>__TITLE__</h1>
<div class="summary">
    <div>Last Update: <span id="last_update">__LAST_UPDATE__</span></div>
    <div>Total Experiments: <span id="total_exp">__TOTAL__</span></div>
</div>
<p class="note">* Numeric parameters are displayed in SI units.</p>
<p class="note">* Search example: valid_loop=1 has_multi=0</p>
<p class="note">* Text values are supported: has_multi=no_data</p>
<input id="search" placeholder="Search...">
<table id="tbl">
<thead>
<tr>
<th>Folder</th>
"""
    escaped_title = html.escape(str(title))
    page_start = page_start.replace("__TITLE__", escaped_title)
    page_start = page_start.replace("__LAST_UPDATE__", last_update)
    page_start = page_start.replace("__TOTAL__", str(len(rows)))

    parts = [page_start]
    for column in columns:
        parts.append(f"<th>{html.escape(column)}</th>\n")

    parts.append(f"<th>{html.escape(tfe_label)}</th>\n")
    parts.append(f"<th>{html.escape(start_time_label)}</th>\n")
    parts.append(f"<th>{html.escape(end_time_label)}</th>\n")
    parts.append(f"<th>{html.escape(elapsed_label)}</th>\n")

    for image_title in image_specs:
        parts.append(f"<th>{html.escape(image_title)}</th>\n")

    parts.append("</tr>\n</thead>\n<tbody>\n")

    for row in rows:
        parts.append("<tr>\n")
        parts.append(f"<td>{html.escape(str(row['folder']))}</td>\n")

        values = row["values"]
        for column in columns:
            text_value = format_parameter(values.get(column))
            parts.append(f"<td>{html.escape(text_value)}</td>\n")

        parts.append(f"<td>{html.escape(format_parameter(row['tfe']))}</td>\n")
        parts.append(f"<td>{html.escape(format_plain(row['start_time']))}</td>\n")
        parts.append(f"<td>{html.escape(format_plain(row['end_time']))}</td>\n")
        parts.append(f"<td>{html.escape(format_plain(row['elapsed']))}</td>\n")

        for image_title in image_specs:
            image_path = row["images"].get(image_title)
            if image_path is None:
                parts.append("<td>No Image</td>\n")
                continue

            href = html.escape(
                image_href(image_path, output_path),
                quote=True,
            )
            parts.append(
                f'<td><a href="{href}" target="_blank">'
                f'<img src="{href}"></a></td>\n'
            )

        parts.append("</tr>\n")

    parts.append(
        r"""
</tbody>
</table>
<script>
const search = document.getElementById("search");

function parseValue(text) {
    const cleaned = text.trim();
    const lowered = cleaned.toLowerCase();

    if (lowered === "true") return 1;
    if (lowered === "false") return 0;

    const numberValue = Number(cleaned);
    if (cleaned !== "" && Number.isFinite(numberValue)) {
        return numberValue;
    }
    return lowered;
}

function parseCondition(text) {
    const match = text.match(/^([^<>=\s]+)(<=|>=|=|<|>)(.+)$/);
    if (!match) return null;

    return {
        key: match[1].trim().toLowerCase(),
        operator: match[2],
        value: parseValue(match[3])
    };
}

function valuesEqual(left, right) {
    if (typeof left === "number" && typeof right === "number") {
        return left === right;
    }
    return String(left).toLowerCase() === String(right).toLowerCase();
}

function updateTable() {
    const filter = search.value.trim();
    const tableRows = document.querySelectorAll("#tbl tbody tr");
    const headers = Array.from(
        document.querySelectorAll("#tbl thead th")
    ).map(header => header.textContent.trim().toLowerCase());

    const conditions = filter === ""
        ? []
        : filter.split(/\s+/).map(parseCondition).filter(Boolean);

    let count = 0;

    tableRows.forEach(tableRow => {
        const cells = tableRow.querySelectorAll("td");
        const rowValues = {};

        headers.forEach((header, index) => {
            rowValues[header] = parseValue(cells[index].textContent);
        });

        let show = true;

        for (const condition of conditions) {
            const left = rowValues[condition.key];
            const right = condition.value;

            if (left === undefined) {
                show = false;
                break;
            }

            if (condition.operator === "=") {
                show = valuesEqual(left, right);
            } else if (
                typeof left !== "number" ||
                typeof right !== "number"
            ) {
                show = false;
            } else if (condition.operator === ">") {
                show = left > right;
            } else if (condition.operator === "<") {
                show = left < right;
            } else if (condition.operator === ">=") {
                show = left >= right;
            } else if (condition.operator === "<=") {
                show = left <= right;
            }

            if (!show) break;
        }

        tableRow.style.display = show ? "" : "none";
        if (show) count += 1;
    });

    document.getElementById("total_exp").textContent = count;
}

search.addEventListener("input", updateTable);
updateTable();
</script>
</body>
</html>
"""
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("".join(parts), encoding="utf-8")


def build_dashboard(
    excel_path=None,
    *,
    data_root=".",
    output_path=DEFAULT_OUTPUT_NAME,
    sheet_name=DEFAULT_SHEET_NAME,
    structure=None,
    header_row=1,
    case_column=None,
    case_glob="*",
    columns=None,
    title=None,
    pv_image=None,
    pz_image=DEFAULT_PZ_IMAGE,
    tfe_column="T_FE",
    start_time_column="Start Time",
    end_time_column="End Time",
    elapsed_column="Elapsed",
):
    """Read one Excel workbook and create its HTML dashboard."""

    root = Path(data_root).expanduser().resolve()
    if not root.is_dir():
        raise NotADirectoryError(f"Data root does not exist: {root}")

    requested_structure = None
    if structure is not None:
        requested_structure = str(structure).strip()
        if not requested_structure:
            raise ValueError("structure cannot be empty")

    if excel_path is None:
        structure_name = requested_structure or DEFAULT_STRUCTURE
        excel = root / f"{structure_name}_dataset.xlsx"
    else:
        excel = resolve_under_root(excel_path, root).resolve()
        structure_name = requested_structure
        if structure_name is None:
            structure_name = infer_structure_from_excel(excel)
        if structure_name is None:
            structure_name = DEFAULT_STRUCTURE

    output = resolve_under_root(output_path, root).resolve()

    if not excel.is_file():
        raise FileNotFoundError(f"Workbook does not exist: {excel}")
    if header_row < 1:
        raise ValueError("header_row must be at least 1")
    if not str(case_glob).strip():
        raise ValueError("case_glob cannot be empty")

    if columns is None:
        columns = DEFAULT_COLUMNS
    columns = cleaned_columns(columns)

    if pv_image is None:
        pv_image = f"figs/{structure_name}_PV_curve.png"

    image_specs = {
        "PV Curve": pv_image,
        "Pz Stack": pz_image,
    }

    if title is None:
        title = f"{structure_name} Dashboard"

    rows = read_dashboard_rows(
        excel,
        root,
        sheet_name,
        header_row=header_row,
        case_column=case_column,
        case_glob=case_glob,
        columns=columns,
        tfe_column=tfe_column,
        start_time_column=start_time_column,
        end_time_column=end_time_column,
        elapsed_column=elapsed_column,
        image_specs=image_specs,
    )

    render_dashboard(
        rows,
        output,
        title=title,
        columns=columns,
        tfe_label=tfe_column,
        start_time_label=start_time_column,
        end_time_label=end_time_column,
        elapsed_label=elapsed_column,
        image_specs=image_specs,
    )

    print(f"Generated {output} ({len(rows)} experiments)")
    return len(rows)


def main(argv=None):
    args = parse_args(argv)
    root = args.root.expanduser().resolve()

    try:
        build_dashboard(
            args.excel,
            data_root=root,
            output_path=args.output,
            sheet_name=args.sheet,
            structure=args.structure,
            header_row=args.header_row,
            case_column=args.case_column,
            case_glob=args.case_glob,
            columns=args.columns,
            title=args.title,
            pv_image=args.pv_image,
            pz_image=args.pz_image,
            tfe_column=args.tfe_column,
            start_time_column=args.start_time_column,
            end_time_column=args.end_time_column,
            elapsed_column=args.elapsed_column,
        )
    except (
        FileNotFoundError,
        NotADirectoryError,
        KeyError,
        ValueError,
        OSError,
    ) as error:
        raise SystemExit(f"ERROR: {error}") from error

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

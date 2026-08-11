#!/usr/bin/env python3
"""Prepare and run a batch of FerroX cases sampled in intrinsic EPR space.

Main features
-------------
1. Sample intrinsic ``(Ec, Pr, rp)`` using reusable global designs
   (uniform/LHS/Sobol/maximin) or local EPR variation around an anchor.
2. Convert every EPR point to sixth-order Landau ``(alpha, beta, gamma)`` and
   write the rounded coefficients used by FerroX.
3. Reject an EPR combination already present at the same ferroelectric
   thickness in ``MFIS_dataset.xlsx`` / ``experiments``. The same EPR point at
   a different thickness is allowed.
4. Create each case with empty ``figs`` and ``plts`` directories, plus copied
   ``inputs`` and notebook files.
5. Schedule at most one simulation on each idle GPU and show completion
   progress. External GPU use is respected.
6. Keep the final 200 lines of each ``run.log`` and record batch status CSV.

Use ``--dry-run`` to generate and inspect the plan without creating folders.
Use ``--launch`` on the real run to prepare the folders and immediately start
GPU scheduling.
"""

from __future__ import annotations

import argparse
import csv
import math
import os
import re
import shutil
import subprocess
import sys
import time
from collections import deque
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable


try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - environment-dependent
    raise SystemExit(
        "openpyxl is required for Excel duplicate checking. Install it with: "
        "conda install openpyxl"
    ) from exc

try:
    from utils.landau_EPR_transformer import (
        EPR_to_landau,
        FIRST_ORDER_RP_BOUNDS,
        SECOND_ORDER_RP_BOUNDS,
        get_transition_info,
        landau_to_EPR,
    )
except ImportError as exc:  # pragma: no cover - environment-dependent
    raise SystemExit(
        "Cannot import utils.landau_EPR_transformer. "
        "Place landau_EPR_transformer.py in ./utils beside this script."
    ) from exc

try:
    from utils.sampling_method import (
        EPRBounds,
        EPRPoint,
        GLOBAL_DESIGNS,
        VARIATION_MODES,
        VariationSpec,
        sample_epr_points,
    )
except ImportError as exc:  # pragma: no cover - environment-dependent
    raise SystemExit(
        "Cannot import utils.sampling_method. "
        "Place sampling_method.py in ./utils beside this script."
    ) from exc


NUMBER_PATTERN = r"[+-]?(?:(?:\d+(?:\.\d*)?)|(?:\.\d+))(?:[Ee][+-]?\d+)?"
PARAMETERS = ("alpha", "beta", "gamma")
EPR_PARAMETERS = ("ec", "pr", "rp")


@dataclass(frozen=True)
class Candidate:
    alpha: float
    beta: float
    gamma: float
    Ec: float
    Pr: float
    rp: float
    design: str

    def values(self) -> dict[str, float]:
        return {
            "alpha": self.alpha,
            "beta": self.beta,
            "gamma": self.gamma,
        }

    def epr_values(self) -> dict[str, float]:
        return {
            "ec": self.Ec,
            "pr": self.Pr,
            "rp": self.rp,
        }

    @property
    def Pc0(self) -> float:
        return self.Pr / self.rp

    @property
    def transition_order(self) -> str:
        return get_transition_info(self.beta).order


@dataclass
class RunningJob:
    folder: Path
    folder_name: str
    gpu_index: int
    process: subprocess.Popen
    log_handle: object
    started_at: datetime
    candidate: Candidate


@dataclass(frozen=True)
class GPUStatus:
    index: int
    memory_used_mb: int
    memory_total_mb: int
    utilization_percent: int


# -----------------------------------------------------------------------------
# inputs parsing and editing
# -----------------------------------------------------------------------------
def read_scalar_parameter(text: str, key: str) -> float:
    pattern = re.compile(
        rf"^\s*{re.escape(key)}\s*=\s*({NUMBER_PATTERN})",
        flags=re.MULTILINE,
    )
    match = pattern.search(text)
    if not match:
        raise ValueError(f"Cannot find scalar parameter {key!r} in inputs file")
    return float(match.group(1))


def replace_scalar_parameter(text: str, key: str, value: float) -> str:
    pattern = re.compile(
        rf"^(\s*{re.escape(key)}\s*=\s*)({NUMBER_PATTERN})(.*)$",
        flags=re.MULTILINE,
    )

    replacement_value = f"{value:.12e}"
    updated, count = pattern.subn(
        lambda match: f"{match.group(1)}{replacement_value}{match.group(3)}",
        text,
        count=1,
    )
    if count != 1:
        raise ValueError(
            f"Expected exactly one {key!r} assignment, but replaced {count}"
        )
    return updated


def rounded_parameter(value: float) -> float:
    """Return exactly the value that will be written with .12e formatting."""
    return float(f"{value:.12e}")


# -----------------------------------------------------------------------------
# Excel duplicate checking
# -----------------------------------------------------------------------------
def normalize_header(value: object) -> str:
    return str(value).strip().lower()


def normalize_t_fe(value: object) -> float | None:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None

    # Support either SI metres (for example 7e-9) or a worksheet stored in nm.
    if abs(number) > 1e-6:
        number *= 1e-9
    return number


def load_existing_parameter_sets(
    workbook_path: Path,
    sheet_name: str,
) -> list[dict[str, float]]:
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Dataset workbook not found: {workbook_path}")

    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {sheet_name!r} not found. "
                f"Available sheets: {workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            return []

        column_map = {
            normalize_header(header): index
            for index, header in enumerate(headers)
            if header is not None
        }

        if "t_fe" not in column_map:
            raise ValueError(
                f"Worksheet {sheet_name!r} is missing the 't_fe' column. "
                f"Available columns: {list(column_map)}"
            )

        has_epr_columns = all(
            key in column_map
            for key in EPR_PARAMETERS
        )
        has_landau_columns = all(
            key in column_map
            for key in PARAMETERS
        )

        if not has_epr_columns and not has_landau_columns:
            raise ValueError(
                f"Worksheet {sheet_name!r} must contain either "
                f"{EPR_PARAMETERS} or {PARAMETERS}. "
                f"Available columns: {list(column_map)}"
            )

        records: list[dict[str, float]] = []
        for row in rows:
            t_fe = normalize_t_fe(row[column_map["t_fe"]])
            if t_fe is None or not math.isfinite(t_fe) or t_fe <= 0.0:
                continue

            epr_values: dict[str, float] | None = None

            if has_epr_columns:
                try:
                    candidate_values = {
                        key: float(row[column_map[key]])
                        for key in EPR_PARAMETERS
                    }
                    if all(
                        math.isfinite(value) and value > 0.0
                        for value in candidate_values.values()
                    ):
                        epr_values = candidate_values
                except (TypeError, ValueError, IndexError):
                    pass

            # Fall back to alpha/beta/gamma when an older or incomplete row
            # has no usable EPR values.
            if epr_values is None and has_landau_columns:
                try:
                    landau_values = {
                        key: float(row[column_map[key]])
                        for key in PARAMETERS
                    }
                    converted = landau_to_EPR(**landau_values)
                    epr_values = {
                        "ec": converted.Ec,
                        "pr": converted.Pr,
                        "rp": converted.rp,
                    }
                except (
                    ArithmeticError,
                    OverflowError,
                    TypeError,
                    ValueError,
                    IndexError,
                ):
                    continue

            if epr_values is not None:
                records.append(
                    {
                        **epr_values,
                        "t_fe": t_fe,
                    }
                )

        return records
    finally:
        workbook.close()


def floats_close(
    left: float,
    right: float,
    *,
    rel_tol: float = 1.0e-8,
    abs_tol: float = 0.0,
) -> bool:
    return math.isclose(
        left,
        right,
        rel_tol=rel_tol,
        abs_tol=abs_tol,
    )


def thicknesses_close(left_m: float, right_m: float) -> bool:
    """Compare metre-scale thicknesses without treating all nm values alike."""
    return floats_close(
        left_m,
        right_m,
        rel_tol=1.0e-7,
        abs_tol=1.0e-15,
    )


def candidate_matches_record(
    candidate: Candidate,
    record: dict[str, float],
    t_fe_m: float,
) -> bool:
    try:
        same_epr = all(
            floats_close(
                candidate.epr_values()[key],
                record[key],
            )
            for key in EPR_PARAMETERS
        )
        same_thickness = thicknesses_close(
            t_fe_m,
            record["t_fe"],
        )
    except KeyError:
        return False

    return same_epr and same_thickness


def is_existing_duplicate(
    candidate: Candidate,
    records: Iterable[dict[str, float]],
    t_fe_m: float,
) -> bool:
    return any(
        candidate_matches_record(candidate, record, t_fe_m)
        for record in records
    )


def candidate_key(candidate: Candidate) -> tuple[float, float, float]:
    return tuple(
        candidate.epr_values()[key]
        for key in EPR_PARAMETERS
    )


# -----------------------------------------------------------------------------
# Sampling bridge: physics/domain checks stay here; generic designs live in
# utils/sampling_method.py.
# -----------------------------------------------------------------------------
def validate_epr_sampling_bounds(
    ec_bounds: tuple[float, float],
    pr_bounds: tuple[float, float],
    rp_bounds: tuple[float, float],
) -> None:
    bounds = EPRBounds(ec=ec_bounds, pr=pr_bounds, rp=rp_bounds)
    bounds.validate_positive()

    valid_rp_low = FIRST_ORDER_RP_BOUNDS[0]
    valid_rp_high = SECOND_ORDER_RP_BOUNDS[1]
    if rp_bounds[0] <= valid_rp_low or rp_bounds[1] >= valid_rp_high:
        raise ValueError(
            "rp bounds must lie strictly inside the alpha < 0, gamma > 0 "
            f"branch ({valid_rp_low:.12g}, {valid_rp_high:.12g}); "
            f"got {rp_bounds}"
        )


def make_epr_candidate(
    Ec: float,
    Pr: float,
    rp: float,
    design: str,
) -> Candidate:
    """Convert EPR to the exact rounded coefficients written to inputs."""
    converted = EPR_to_landau(
        Ec=Ec,
        Pr=Pr,
        rp=rp,
    )
    alpha = rounded_parameter(converted.alpha)
    beta = rounded_parameter(converted.beta)
    gamma = rounded_parameter(converted.gamma)

    # The .12e output rounding changes EPR by a tiny amount. Store the EPR
    # corresponding to the coefficients FerroX will actually receive.
    actual_epr = landau_to_EPR(
        alpha=alpha,
        beta=beta,
        gamma=gamma,
    )

    return Candidate(
        alpha=alpha,
        beta=beta,
        gamma=gamma,
        Ec=actual_epr.Ec,
        Pr=actual_epr.Pr,
        rp=actual_epr.rp,
        design=design,
    )


def generate_candidates(
    count: int | None,
    existing_records: list[dict[str, float]],
    t_fe_m: float,
    seed: int,
    design: str,
    ec_bounds: tuple[float, float],
    pr_bounds: tuple[float, float],
    rp_bounds: tuple[float, float],
    ec_scale: str = "linear",
    pr_scale: str = "linear",
    maximin_pool_size: int | None = None,
    max_abs_alpha: float | None = None,
    variation_spec: VariationSpec | None = None,
) -> list[Candidate]:
    """Generate FerroX-ready candidates via the reusable sampling module.

    ``utils.sampling_method`` owns the geometry of uniform/LHS/Sobol/maximin
    and local variation.  This wrapper owns FerroX-specific conversion,
    coefficient rounding, |alpha| filtering, and Excel duplicate rejection.
    """
    if max_abs_alpha is not None and max_abs_alpha <= 0:
        raise ValueError("max_abs_alpha must be positive")

    validate_epr_sampling_bounds(ec_bounds, pr_bounds, rp_bounds)
    bounds = EPRBounds(ec=ec_bounds, pr=pr_bounds, rp=rp_bounds)

    # Maximin needs only same-thickness records as reference geometry.  Points
    # outside the active global/local box are ignored inside sampling_method.
    reference_points: list[tuple[float, float, float]] = []
    for record in existing_records:
        try:
            if not thicknesses_close(t_fe_m, record["t_fe"]):
                continue
            reference_points.append(
                (float(record["ec"]), float(record["pr"]), float(record["rp"]))
            )
        except (KeyError, TypeError, ValueError):
            continue

    candidate_cache: dict[tuple[float, float, float], Candidate] = {}
    seen_candidate_keys: set[tuple[float, float, float]] = set()
    rejected = {
        "conversion": 0,
        "max_abs_alpha": 0,
        "existing_duplicate": 0,
        "rounded_duplicate": 0,
    }

    def accept_point(point: EPRPoint) -> bool:
        raw_key = point.as_tuple()
        try:
            candidate = make_epr_candidate(
                Ec=point.Ec,
                Pr=point.Pr,
                rp=point.rp,
                design=point.design,
            )
        except (ArithmeticError, OverflowError, ValueError):
            rejected["conversion"] += 1
            return False

        if max_abs_alpha is not None and abs(candidate.alpha) > max_abs_alpha:
            rejected["max_abs_alpha"] += 1
            return False

        if is_existing_duplicate(candidate, existing_records, t_fe_m):
            rejected["existing_duplicate"] += 1
            return False

        rounded_key = candidate_key(candidate)
        if rounded_key in seen_candidate_keys:
            rejected["rounded_duplicate"] += 1
            return False

        seen_candidate_keys.add(rounded_key)
        candidate_cache[raw_key] = candidate
        return True

    sampled_points = sample_epr_points(
        count=count,
        design=design,
        bounds=bounds,
        ec_scale=ec_scale,
        pr_scale=pr_scale,
        seed=seed,
        reference_points=reference_points,
        maximin_pool_size=maximin_pool_size,
        variation=variation_spec,
        accept=accept_point,
    )

    candidates = [candidate_cache[point.as_tuple()] for point in sampled_points]

    if design == "maximin" or (
        design == "variation"
        and variation_spec is not None
        and variation_spec.mode == "maximin"
    ):
        print(
            f"Maximin reference set: {len(reference_points)} existing points "
            f"at t_FE={t_fe_m * 1e9:.6g} nm"
        )

    if any(rejected.values()):
        summary = ", ".join(
            f"{name}={value}" for name, value in rejected.items() if value
        )
        print(f"Sampling rejections while building candidate pool: {summary}")

    return candidates


# -----------------------------------------------------------------------------
# Case preparation
# -----------------------------------------------------------------------------
def infer_t_fe_m(prefix: str) -> float | None:
    match = re.search(r"(?:^|_)t_([0-9]+(?:\.[0-9]+)?)(?:_|$)", prefix)
    if not match:
        return None
    return float(match.group(1)) * 1e-9


def resolve_exec_subdirectory(
    exec_dir: Path,
    relative_path: Path,
    option_name: str,
) -> Path:
    """Resolve a user-supplied output directory strictly below exec_dir."""
    if relative_path.is_absolute():
        raise ValueError(
            f"{option_name} must be relative to --exec-dir; "
            f"got absolute path {relative_path}"
        )

    resolved = (exec_dir / relative_path).resolve()
    try:
        resolved.relative_to(exec_dir)
    except ValueError as exc:
        raise ValueError(
            f"{option_name} must stay inside --exec-dir; "
            f"got {relative_path}"
        ) from exc

    return resolved


def prepare_cases(
    exec_dir: Path,
    template_folder: str,
    notebook_template_folder: str | None,
    prefix: str,
    start_index: int,
    candidates: list[Candidate],
    notebook_name: str,
    source_notebook_name: str | None,
    allow_existing_empty: bool,
) -> list[tuple[Path, Candidate]]:
    template_dir = exec_dir / template_folder
    notebook_template_dir = exec_dir / (notebook_template_folder or template_folder)
    template_inputs = template_dir / "inputs"
    template_notebook = notebook_template_dir / (source_notebook_name or notebook_name)

    if not template_inputs.is_file():
        raise FileNotFoundError(f"Template inputs not found: {template_inputs}")
    if not template_notebook.is_file():
        raise FileNotFoundError(f"Template notebook not found: {template_notebook}")

    source_inputs_text = template_inputs.read_text(encoding="utf-8")
    prepared: list[tuple[Path, Candidate]] = []
    clean_prefix = prefix.rstrip("_")

    target_dirs = [
        exec_dir / f"{clean_prefix}_{start_index + offset}"
        for offset in range(len(candidates))
    ]

    existing_nonempty: list[Path] = []
    for target_dir in target_dirs:
        if target_dir.exists():
            if not allow_existing_empty or any(target_dir.iterdir()):
                existing_nonempty.append(target_dir)
    if existing_nonempty:
        preview = "\n".join(f"  - {path}" for path in existing_nonempty[:10])
        raise FileExistsError(
            "Target folders already exist and will not be overwritten:\n"
            f"{preview}"
        )

    for target_dir, candidate in zip(target_dirs, candidates, strict=True):
        target_dir.mkdir(parents=False, exist_ok=allow_existing_empty)
        (target_dir / "figs").mkdir(exist_ok=False)
        (target_dir / "plts").mkdir(exist_ok=False)

        updated_inputs = source_inputs_text
        for key, value in candidate.values().items():
            updated_inputs = replace_scalar_parameter(updated_inputs, key, value)

        (target_dir / "inputs").write_text(updated_inputs, encoding="utf-8")
        shutil.copy2(template_notebook, target_dir / notebook_name)

        prepared.append((target_dir, candidate))

    return prepared


def write_plan_csv(
    path: Path,
    prepared_cases: list[tuple[Path, Candidate]],
    nominal: dict[str, float],
    seed: int,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "folder",
                "design",
                "seed",
                "Ec",
                "Pr",
                "Pc0",
                "rp",
                "transition_order",
                "alpha",
                "beta",
                "gamma",
                "alpha_change_percent",
                "beta_change_percent",
                "gamma_change_percent",
            ],
        )

        writer.writeheader()

        for folder, candidate in prepared_cases:
            writer.writerow(
                {
                    "folder": folder.name,
                    "design": candidate.design,
                    "seed": seed,
                    "Ec": f"{candidate.Ec:.12e}",
                    "Pr": f"{candidate.Pr:.12e}",
                    "Pc0": f"{candidate.Pc0:.12e}",
                    "rp": f"{candidate.rp:.12e}",
                    "transition_order": candidate.transition_order,
                    "alpha": f"{candidate.alpha:.12e}",
                    "beta": f"{candidate.beta:.12e}",
                    "gamma": f"{candidate.gamma:.12e}",
                    "alpha_change_percent": (
                        100.0
                        * (candidate.alpha / nominal["alpha"] - 1.0)
                    ),
                    "beta_change_percent": (
                        100.0
                        * (candidate.beta / nominal["beta"] - 1.0)
                    ),
                    "gamma_change_percent": (
                        100.0
                        * (candidate.gamma / nominal["gamma"] - 1.0)
                    ),
                }
            )


# -----------------------------------------------------------------------------
# GPU scheduling and execution
# -----------------------------------------------------------------------------
def query_gpus() -> list[GPUStatus]:
    command = [
        "nvidia-smi",
        "--query-gpu=index,memory.used,memory.total,utilization.gpu",
        "--format=csv,noheader,nounits",
    ]
    try:
        result = subprocess.run(
            command,
            check=True,
            capture_output=True,
            text=True,
        )
    except FileNotFoundError as exc:
        raise RuntimeError("nvidia-smi was not found") from exc
    except subprocess.CalledProcessError as exc:
        raise RuntimeError(f"nvidia-smi failed: {exc.stderr}") from exc

    statuses: list[GPUStatus] = []
    for line in result.stdout.splitlines():
        if not line.strip():
            continue
        fields = [field.strip() for field in line.split(",")]
        if len(fields) != 4:
            continue
        statuses.append(
            GPUStatus(
                index=int(fields[0]),
                memory_used_mb=int(fields[1]),
                memory_total_mb=int(fields[2]),
                utilization_percent=int(fields[3]),
            )
        )
    return statuses


def idle_gpu_indices(
    statuses: list[GPUStatus],
    occupied_by_batch: set[int],
    max_idle_memory_mb: int,
    max_idle_utilization: int,
) -> list[int]:
    idle = [
        status
        for status in statuses
        if status.index not in occupied_by_batch
        and status.memory_used_mb <= max_idle_memory_mb
        and status.utilization_percent <= max_idle_utilization
    ]
    idle.sort(
        key=lambda status: (
            status.memory_used_mb / max(status.memory_total_mb, 1),
            status.utilization_percent,
            status.index,
        )
    )
    return [status.index for status in idle]


def trim_file_to_last_lines(path: Path, max_lines: int = 200) -> None:
    if not path.is_file():
        return
    with path.open("r", encoding="utf-8", errors="replace") as file:
        final_lines = deque(file, maxlen=max_lines)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8") as file:
        file.writelines(final_lines)
    temporary.replace(path)


def render_progress(completed: int, total: int, running: int, failed: int) -> None:
    width = 30
    fraction = completed / total if total else 1.0
    filled = min(width, int(round(width * fraction)))
    bar = "#" * filled + "-" * (width - filled)
    message = (
        f"\r[{bar}] {completed}/{total} completed | "
        f"running={running} | failed={failed}"
    )
    print(message, end="", flush=True)
    if completed >= total:
        print()


def append_status_row(
    status_csv: Path,
    row: dict[str, object],
) -> None:
    fieldnames = [
        "folder",
        "gpu",
        "status",
        "return_code",
        "started_at",
        "ended_at",
        "elapsed_seconds",
        "Ec",
        "Pr",
        "rp",
        "alpha",
        "beta",
        "gamma",
    ]
    new_file = not status_csv.exists()
    with status_csv.open("a", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        if new_file:
            writer.writeheader()
        writer.writerow(row)


def execute_notebook(folder: Path, notebook_name: str) -> int:
    log_path = folder / "analyze.log"
    command = [
        "jupyter",
        "nbconvert",
        "--to",
        "notebook",
        "--execute",
        "--inplace",
        "--ExecutePreprocessor.timeout=-1",
        notebook_name,
    ]
    with log_path.open("w", encoding="utf-8") as log_file:
        result = subprocess.run(
            command,
            cwd=folder,
            stdout=log_file,
            stderr=subprocess.STDOUT,
            check=False,
        )
    return result.returncode


def run_batch(
    exec_dir: Path,
    prepared_cases: list[tuple[Path, Candidate]],
    executable_name: str,
    notebook_name: str,
    execute_notebook_after: bool,
    poll_seconds: float,
    idle_memory_mb: int,
    idle_utilization: int,
    max_parallel: int | None,
    status_csv: Path,
) -> None:
    executable = exec_dir / executable_name
    if not executable.is_file():
        raise FileNotFoundError(f"FerroX executable not found: {executable}")
    if not os.access(executable, os.X_OK):
        raise PermissionError(f"FerroX executable is not executable: {executable}")

    initial_gpus = query_gpus()
    if not initial_gpus:
        raise RuntimeError("No NVIDIA GPUs were reported by nvidia-smi")

    hardware_limit = len(initial_gpus)
    parallel_limit = hardware_limit if max_parallel is None else min(max_parallel, hardware_limit)
    if parallel_limit <= 0:
        raise ValueError("max_parallel must be positive")

    pending = deque(prepared_cases)
    running: dict[int, RunningJob] = {}
    total = len(prepared_cases)
    completed = 0
    failures = 0
    wait_counter = 0

    print(
        f"Detected GPUs: {[gpu.index for gpu in initial_gpus]}; "
        f"maximum batch parallelism: {parallel_limit}"
    )
    render_progress(completed, total, len(running), failures)

    try:
        while pending or running:
            # Collect completed jobs.
            for gpu_index, job in list(running.items()):
                return_code = job.process.poll()
                if return_code is None:
                    continue

                ended_at = datetime.now().astimezone()
                elapsed = (ended_at - job.started_at).total_seconds()
                job.log_handle.write(f"\nEXIT_CODE={return_code}\n")
                job.log_handle.write(ended_at.strftime("%a %b %d %H:%M:%S %Z %Y") + "\n")
                job.log_handle.flush()
                job.log_handle.close()

                run_log = job.folder / "run.log"
                trim_file_to_last_lines(run_log, max_lines=200)

                final_status = "simulation_ok" if return_code == 0 else "simulation_failed"
                if return_code == 0 and execute_notebook_after:
                    notebook_return_code = execute_notebook(job.folder, notebook_name)
                    if notebook_return_code != 0:
                        final_status = "analysis_failed"
                        return_code = notebook_return_code

                if final_status != "simulation_ok":
                    failures += 1

                append_status_row(
                    status_csv,
                    {
                        "folder": job.folder_name,
                        "gpu": gpu_index,
                        "status": final_status,
                        "return_code": return_code,
                        "started_at": job.started_at.isoformat(),
                        "ended_at": ended_at.isoformat(),
                        "elapsed_seconds": f"{elapsed:.3f}",
                        "Ec": f"{job.candidate.Ec:.12e}",
                        "Pr": f"{job.candidate.Pr:.12e}",
                        "rp": f"{job.candidate.rp:.12e}",
                        "alpha": f"{job.candidate.alpha:.12e}",
                        "beta": f"{job.candidate.beta:.12e}",
                        "gamma": f"{job.candidate.gamma:.12e}",
                    },
                )

                del running[gpu_index]
                completed += 1
                render_progress(completed, total, len(running), failures)
                print(
                    f"  {job.folder_name}: {final_status}, "
                    f"GPU {gpu_index}, elapsed {elapsed / 60.0:.1f} min"
                )

            # Start jobs only on GPUs that are currently idle and not already
            # occupied by this scheduler.
            if pending and len(running) < parallel_limit:
                statuses = query_gpus()
                available = idle_gpu_indices(
                    statuses,
                    occupied_by_batch=set(running),
                    max_idle_memory_mb=idle_memory_mb,
                    max_idle_utilization=idle_utilization,
                )

                slots = parallel_limit - len(running)
                for gpu_index in available[:slots]:
                    if not pending:
                        break
                    folder, candidate = pending.popleft()
                    plts_dir = folder / "plts"
                    run_log = folder / "run.log"
                    log_handle = run_log.open("w", encoding="utf-8")

                    environment = os.environ.copy()
                    environment["CUDA_VISIBLE_DEVICES"] = str(gpu_index)

                    command = [
                        "mpirun",
                        "-x",
                        "CUDA_VISIBLE_DEVICES",
                        "-n",
                        "1",
                        str(executable),
                        "../inputs",
                    ]
                    started_at = datetime.now().astimezone()
                    log_handle.write(
                        f"START: {started_at.isoformat()}\n"
                        f"PHYSICAL_GPU: {gpu_index}\n"
                        f"COMMAND: {' '.join(command)}\n\n"
                    )
                    log_handle.flush()

                    try:
                        process = subprocess.Popen(
                            command,
                            cwd=plts_dir,
                            stdout=log_handle,
                            stderr=subprocess.STDOUT,
                            env=environment,
                            text=True,
                        )
                    except Exception:
                        log_handle.close()
                        raise

                    running[gpu_index] = RunningJob(
                        folder=folder,
                        folder_name=folder.name,
                        gpu_index=gpu_index,
                        process=process,
                        log_handle=log_handle,
                        started_at=started_at,
                        candidate=candidate,
                    )
                    print(f"\nStarted {folder.name} on physical GPU {gpu_index}")
                    render_progress(completed, total, len(running), failures)

                if not available:
                    wait_counter += 1
                    if wait_counter == 1 or wait_counter % 10 == 0:
                        summary = ", ".join(
                            f"GPU {gpu.index}: {gpu.memory_used_mb} MiB, "
                            f"{gpu.utilization_percent}%"
                            for gpu in statuses
                        )
                        print(f"\nWaiting for an idle GPU. {summary}")
                        render_progress(completed, total, len(running), failures)
                else:
                    wait_counter = 0

            if pending or running:
                time.sleep(poll_seconds)

    except KeyboardInterrupt:
        print("\nInterrupted. Terminating simulations started by this batch...")
        for job in running.values():
            job.process.terminate()
        for job in running.values():
            try:
                job.process.wait(timeout=20)
            except subprocess.TimeoutExpired:
                job.process.kill()
            job.log_handle.close()
        raise

    print(f"Batch finished: {total - failures} succeeded, {failures} failed")
    print(f"Status file: {status_csv}")


# -----------------------------------------------------------------------------
# CLI
# -----------------------------------------------------------------------------
EXAMPLES = r"""
Examples
========
Global designs (old behavior remains available)
-----------------------------------------------
1) Random uniform in the requested EPR box:
   python batch_ferrox_epr.py \
     --template-folder MFIS_t_8_nomi_29 \
     --prefix MFIS_t_8_epr_uniform --start-index 1 --count 10 \
     --design uniform \
     --ec-range 5e7 2e9 --pr-range 0.02 0.5 --rp-range 1.4954 1.7320 \
     --dry-run

2) Latin hypercube with logarithmic Ec and Pr coordinates:
   python batch_ferrox_epr.py \
     --template-folder MFIS_t_8_nomi_29 \
     --prefix MFIS_t_8_epr_lhs --start-index 1 --count 16 \
     --design lhs --ec-scale log --pr-scale log \
     --ec-range 5e7 2e9 --pr-range 0.02 0.5 --rp-range 1.4954 1.7320 \
     --dry-run

3) Global maximin augmentation.  Existing same-thickness cases are reference
   points, so the new batch fills large holes instead of merely spacing itself:
   python batch_ferrox_epr.py \
     --template-folder MFIS_t_8_nomi_29 \
     --prefix MFIS_t_8_epr_maximin --start-index 1 --count 10 \
     --design maximin --ec-scale log --pr-scale log \
     --ec-range 5e7 2e9 --pr-range 0.02 0.5 --rp-range 1.4954 1.7320 \
     --maximin-pool-size 10000 --dry-run

4) Sobol low-discrepancy design (requires SciPy):
   python batch_ferrox_epr.py \
     --template-folder MFIS_t_8_nomi_29 \
     --prefix MFIS_t_8_epr_sobol --start-index 1 --count 16 \
     --design sobol --ec-scale log --pr-scale log \
     --ec-range 5e7 2e9 --pr-range 0.02 0.5 --rp-range 1.4954 1.7320 \
     --dry-run

Local EPR variation (recommended for targeted data supplementation)
-------------------------------------------------------------------
5) OAT sensitivity study around the template's own nominal EPR.
   Default spans are Ec +/-10%, Pr +/-10%, rp +/-0.02.  With three parameters
   and one level this creates 6 deterministic cases, so --count can be omitted:
   python batch_ferrox_epr.py \
     --template-folder MFIS_t_8_nomi_29 \
     --prefix MFIS_t_8_oat --start-index 1 \
     --design variation --variation-mode oat \
     --variation-ec-fraction 0.10 --variation-pr-fraction 0.10 \
     --variation-rp-delta 0.02 \
     --ec-range 5e7 2e9 --pr-range 0.02 0.5 --rp-range 1.4954 1.7320 \
     --dry-run

6) Multi-level OAT, e.g. +/-5% and +/-10% for Ec/Pr and proportional rp
   offsets.  --variation-levels multiplies the configured span.  Here levels
   0.5 and 1.0 produce 12 perturbed cases; --variation-include-center adds the
   anchor for a total of 13 (unless a point is rejected as an existing case):
   python batch_ferrox_epr.py \
     --template-folder MFIS_t_8_nomi_29 \
     --prefix MFIS_t_8_oat2 --start-index 1 \
     --design variation --variation-mode oat \
     --variation-levels 0.5 1.0 --variation-include-center \
     --variation-ec-fraction 0.10 --variation-pr-fraction 0.10 \
     --variation-rp-delta 0.02 \
     --ec-range 5e7 2e9 --pr-range 0.02 0.5 --rp-range 1.4954 1.7320 \
     --dry-run

7) Vary only rp around an explicit bad-case anchor.  This is useful when the
   inverse model has a localized rp error and you first want a clean sensitivity
   experiment rather than a 3-D training-data fill:
   python batch_ferrox_epr.py \
     --template-folder MFIS_t_8_nomi_29 \
     --prefix MFIS_t_8_rp_oat --start-index 1 \
     --design variation --variation-mode oat \
     --variation-anchor 1.4e9 0.39 1.62 \
     --variation-parameters rp --variation-levels 0.5 1.0 \
     --variation-rp-delta 0.024 \
     --ec-range 5e7 2e9 --pr-range 0.02 0.5 --rp-range 1.4954 1.7320 \
     --dry-run

8) Local joint LHS around an explicit anchor.  Unlike OAT, all EPR coordinates
   vary together and therefore this is appropriate for adding training data
   after the local sensitivity test says the region is identifiable:
   python batch_ferrox_epr.py \
     --template-folder MFIS_t_8_nomi_29 \
     --prefix MFIS_t_8_local_lhs --start-index 1 --count 10 \
     --design variation --variation-mode lhs \
     --variation-anchor 1.4e9 0.39 1.62 \
     --variation-ec-fraction 0.15 --variation-pr-fraction 0.15 \
     --variation-rp-delta 0.03 \
     --ec-scale log --pr-scale log \
     --ec-range 5e7 2e9 --pr-range 0.02 0.5 --rp-range 1.4954 1.7320 \
     --dry-run

9) Local maximin around a bad-case anchor.  This is the recommended small-batch
   "fill this local hole" mode: the candidate pool is local, while existing
   same-thickness points still repel the new selections:
   python batch_ferrox_epr.py \
     --template-folder MFIS_t_8_nomi_29 \
     --prefix MFIS_t_8_local_maximin --start-index 1 --count 8 \
     --design variation --variation-mode maximin \
     --variation-anchor 6.2e8 0.20 1.72 \
     --variation-ec-fraction 0.20 --variation-pr-fraction 0.20 \
     --variation-rp-delta 0.012 \
     --ec-scale log --pr-scale log --maximin-pool-size 10000 \
     --ec-range 5e7 2e9 --pr-range 0.02 0.5 --rp-range 1.4954 1.7320 \
     --dry-run

10) Specify rp span as a fraction of the FULL --rp-range instead of an absolute
    delta.  For example 0.05 means +/-5% of the full rp interval around anchor:
    python batch_ferrox_epr.py \
      --template-folder MFIS_t_8_nomi_29 \
      --prefix MFIS_t_8_local_sobol --start-index 1 --count 8 \
      --design variation --variation-mode sobol \
      --variation-rp-range-fraction 0.05 \
      --ec-range 5e7 2e9 --pr-range 0.02 0.5 --rp-range 1.4954 1.7320 \
      --dry-run

Multiple template folders
-------------------------
11) Apply the same sampling command to several thickness templates.  If
    --variation-anchor is omitted, each template uses its own nominal EPR as
    the local anchor.  If an explicit anchor is supplied, the same EPR anchor
    is used for every thickness:
    python batch_ferrox_epr.py \
      --template-folder MFIS_t_5_nomi_29 MFIS_t_6_nomi_29 MFIS_t_7_nomi_29 MFIS_t_8_nomi_29 \
      --prefix local --start-index 1 --count 6 \
      --design variation --variation-mode lhs \
      --variation-anchor 6.2e8 0.20 1.72 \
      --variation-ec-fraction 0.10 --variation-pr-fraction 0.10 \
      --variation-rp-delta 0.01 \
      --ec-range 5e7 2e9 --pr-range 0.02 0.5 --rp-range 1.4954 1.7320 \
      --dry-run

Python API
----------
Sampling geometry is reusable without FerroX:

    from utils.sampling_method import EPRBounds, VariationSpec, sample_epr_points

    bounds = EPRBounds(
        ec=(5e7, 2e9),
        pr=(0.02, 0.5),
        rp=(1.4954, 1.7320),
    )
    spec = VariationSpec(
        anchor=(1.4e9, 0.39, 1.62),
        mode="lhs",
        ec_fraction=0.15,
        pr_fraction=0.15,
        rp_delta=0.03,
    )
    points = sample_epr_points(
        count=10,
        design="variation",
        bounds=bounds,
        ec_scale="log",
        pr_scale="log",
        seed=20260712,
        variation=spec,
    )
"""


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=(
            "Sample intrinsic Ec/Pr/rp, convert to Landau coefficients, and "
            "optionally run a FerroX batch. Sampling implementations live in "
            "./utils/sampling_method.py so notebooks and other scripts can "
            "reuse exactly the same designs."
        ),
        epilog=EXAMPLES,
    )
    parser.add_argument(
        "--exec-dir",
        type=Path,
        default=Path("/home/bowei/FAM/FerroX/Exec"),
        help="FerroX Exec directory. Default: %(default)s",
    )
    parser.add_argument(
        "--template-folder",
        nargs="+",
        required=True,
        help=(
            "One or more template folders supplying inputs and structure. "
            "When multiple folders are supplied, --prefix is appended to each "
            "template-folder name."
        ),
    )
    parser.add_argument(
        "--notebook-template-folder",
        default=None,
        help=(
            "Optional folder supplying only the notebook. When omitted, each "
            "--template-folder supplies both inputs and notebook."
        ),
    )
    parser.add_argument(
        "--source-notebook-name",
        default=None,
        help="Source notebook filename; defaults to --notebook-name.",
    )
    parser.add_argument(
        "--notebook-name",
        default="analyze.ipynb",
        help="Notebook filename created inside every new case. Default: %(default)s",
    )
    parser.add_argument(
        "--prefix",
        default="",
        help="Folder name prefix/suffix. Example: MFIS_t_7_epr. Default: empty.",
    )
    parser.add_argument(
        "--start-index",
        required=True,
        type=int,
        help="The first created numeric suffix. Use 18 to create ..._18 first.",
    )
    parser.add_argument(
        "--count",
        type=int,
        default=None,
        help=(
            "Number of cases. Required for global and joint-variation designs. "
            "For variation/oat it is normally omitted because OAT count is "
            "determined by --variation-parameters, --variation-levels and "
            "--variation-include-center."
        ),
    )
    parser.add_argument(
        "--t-fe-nm",
        type=float,
        default=None,
        help=(
            "Thickness used for Excel duplicate comparison; inferred from each "
            "template-folder name when omitted."
        ),
    )
    parser.add_argument("--seed", type=int, default=20260712, help="Sampling RNG seed.")
    parser.add_argument(
        "--design",
        choices=(*GLOBAL_DESIGNS, "variation"),
        default="maximin",
        help=(
            "Top-level EPR design. Existing modes uniform/lhs/maximin are kept; "
            "sobol adds low-discrepancy global sampling; variation activates a "
            "local study around an anchor. Default: %(default)s"
        ),
    )
    parser.add_argument(
        "--ec-range",
        nargs=2,
        required=True,
        type=float,
        metavar=("LOW", "HIGH"),
        help="Hard global Ec bounds in V/m. Local variation is clipped/checked against them.",
    )
    parser.add_argument(
        "--pr-range",
        nargs=2,
        required=True,
        type=float,
        metavar=("LOW", "HIGH"),
        help="Hard global positive Pr bounds in C/m^2.",
    )
    parser.add_argument(
        "--rp-range",
        nargs=2,
        required=True,
        type=float,
        metavar=("LOW", "HIGH"),
        help=(
            "Hard global bounds for rp=Pr/Pc0. They must lie strictly inside "
            f"({FIRST_ORDER_RP_BOUNDS[0]:.8g}, {SECOND_ORDER_RP_BOUNDS[1]:.8g})."
        ),
    )
    parser.add_argument(
        "--ec-scale",
        choices=("linear", "log"),
        default="linear",
        help=(
            "Coordinate used when sampling/normalizing Ec. For broad ranges, "
            "--ec-scale log is usually preferable. Default: %(default)s"
        ),
    )
    parser.add_argument(
        "--pr-scale",
        choices=("linear", "log"),
        default="linear",
        help=(
            "Coordinate used when sampling/normalizing Pr. For broad positive "
            "ranges, --pr-scale log is often preferable. Default: %(default)s"
        ),
    )
    parser.add_argument(
        "--maximin-pool-size",
        type=int,
        default=None,
        help=(
            "Candidate-pool size for global or local maximin. Automatic value "
            "is max(4096, 50*count). Larger values improve the discrete search "
            "at extra CPU cost."
        ),
    )

    variation_group = parser.add_argument_group(
        "local EPR variation",
        "Used only with --design variation. The anchor defaults to the EPR "
        "converted from each template's alpha/beta/gamma.",
    )
    variation_group.add_argument(
        "--variation-mode",
        choices=VARIATION_MODES,
        default="oat",
        help=(
            "oat = one-at-a-time sensitivity points; uniform/lhs/sobol/maximin "
            "= joint sampling inside the local variation box. Default: %(default)s"
        ),
    )
    variation_group.add_argument(
        "--variation-anchor",
        nargs=3,
        type=float,
        metavar=("EC", "PR", "RP"),
        default=None,
        help=(
            "Explicit local anchor (Ec[V/m], Pr[C/m^2], rp). If omitted, each "
            "template folder's nominal alpha/beta/gamma is converted to EPR and "
            "used as its anchor."
        ),
    )
    variation_group.add_argument(
        "--variation-ec-fraction",
        type=float,
        default=0.10,
        help="Ec fractional half-width. 0.10 means anchor Ec +/-10%%. Default: %(default)s",
    )
    variation_group.add_argument(
        "--variation-pr-fraction",
        type=float,
        default=0.10,
        help="Pr fractional half-width. 0.10 means anchor Pr +/-10%%. Default: %(default)s",
    )
    rp_span = variation_group.add_mutually_exclusive_group()
    rp_span.add_argument(
        "--variation-rp-delta",
        type=float,
        default=None,
        help=(
            "Absolute rp half-width, e.g. 0.02 means anchor rp +/-0.02. If neither "
            "rp-span option is given, 0.02 is used."
        ),
    )
    rp_span.add_argument(
        "--variation-rp-range-fraction",
        type=float,
        default=None,
        help=(
            "rp half-width as a fraction of the FULL --rp-range width. Example: "
            "0.05 means +/-5%% of (RP_HIGH-RP_LOW)."
        ),
    )
    variation_group.add_argument(
        "--variation-parameters",
        nargs="+",
        choices=("ec", "pr", "rp"),
        default=["ec", "pr", "rp"],
        help=(
            "Parameters perturbed by OAT. Example: --variation-parameters rp; "
            "or --variation-parameters ec rp. Ignored by joint variation modes."
        ),
    )
    variation_group.add_argument(
        "--variation-levels",
        nargs="+",
        type=float,
        default=[1.0],
        help=(
            "Positive multipliers for OAT spans. With Ec fraction=0.10, levels "
            "0.5 1.0 produce +/-5%% and +/-10%% Ec points. Default: %(default)s"
        ),
    )
    variation_group.add_argument(
        "--variation-include-center",
        action="store_true",
        help=(
            "For OAT, include the anchor itself. Usually unnecessary if the "
            "anchor already exists in the dataset; duplicate checking may reject it."
        ),
    )

    parser.add_argument(
        "--dataset",
        type=Path,
        default=Path("/home/bowei/FAM/FerroX/Exec/MFIS_dataset.xlsx"),
        help="Dataset used for same-thickness EPR duplicate/reference checks. Default: %(default)s",
    )
    parser.add_argument("--dataset-sheet", default="experiments")
    parser.add_argument(
        "--plan-csv-dir",
        type=Path,
        default=Path("epr_plan_csv"),
        help="Relative directory under --exec-dir for batch plan CSV files. Default: %(default)s",
    )
    parser.add_argument(
        "--status-csv-dir",
        type=Path,
        default=Path("epr_status_csv"),
        help=(
            "Relative directory under --exec-dir for combined status CSV files. "
            "Created only with --launch. Default: %(default)s"
        ),
    )
    parser.add_argument(
        "--allow-existing-empty",
        action="store_true",
        help="Allow a target folder only if it already exists and is completely empty.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Generate the plan CSV only; do not create folders or launch simulations.",
    )
    parser.add_argument(
        "--launch",
        action="store_true",
        help="Run simulations immediately after creating the folders.",
    )
    parser.add_argument(
        "--executable",
        default="main3d.gnu.TPROF.MPI.CUDA.ex",
        help="FerroX executable under --exec-dir. Default: %(default)s",
    )
    parser.add_argument(
        "--execute-notebook",
        action="store_true",
        help="Execute the copied notebook after each successful simulation.",
    )
    parser.add_argument("--poll-seconds", type=float, default=15.0)
    parser.add_argument(
        "--gpu-idle-memory-mb",
        type=int,
        default=1000,
        help="A GPU above this used-memory threshold is considered externally busy.",
    )
    parser.add_argument(
        "--gpu-idle-utilization",
        type=int,
        default=10,
        help="A GPU above this utilization percentage is considered externally busy.",
    )
    parser.add_argument(
        "--max-parallel",
        type=int,
        default=None,
        help="Maximum concurrent simulations; default is detected GPU count.",
    )
    parser.add_argument(
        "--max-abs-alpha",
        type=float,
        default=None,
        help=(
            "Maximum allowed |alpha| after coefficient rounding. Points beyond "
            "the limit are rejected before final selection. Omit to disable."
        ),
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()

    exec_dir = args.exec_dir.expanduser().resolve()
    dataset_path = args.dataset.expanduser().resolve()

    if not exec_dir.is_dir():
        raise FileNotFoundError(
            f"Exec directory not found: {exec_dir}"
        )

    plan_csv_dir = resolve_exec_subdirectory(
        exec_dir,
        args.plan_csv_dir,
        "--plan-csv-dir",
    )
    status_csv_dir = resolve_exec_subdirectory(
        exec_dir,
        args.status_csv_dir,
        "--status-csv-dir",
    )

    template_folders: list[str] = args.template_folder
    ec_bounds = (float(args.ec_range[0]), float(args.ec_range[1]))
    pr_bounds = (float(args.pr_range[0]), float(args.pr_range[1]))
    rp_bounds = (float(args.rp_range[0]), float(args.rp_range[1]))

    validate_epr_sampling_bounds(
        ec_bounds,
        pr_bounds,
        rp_bounds,
    )

    if args.count is not None and args.count <= 0:
        raise ValueError("--count must be positive when supplied")
    if args.design != "variation" and args.count is None:
        raise ValueError(f"--design {args.design} requires --count")
    if (
        args.design == "variation"
        and args.variation_mode != "oat"
        and args.count is None
    ):
        raise ValueError(
            f"--design variation --variation-mode {args.variation_mode} requires --count"
        )

    # Excel 只需要讀一次
    existing_records = load_existing_parameter_sets(
        dataset_path,
        args.dataset_sheet,
    )

    print(
        f"Loaded {len(existing_records)} "
        "usable existing EPR records"
    )

    all_prepared: list[tuple[Path, Candidate]] = []

    # --------------------------------------------------------
    # Process each template folder
    # --------------------------------------------------------

    for template_number, template_folder in enumerate(
        template_folders,
        start=1,
    ):
        print()
        print("=" * 70)
        print(
            f"Template "
            f"{template_number}/{len(template_folders)}: "
            f"{template_folder}"
        )
        print("=" * 70)

        # ----------------------------------------------------
        # Output prefix
        #
        # 單一 template：
        #   保持原本 --prefix 的完整意義
        #
        # 多個 templates：
        #   把 --prefix 當 suffix
        #
        # Example:
        #   template_folder = MFIS_t_8_nomi_29
        #   args.prefix     = gamma
        #
        #   case_prefix =
        #   MFIS_t_8_nomi_29_gamma
        # ----------------------------------------------------

        if len(template_folders) == 1:
            case_prefix = args.prefix
        else:
            clean_template = template_folder.rstrip("_")
            clean_suffix = args.prefix.strip("_")

            case_prefix = (
                f"{clean_template}_{clean_suffix}"
                if clean_suffix
                else clean_template
            )

        # ----------------------------------------------------
        # Determine t_FE for this nominal
        # ----------------------------------------------------

        t_fe_m = (
            args.t_fe_nm * 1e-9
            if args.t_fe_nm is not None
            else infer_t_fe_m(template_folder)
        )

        if t_fe_m is None:
            raise ValueError(
                "Cannot infer t_FE from template folder "
                f"{template_folder!r}. "
                "Supply --t-fe-nm explicitly."
            )

        # ----------------------------------------------------
        # Read this nominal inputs file
        # ----------------------------------------------------

        template_inputs = (
            exec_dir
            / template_folder
            / "inputs"
        )

        if not template_inputs.is_file():
            raise FileNotFoundError(
                "Template inputs not found: "
                f"{template_inputs}"
            )

        template_text = template_inputs.read_text(
            encoding="utf-8"
        )

        nominal = {
            key: rounded_parameter(
                read_scalar_parameter(
                    template_text,
                    key,
                )
            )
            for key in PARAMETERS
        }
        nominal_epr = landau_to_EPR(**nominal)

        variation_spec: VariationSpec | None = None
        if args.design == "variation":
            if args.variation_anchor is None:
                variation_anchor = (
                    nominal_epr.Ec,
                    nominal_epr.Pr,
                    nominal_epr.rp,
                )
                anchor_source = f"template {template_folder}"
            else:
                variation_anchor = tuple(float(v) for v in args.variation_anchor)
                anchor_source = "--variation-anchor"

            global_bounds = EPRBounds(
                ec=ec_bounds,
                pr=pr_bounds,
                rp=rp_bounds,
            )
            if not global_bounds.contains(variation_anchor):
                raise ValueError(
                    f"Variation anchor {variation_anchor} from {anchor_source} lies "
                    "outside the hard --ec-range/--pr-range/--rp-range bounds."
                )

            if args.variation_rp_range_fraction is not None:
                if args.variation_rp_range_fraction < 0:
                    raise ValueError("--variation-rp-range-fraction must be >= 0")
                rp_delta = (
                    args.variation_rp_range_fraction
                    * (rp_bounds[1] - rp_bounds[0])
                )
            else:
                rp_delta = (
                    0.02
                    if args.variation_rp_delta is None
                    else args.variation_rp_delta
                )

            variation_spec = VariationSpec(
                anchor=variation_anchor,
                mode=args.variation_mode,
                ec_fraction=args.variation_ec_fraction,
                pr_fraction=args.variation_pr_fraction,
                rp_delta=rp_delta,
                parameters=tuple(args.variation_parameters),
                levels=tuple(args.variation_levels),
                include_center=args.variation_include_center,
            )
            variation_spec.validate()

        print(f"Exec directory: {exec_dir}")
        print(
            f"Inputs template folder: "
            f"{template_folder}"
        )

        print(
            "Notebook template folder: "
            f"{args.notebook_template_folder or template_folder}"
        )

        print(f"Output prefix: {case_prefix}")
        print(f"Nominal Landau parameters: {nominal}")
        print(
            "Nominal EPR: "
            f"Ec={nominal_epr.Ec:.6e}, "
            f"Pr={nominal_epr.Pr:.6e}, "
            f"rp={nominal_epr.rp:.8f}"
        )

        print(
            "Hard EPR sampling bounds: "
            f"Ec={ec_bounds} V/m ({args.ec_scale}), "
            f"Pr={pr_bounds} C/m^2 ({args.pr_scale}), "
            f"rp={rp_bounds} (linear)"
        )
        print(f"Sampling design: {args.design}")
        if variation_spec is not None:
            print(
                "Variation: "
                f"mode={variation_spec.mode}, "
                f"anchor={variation_spec.anchor}, "
                f"Ec +/-{100 * variation_spec.ec_fraction:.3g}%, "
                f"Pr +/-{100 * variation_spec.pr_fraction:.3g}%, "
                f"rp +/-{variation_spec.rp_delta:.6g}, "
                f"parameters={variation_spec.parameters}, "
                f"levels={variation_spec.levels}, "
                f"include_center={variation_spec.include_center}"
            )

        print(
            "Duplicate-check thickness: "
            f"{t_fe_m:.6e} m"
        )

        # ----------------------------------------------------
        # Generate candidates in the requested EPR box. The template's
        # nominal coefficients are retained only for plan-file comparisons.
        # ----------------------------------------------------

        candidates = generate_candidates(
            count=args.count,
            existing_records=existing_records,
            t_fe_m=t_fe_m,
            seed=args.seed,
            design=args.design,
            ec_bounds=ec_bounds,
            pr_bounds=pr_bounds,
            rp_bounds=rp_bounds,
            ec_scale=args.ec_scale,
            pr_scale=args.pr_scale,
            maximin_pool_size=args.maximin_pool_size,
            max_abs_alpha=args.max_abs_alpha,
            variation_spec=variation_spec,
        )

        final_index = (
            args.start_index
            + len(candidates)
            - 1
        )

        safe_prefix = re.sub(
            r"[^A-Za-z0-9_.-]+",
            "_",
            case_prefix,
        )

        plan_csv = (
            plan_csv_dir
            / (
                f"batch_plan_{safe_prefix}_"
                f"{args.start_index}_{final_index}.csv"
            )
        )

        clean_prefix = case_prefix.rstrip("_")

        planned_cases = [
            (
                exec_dir
                / (
                    f"{clean_prefix}_"
                    f"{args.start_index + offset}"
                ),
                candidate,
            )
            for offset, candidate
            in enumerate(candidates)
        ]

        write_plan_csv(
            plan_csv,
            planned_cases,
            nominal,
            args.seed,
        )

        print(f"Plan file: {plan_csv}")

        for folder, candidate in planned_cases:
            print(
                f"  {folder.name}: "
                f"Ec={candidate.Ec:.6e}, "
                f"Pr={candidate.Pr:.6e}, "
                f"rp={candidate.rp:.8f}, "
                f"alpha={candidate.alpha:.6e}, "
                f"beta={candidate.beta:.6e}, "
                f"gamma={candidate.gamma:.6e}, "
                f"design={candidate.design}"
            )

        # ----------------------------------------------------
        # Add candidates to in-memory duplicate records
        #
        # 這些 cases 尚未寫入 Excel，但下一個 nominal
        # 也應將它們視為已存在，避免同一次 batch 在相同厚度產生
        # 重複 EPR。不同厚度仍允許使用相同 EPR。
        # ----------------------------------------------------

        for candidate in candidates:
            planned_record = {
                **candidate.epr_values(),
                "alpha": candidate.alpha,
                "beta": candidate.beta,
                "gamma": candidate.gamma,
                "t_fe": t_fe_m,
            }

            existing_records.append(
                planned_record
            )

        if args.dry_run:
            continue

        # ----------------------------------------------------
        # Prepare folders for this nominal
        # ----------------------------------------------------

        prepared = prepare_cases(
            exec_dir=exec_dir,
            template_folder=template_folder,
            notebook_template_folder=(
                args.notebook_template_folder
            ),
            prefix=case_prefix,
            start_index=args.start_index,
            candidates=candidates,
            notebook_name=args.notebook_name,
            source_notebook_name=(
                args.source_notebook_name
            ),
            allow_existing_empty=(
                args.allow_existing_empty
            ),
        )

        print(
            f"Prepared {len(prepared)} cases "
            f"from {template_folder}"
        )

        all_prepared.extend(prepared)

    # --------------------------------------------------------
    # Finished all nominal folders
    # --------------------------------------------------------

    if args.dry_run:
        print()
        print(
            "Dry run complete: no case folders were created "
            "and no jobs were launched."
        )
        return 0

    print()
    print(
        f"Prepared {len(all_prepared)} total cases "
        f"from {len(template_folders)} nominal folders."
    )

    if not args.launch:
        print(
            "Preparation complete; simulations were not launched."
        )
        return 0

    # 所有 nominal folders 共用一份 status CSV
    status_csv_dir.mkdir(
        parents=True,
        exist_ok=True,
    )
    status_csv = (
        status_csv_dir
        / (
            f"batch_status_multi_"
            f"n{len(template_folders)}_"
            f"seed{args.seed}.csv"
        )
    )

    print(f"Combined status file: {status_csv}")

    run_batch(
        exec_dir=exec_dir,
        prepared_cases=all_prepared,
        executable_name=args.executable,
        notebook_name=args.notebook_name,
        execute_notebook_after=args.execute_notebook,
        poll_seconds=args.poll_seconds,
        idle_memory_mb=args.gpu_idle_memory_mb,
        idle_utilization=args.gpu_idle_utilization,
        max_parallel=args.max_parallel,
        status_csv=status_csv,
    )

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(f"ERROR: {error}", file=sys.stderr)
        raise

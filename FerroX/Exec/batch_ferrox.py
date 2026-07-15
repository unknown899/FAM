#!/usr/bin/env python3
"""Prepare and run a batch of FerroX cases.

Main features
-------------
1. Read nominal alpha/beta/gamma from a template ``inputs`` file.
2. Generate a hybrid design:
   - one-at-a-time end points for alpha, beta and gamma;
   - Latin-hypercube samples for the remaining cases.
3. Reject parameter combinations already present in
   ``MFIS_dataset.xlsx`` / ``experiments``.
4. Create each case with empty ``figs`` and ``plts`` directories, plus copied
   ``inputs`` and notebook files.
5. Schedule at most one simulation on each idle GPU and show completion
   progress. External GPU use is respected.
6. Keep the final 200 lines of each ``run.log`` and record batch status CSV.

Use ``--dry-run`` to generate and inspect the plan without creating folders.
Use ``--launch`` on the real run to prepare the folders and immediately start
GPU scheduling.
"""

from __future__ import annotations

import argparse
import csv
import math
import os
import re
import shutil
import subprocess
import sys
import time
from collections import deque
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

try:
    import numpy as np
except ImportError as exc:  # pragma: no cover - environment-dependent
    raise SystemExit(
        "NumPy is required. Activate the FerroX Conda environment first."
    ) from exc

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - environment-dependent
    raise SystemExit(
        "openpyxl is required for Excel duplicate checking. Install it with: "
        "conda install openpyxl"
    ) from exc


NUMBER_PATTERN = r"[+-]?(?:(?:\d+(?:\.\d*)?)|(?:\.\d+))(?:[Ee][+-]?\d+)?"
PARAMETERS = ("alpha", "beta", "gamma")
OPTIONAL_FIXED_DUPLICATE_KEYS = ("BigGamma", "g11", "g44")
VARIATION_FRACTIONS = {
    "alpha": 0.15,
    "beta": 0.30,
    "gamma": 0.70,
}


@dataclass(frozen=True)
class Candidate:
    alpha: float
    beta: float
    gamma: float
    design: str

    def values(self) -> dict[str, float]:
        return {
            "alpha": self.alpha,
            "beta": self.beta,
            "gamma": self.gamma,
        }


@dataclass
class RunningJob:
    folder: Path
    folder_name: str
    gpu_index: int
    process: subprocess.Popen
    log_handle: object
    started_at: datetime
    candidate: Candidate


@dataclass(frozen=True)
class GPUStatus:
    index: int
    memory_used_mb: int
    memory_total_mb: int
    utilization_percent: int


# -----------------------------------------------------------------------------
# inputs parsing and editing
# -----------------------------------------------------------------------------
def read_scalar_parameter(text: str, key: str) -> float:
    pattern = re.compile(
        rf"^\s*{re.escape(key)}\s*=\s*({NUMBER_PATTERN})",
        flags=re.MULTILINE,
    )
    match = pattern.search(text)
    if not match:
        raise ValueError(f"Cannot find scalar parameter {key!r} in inputs file")
    return float(match.group(1))


def replace_scalar_parameter(text: str, key: str, value: float) -> str:
    pattern = re.compile(
        rf"^(\s*{re.escape(key)}\s*=\s*)({NUMBER_PATTERN})(.*)$",
        flags=re.MULTILINE,
    )

    replacement_value = f"{value:.12e}"
    updated, count = pattern.subn(
        lambda match: f"{match.group(1)}{replacement_value}{match.group(3)}",
        text,
        count=1,
    )
    if count != 1:
        raise ValueError(
            f"Expected exactly one {key!r} assignment, but replaced {count}"
        )
    return updated


def rounded_parameter(value: float) -> float:
    """Return exactly the value that will be written with .12e formatting."""
    return float(f"{value:.12e}")


# -----------------------------------------------------------------------------
# Excel duplicate checking
# -----------------------------------------------------------------------------
def normalize_header(value: object) -> str:
    return str(value).strip().lower()


def normalize_t_fe(value: object) -> float | None:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None

    # Support either SI metres (for example 7e-9) or a worksheet stored in nm.
    if abs(number) > 1e-6:
        number *= 1e-9
    return number


def load_existing_parameter_sets(
    workbook_path: Path,
    sheet_name: str,
) -> list[dict[str, float]]:
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Dataset workbook not found: {workbook_path}")

    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {sheet_name!r} not found. "
                f"Available sheets: {workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            return []

        column_map = {
            normalize_header(header): index
            for index, header in enumerate(headers)
            if header is not None
        }

        missing = [key for key in PARAMETERS if key not in column_map]
        if missing:
            raise ValueError(
                f"Worksheet {sheet_name!r} is missing columns: {missing}. "
                f"Available columns: {list(column_map)}"
            )

        records: list[dict[str, float]] = []
        for row in rows:
            try:
                record = {
                    key: float(row[column_map[key]])
                    for key in PARAMETERS
                }
            except (TypeError, ValueError, IndexError):
                continue

            for original_key in OPTIONAL_FIXED_DUPLICATE_KEYS:
                normalized_key = original_key.lower()
                if normalized_key not in column_map:
                    continue
                try:
                    record[normalized_key] = float(row[column_map[normalized_key]])
                except (TypeError, ValueError, IndexError):
                    pass

            if "t_fe" in column_map:
                t_fe = normalize_t_fe(row[column_map["t_fe"]])
                if t_fe is not None:
                    record["t_fe"] = t_fe

            records.append(record)

        return records
    finally:
        workbook.close()


def floats_close(left: float, right: float, rel_tol: float = 1e-8) -> bool:
    scale = max(abs(left), abs(right), 1.0)
    return abs(left - right) <= rel_tol * scale


def candidate_matches_record(
    candidate: Candidate,
    record: dict[str, float],
    t_fe_m: float,
    fixed_values: dict[str, float],
) -> bool:
    candidate_values = candidate.values()
    if not all(
        floats_close(candidate_values[key], record[key])
        for key in PARAMETERS
    ):
        return False

    # Compare fixed parameters whenever both the inputs template and worksheet
    # provide them. Keys are normalized to lower case in the Excel records.
    for key, value in fixed_values.items():
        normalized_key = key.lower()
        if normalized_key in record and not floats_close(value, record[normalized_key]):
            return False

    # Compare thickness whenever it is available in the worksheet.
    if "t_fe" in record:
        return floats_close(t_fe_m, record["t_fe"], rel_tol=1e-7)
    return True


def is_existing_duplicate(
    candidate: Candidate,
    records: Iterable[dict[str, float]],
    t_fe_m: float,
    fixed_values: dict[str, float],
) -> bool:
    return any(
        candidate_matches_record(candidate, record, t_fe_m, fixed_values)
        for record in records
    )


def candidate_key(candidate: Candidate) -> tuple[float, float, float]:
    return tuple(candidate.values()[key] for key in PARAMETERS)


# -----------------------------------------------------------------------------
# Sampling
# -----------------------------------------------------------------------------
def lhs_unit(n_samples: int, dimensions: int, rng: np.random.Generator) -> np.ndarray:
    """Dependency-free Latin-hypercube points in [0, 1)."""
    points = np.empty((n_samples, dimensions), dtype=float)
    for dimension in range(dimensions):
        values = (np.arange(n_samples) + rng.random(n_samples)) / n_samples
        rng.shuffle(values)
        points[:, dimension] = values
    return points


def make_candidate(
    nominal: dict[str, float],
    factors: dict[str, float],
    design: str,
) -> Candidate:
    values = {
        key: rounded_parameter(nominal[key] * factors[key])
        for key in PARAMETERS
    }
    return Candidate(
        alpha=values["alpha"],
        beta=values["beta"],
        gamma=values["gamma"],
        design=design,
    )


def one_at_a_time_candidates(nominal: dict[str, float]) -> list[Candidate]:
    candidates: list[Candidate] = []
    for key in PARAMETERS:
        variation = VARIATION_FRACTIONS[key]
        for direction, factor in (("low", 1.0 - variation), ("high", 1.0 + variation)):
            factors = {parameter: 1.0 for parameter in PARAMETERS}
            factors[key] = factor
            candidates.append(
                make_candidate(
                    nominal,
                    factors,
                    design=f"OAT_{key}_{direction}",
                )
            )
    return candidates


def landau_spontaneous_p(alpha: float, beta: float, gamma: float) -> float | None:
    """Positive stable stationary polarization of a sixth-order Landau model."""
    if gamma <= 0:
        return None
    discriminant = beta * beta - 4.0 * gamma * alpha
    if discriminant < 0:
        return None

    roots_y = [
        (-beta + math.sqrt(discriminant)) / (2.0 * gamma),
        (-beta - math.sqrt(discriminant)) / (2.0 * gamma),
    ]
    stable: list[float] = []
    for y in roots_y:
        if y <= 0:
            continue
        p = math.sqrt(y)
        second_derivative = alpha + 3.0 * beta * p * p + 5.0 * gamma * p**4
        if second_derivative > 0:
            stable.append(p)
    return max(stable) if stable else None


def landau_intrinsic_ec(alpha: float, beta: float, gamma: float) -> float | None:
    """Approximate intrinsic coercive field from extrema of E(P)."""
    if gamma == 0:
        return None
    discriminant = (3.0 * beta) ** 2 - 20.0 * gamma * alpha
    if discriminant < 0:
        return None

    roots_y = [
        (-3.0 * beta + math.sqrt(discriminant)) / (10.0 * gamma),
        (-3.0 * beta - math.sqrt(discriminant)) / (10.0 * gamma),
    ]
    fields: list[float] = []
    for y in roots_y:
        if y <= 0:
            continue
        p = math.sqrt(y)
        field = alpha * p + beta * p**3 + gamma * p**5
        if math.isfinite(field) and field != 0:
            fields.append(abs(field))
    return min(fields) if fields else None


def passes_optional_screen(
    candidate: Candidate,
    nominal_ps: float | None,
    nominal_ec: float | None,
    ps_ratio: tuple[float, float] | None,
    ec_ratio: tuple[float, float] | None,
) -> bool:
    values = candidate.values()
    ps = landau_spontaneous_p(**values)
    ec = landau_intrinsic_ec(**values)

    if ps_ratio is not None:
        if ps is None or nominal_ps in (None, 0):
            return False
        ratio = ps / nominal_ps
        if not (ps_ratio[0] <= ratio <= ps_ratio[1]):
            return False

    if ec_ratio is not None:
        if ec is None or nominal_ec in (None, 0):
            return False
        ratio = ec / nominal_ec
        if not (ec_ratio[0] <= ratio <= ec_ratio[1]):
            return False

    return True


def generate_candidates(
    count: int,
    nominal: dict[str, float],
    existing_records: list[dict[str, float]],
    t_fe_m: float,
    seed: int,
    design: str,
    fixed_values: dict[str, float],
    ps_ratio: tuple[float, float] | None,
    ec_ratio: tuple[float, float] | None,
) -> list[Candidate]:
    if count <= 0:
        raise ValueError("count must be positive")

    rng = np.random.default_rng(seed)
    accepted: list[Candidate] = []
    accepted_keys: set[tuple[float, float, float]] = set()

    nominal_ps = landau_spontaneous_p(**nominal)
    nominal_ec = landau_intrinsic_ec(**nominal)

    def try_add(candidate: Candidate) -> bool:
        key = candidate_key(candidate)
        if key in accepted_keys:
            return False
        if is_existing_duplicate(candidate, existing_records, t_fe_m, fixed_values):
            return False
        if not passes_optional_screen(
            candidate,
            nominal_ps,
            nominal_ec,
            ps_ratio,
            ec_ratio,
        ):
            return False
        accepted.append(candidate)
        accepted_keys.add(key)
        return True

    if design == "hybrid":
        for candidate in one_at_a_time_candidates(nominal):
            if len(accepted) >= count:
                break
            try_add(candidate)

    attempts = 0
    max_attempts = max(10_000, count * 2_000)
    while len(accepted) < count and attempts < max_attempts:
        remaining = count - len(accepted)
        batch_size = max(remaining * 4, 32)
        unit_points = lhs_unit(batch_size, len(PARAMETERS), rng)

        for point in unit_points:
            factors: dict[str, float] = {}
            for index, key in enumerate(PARAMETERS):
                variation = VARIATION_FRACTIONS[key]
                factors[key] = 1.0 + (2.0 * point[index] - 1.0) * variation

            candidate = make_candidate(
                nominal,
                factors,
                design="LHS",
            )
            attempts += 1
            try_add(candidate)
            if len(accepted) >= count or attempts >= max_attempts:
                break

    if len(accepted) != count:
        raise RuntimeError(
            f"Could generate only {len(accepted)} unique cases out of {count}. "
            "Relax the optional Ps/Ec screening ranges or change the seed."
        )

    return accepted


# -----------------------------------------------------------------------------
# Case preparation
# -----------------------------------------------------------------------------
def infer_t_fe_m(prefix: str) -> float | None:
    match = re.search(r"(?:^|_)t_([0-9]+(?:\.[0-9]+)?)(?:_|$)", prefix)
    if not match:
        return None
    return float(match.group(1)) * 1e-9


def prepare_cases(
    exec_dir: Path,
    template_folder: str,
    notebook_template_folder: str | None,
    prefix: str,
    start_index: int,
    candidates: list[Candidate],
    notebook_name: str,
    source_notebook_name: str | None,
    allow_existing_empty: bool,
) -> list[tuple[Path, Candidate]]:
    template_dir = exec_dir / template_folder
    notebook_template_dir = exec_dir / (notebook_template_folder or template_folder)
    template_inputs = template_dir / "inputs"
    template_notebook = notebook_template_dir / (source_notebook_name or notebook_name)

    if not template_inputs.is_file():
        raise FileNotFoundError(f"Template inputs not found: {template_inputs}")
    if not template_notebook.is_file():
        raise FileNotFoundError(f"Template notebook not found: {template_notebook}")

    source_inputs_text = template_inputs.read_text(encoding="utf-8")
    prepared: list[tuple[Path, Candidate]] = []
    clean_prefix = prefix.rstrip("_")

    target_dirs = [
        exec_dir / f"{clean_prefix}_{start_index + offset}"
        for offset in range(len(candidates))
    ]

    existing_nonempty: list[Path] = []
    for target_dir in target_dirs:
        if target_dir.exists():
            if not allow_existing_empty or any(target_dir.iterdir()):
                existing_nonempty.append(target_dir)
    if existing_nonempty:
        preview = "\n".join(f"  - {path}" for path in existing_nonempty[:10])
        raise FileExistsError(
            "Target folders already exist and will not be overwritten:\n"
            f"{preview}"
        )

    for target_dir, candidate in zip(target_dirs, candidates, strict=True):
        target_dir.mkdir(parents=False, exist_ok=allow_existing_empty)
        (target_dir / "figs").mkdir(exist_ok=False)
        (target_dir / "plts").mkdir(exist_ok=False)

        updated_inputs = source_inputs_text
        for key, value in candidate.values().items():
            updated_inputs = replace_scalar_parameter(updated_inputs, key, value)

        (target_dir / "inputs").write_text(updated_inputs, encoding="utf-8")
        shutil.copy2(template_notebook, target_dir / notebook_name)

        prepared.append((target_dir, candidate))

    return prepared


def write_plan_csv(
    path: Path,
    prepared_cases: list[tuple[Path, Candidate]],
    nominal: dict[str, float],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "folder",
                "design",
                "alpha",
                "beta",
                "gamma",
                "alpha_change_percent",
                "beta_change_percent",
                "gamma_change_percent",
                "landau_ps_estimate",
                "landau_ec_estimate",
            ],
        )
        writer.writeheader()
        for folder, candidate in prepared_cases:
            values = candidate.values()
            writer.writerow(
                {
                    "folder": folder.name,
                    "design": candidate.design,
                    "alpha": f"{candidate.alpha:.12e}",
                    "beta": f"{candidate.beta:.12e}",
                    "gamma": f"{candidate.gamma:.12e}",
                    "alpha_change_percent": 100.0 * (candidate.alpha / nominal["alpha"] - 1.0),
                    "beta_change_percent": 100.0 * (candidate.beta / nominal["beta"] - 1.0),
                    "gamma_change_percent": 100.0 * (candidate.gamma / nominal["gamma"] - 1.0),
                    "landau_ps_estimate": landau_spontaneous_p(**values),
                    "landau_ec_estimate": landau_intrinsic_ec(**values),
                }
            )


# -----------------------------------------------------------------------------
# GPU scheduling and execution
# -----------------------------------------------------------------------------
def query_gpus() -> list[GPUStatus]:
    command = [
        "nvidia-smi",
        "--query-gpu=index,memory.used,memory.total,utilization.gpu",
        "--format=csv,noheader,nounits",
    ]
    try:
        result = subprocess.run(
            command,
            check=True,
            capture_output=True,
            text=True,
        )
    except FileNotFoundError as exc:
        raise RuntimeError("nvidia-smi was not found") from exc
    except subprocess.CalledProcessError as exc:
        raise RuntimeError(f"nvidia-smi failed: {exc.stderr}") from exc

    statuses: list[GPUStatus] = []
    for line in result.stdout.splitlines():
        if not line.strip():
            continue
        fields = [field.strip() for field in line.split(",")]
        if len(fields) != 4:
            continue
        statuses.append(
            GPUStatus(
                index=int(fields[0]),
                memory_used_mb=int(fields[1]),
                memory_total_mb=int(fields[2]),
                utilization_percent=int(fields[3]),
            )
        )
    return statuses


def idle_gpu_indices(
    statuses: list[GPUStatus],
    occupied_by_batch: set[int],
    max_idle_memory_mb: int,
    max_idle_utilization: int,
) -> list[int]:
    idle = [
        status
        for status in statuses
        if status.index not in occupied_by_batch
        and status.memory_used_mb <= max_idle_memory_mb
        and status.utilization_percent <= max_idle_utilization
    ]
    idle.sort(
        key=lambda status: (
            status.memory_used_mb / max(status.memory_total_mb, 1),
            status.utilization_percent,
            status.index,
        )
    )
    return [status.index for status in idle]


def trim_file_to_last_lines(path: Path, max_lines: int = 200) -> None:
    if not path.is_file():
        return
    with path.open("r", encoding="utf-8", errors="replace") as file:
        final_lines = deque(file, maxlen=max_lines)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8") as file:
        file.writelines(final_lines)
    temporary.replace(path)


def render_progress(completed: int, total: int, running: int, failed: int) -> None:
    width = 30
    fraction = completed / total if total else 1.0
    filled = min(width, int(round(width * fraction)))
    bar = "#" * filled + "-" * (width - filled)
    message = (
        f"\r[{bar}] {completed}/{total} completed | "
        f"running={running} | failed={failed}"
    )
    print(message, end="", flush=True)
    if completed >= total:
        print()


def append_status_row(
    status_csv: Path,
    row: dict[str, object],
) -> None:
    fieldnames = [
        "folder",
        "gpu",
        "status",
        "return_code",
        "started_at",
        "ended_at",
        "elapsed_seconds",
        "alpha",
        "beta",
        "gamma",
    ]
    new_file = not status_csv.exists()
    with status_csv.open("a", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        if new_file:
            writer.writeheader()
        writer.writerow(row)


def execute_notebook(folder: Path, notebook_name: str) -> int:
    log_path = folder / "analyze.log"
    command = [
        "jupyter",
        "nbconvert",
        "--to",
        "notebook",
        "--execute",
        "--inplace",
        "--ExecutePreprocessor.timeout=-1",
        notebook_name,
    ]
    with log_path.open("w", encoding="utf-8") as log_file:
        result = subprocess.run(
            command,
            cwd=folder,
            stdout=log_file,
            stderr=subprocess.STDOUT,
            check=False,
        )
    return result.returncode


def run_batch(
    exec_dir: Path,
    prepared_cases: list[tuple[Path, Candidate]],
    executable_name: str,
    notebook_name: str,
    execute_notebook_after: bool,
    poll_seconds: float,
    idle_memory_mb: int,
    idle_utilization: int,
    max_parallel: int | None,
    status_csv: Path,
) -> None:
    executable = exec_dir / executable_name
    if not executable.is_file():
        raise FileNotFoundError(f"FerroX executable not found: {executable}")
    if not os.access(executable, os.X_OK):
        raise PermissionError(f"FerroX executable is not executable: {executable}")

    initial_gpus = query_gpus()
    if not initial_gpus:
        raise RuntimeError("No NVIDIA GPUs were reported by nvidia-smi")

    hardware_limit = len(initial_gpus)
    parallel_limit = hardware_limit if max_parallel is None else min(max_parallel, hardware_limit)
    if parallel_limit <= 0:
        raise ValueError("max_parallel must be positive")

    pending = deque(prepared_cases)
    running: dict[int, RunningJob] = {}
    total = len(prepared_cases)
    completed = 0
    failures = 0
    wait_counter = 0

    print(
        f"Detected GPUs: {[gpu.index for gpu in initial_gpus]}; "
        f"maximum batch parallelism: {parallel_limit}"
    )
    render_progress(completed, total, len(running), failures)

    try:
        while pending or running:
            # Collect completed jobs.
            for gpu_index, job in list(running.items()):
                return_code = job.process.poll()
                if return_code is None:
                    continue

                ended_at = datetime.now().astimezone()
                elapsed = (ended_at - job.started_at).total_seconds()
                job.log_handle.write(f"\nEXIT_CODE={return_code}\n")
                job.log_handle.write(ended_at.strftime("%a %b %d %H:%M:%S %Z %Y") + "\n")
                job.log_handle.flush()
                job.log_handle.close()

                run_log = job.folder / "run.log"
                trim_file_to_last_lines(run_log, max_lines=200)

                final_status = "simulation_ok" if return_code == 0 else "simulation_failed"
                if return_code == 0 and execute_notebook_after:
                    notebook_return_code = execute_notebook(job.folder, notebook_name)
                    if notebook_return_code != 0:
                        final_status = "analysis_failed"
                        return_code = notebook_return_code

                if final_status != "simulation_ok":
                    failures += 1

                append_status_row(
                    status_csv,
                    {
                        "folder": job.folder_name,
                        "gpu": gpu_index,
                        "status": final_status,
                        "return_code": return_code,
                        "started_at": job.started_at.isoformat(),
                        "ended_at": ended_at.isoformat(),
                        "elapsed_seconds": f"{elapsed:.3f}",
                        "alpha": f"{job.candidate.alpha:.12e}",
                        "beta": f"{job.candidate.beta:.12e}",
                        "gamma": f"{job.candidate.gamma:.12e}",
                    },
                )

                del running[gpu_index]
                completed += 1
                render_progress(completed, total, len(running), failures)
                print(
                    f"  {job.folder_name}: {final_status}, "
                    f"GPU {gpu_index}, elapsed {elapsed / 60.0:.1f} min"
                )

            # Start jobs only on GPUs that are currently idle and not already
            # occupied by this scheduler.
            if pending and len(running) < parallel_limit:
                statuses = query_gpus()
                available = idle_gpu_indices(
                    statuses,
                    occupied_by_batch=set(running),
                    max_idle_memory_mb=idle_memory_mb,
                    max_idle_utilization=idle_utilization,
                )

                slots = parallel_limit - len(running)
                for gpu_index in available[:slots]:
                    if not pending:
                        break
                    folder, candidate = pending.popleft()
                    plts_dir = folder / "plts"
                    run_log = folder / "run.log"
                    log_handle = run_log.open("w", encoding="utf-8")

                    environment = os.environ.copy()
                    environment["CUDA_VISIBLE_DEVICES"] = str(gpu_index)

                    command = [
                        "mpirun",
                        "-x",
                        "CUDA_VISIBLE_DEVICES",
                        "-n",
                        "1",
                        str(executable),
                        "../inputs",
                    ]
                    started_at = datetime.now().astimezone()
                    log_handle.write(
                        f"START: {started_at.isoformat()}\n"
                        f"PHYSICAL_GPU: {gpu_index}\n"
                        f"COMMAND: {' '.join(command)}\n\n"
                    )
                    log_handle.flush()

                    try:
                        process = subprocess.Popen(
                            command,
                            cwd=plts_dir,
                            stdout=log_handle,
                            stderr=subprocess.STDOUT,
                            env=environment,
                            text=True,
                        )
                    except Exception:
                        log_handle.close()
                        raise

                    running[gpu_index] = RunningJob(
                        folder=folder,
                        folder_name=folder.name,
                        gpu_index=gpu_index,
                        process=process,
                        log_handle=log_handle,
                        started_at=started_at,
                        candidate=candidate,
                    )
                    print(f"\nStarted {folder.name} on physical GPU {gpu_index}")
                    render_progress(completed, total, len(running), failures)

                if not available:
                    wait_counter += 1
                    if wait_counter == 1 or wait_counter % 10 == 0:
                        summary = ", ".join(
                            f"GPU {gpu.index}: {gpu.memory_used_mb} MiB, "
                            f"{gpu.utilization_percent}%"
                            for gpu in statuses
                        )
                        print(f"\nWaiting for an idle GPU. {summary}")
                        render_progress(completed, total, len(running), failures)
                else:
                    wait_counter = 0

            if pending or running:
                time.sleep(poll_seconds)

    except KeyboardInterrupt:
        print("\nInterrupted. Terminating simulations started by this batch...")
        for job in running.values():
            job.process.terminate()
        for job in running.values():
            try:
                job.process.wait(timeout=20)
            except subprocess.TimeoutExpired:
                job.process.kill()
            job.log_handle.close()
        raise

    print(f"Batch finished: {total - failures} succeeded, {failures} failed")
    print(f"Status file: {status_csv}")


# -----------------------------------------------------------------------------
# CLI
# -----------------------------------------------------------------------------
def ratio_argument(values: list[float] | None) -> tuple[float, float] | None:
    if values is None:
        return None
    low, high = values
    if low <= 0 or high < low:
        raise ValueError("Ratio limits must satisfy 0 < LOW <= HIGH")
    return float(low), float(high)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        description="Prepare and optionally run a FerroX parameter-variation batch.",
    )
    parser.add_argument(
        "--exec-dir",
        type=Path,
        default=Path("/home/bowei/FAM/FerroX/Exec"),
    )
    parser.add_argument(
        "--template-folder",
        default="MFIS_t_8_nomi_33",
        help=(
            "Folder supplying inputs and the notebook. Choose a template with "
            "the correct geometry/thickness for the new batch."
        ),
    )
    parser.add_argument(
        "--notebook-template-folder",
        default=None,
        help=(
            "Optional folder supplying only the notebook. When omitted, "
            "--template-folder supplies both inputs and notebook."
        ),
    )
    parser.add_argument(
        "--source-notebook-name",
        default=None,
        help="Source notebook filename; defaults to --notebook-name.",
    )
    parser.add_argument(
        "--notebook-name",
        default="analyze.ipynb",
        help="Notebook filename created inside every new case.",
    )
    parser.add_argument("--prefix", required=True, help="For example MFIS_t_7_nomi")
    parser.add_argument(
        "--start-index",
        required=True,
        type=int,
        help="The first created suffix. Use 18 to create ..._18 first.",
    )
    parser.add_argument("--count", required=True, type=int)
    parser.add_argument(
        "--t-fe-nm",
        type=float,
        default=None,
        help="Thickness used for Excel duplicate comparison; inferred from prefix when omitted.",
    )
    parser.add_argument("--seed", type=int, default=20260712)
    parser.add_argument(
        "--design",
        choices=("lhs", "hybrid"),
        default="hybrid",
        help="hybrid adds six one-at-a-time endpoints before LHS samples.",
    )
    parser.add_argument(
        "--dataset",
        type=Path,
        default=Path("/home/bowei/FAM/FerroX/Exec/MFIS_dataset.xlsx"),
    )
    parser.add_argument("--dataset-sheet", default="experiments")
    parser.add_argument(
        "--ps-ratio",
        nargs=2,
        type=float,
        metavar=("LOW", "HIGH"),
        default=None,
        help="Optional pre-screen using Landau Ps relative to the nominal value.",
    )
    parser.add_argument(
        "--ec-ratio",
        nargs=2,
        type=float,
        metavar=("LOW", "HIGH"),
        default=None,
        help="Optional pre-screen using intrinsic Ec relative to the nominal value.",
    )
    parser.add_argument(
        "--allow-existing-empty",
        action="store_true",
        help="Allow a target folder only when it already exists but is completely empty.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Generate the plan CSV only; do not create folders or launch simulations.",
    )
    parser.add_argument(
        "--launch",
        action="store_true",
        help="Run simulations immediately after creating the folders.",
    )
    parser.add_argument(
        "--executable",
        default="main3d.gnu.TPROF.MPI.CUDA.ex",
    )
    parser.add_argument(
        "--execute-notebook",
        action="store_true",
        help="Execute the copied notebook after each successful simulation.",
    )
    parser.add_argument("--poll-seconds", type=float, default=15.0)
    parser.add_argument(
        "--gpu-idle-memory-mb",
        type=int,
        default=1000,
        help="A GPU above this used-memory threshold is considered externally busy.",
    )
    parser.add_argument(
        "--gpu-idle-utilization",
        type=int,
        default=10,
        help="A GPU above this utilization percentage is considered externally busy.",
    )
    parser.add_argument(
        "--max-parallel",
        type=int,
        default=None,
        help="Maximum concurrent simulations; default is the GPU count.",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    exec_dir = args.exec_dir.expanduser().resolve()
    dataset_path = args.dataset.expanduser().resolve()

    if not exec_dir.is_dir():
        raise FileNotFoundError(f"Exec directory not found: {exec_dir}")

    t_fe_m = (
        args.t_fe_nm * 1e-9
        if args.t_fe_nm is not None
        else infer_t_fe_m(args.prefix)
    )
    if t_fe_m is None:
        raise ValueError(
            "Cannot infer t_FE from prefix. Supply --t-fe-nm explicitly."
        )

    template_inputs = exec_dir / args.template_folder / "inputs"
    if not template_inputs.is_file():
        raise FileNotFoundError(f"Template inputs not found: {template_inputs}")
    template_text = template_inputs.read_text(encoding="utf-8")
    nominal = {
        key: rounded_parameter(read_scalar_parameter(template_text, key))
        for key in PARAMETERS
    }
    fixed_values: dict[str, float] = {}
    for key in OPTIONAL_FIXED_DUPLICATE_KEYS:
        try:
            fixed_values[key] = rounded_parameter(read_scalar_parameter(template_text, key))
        except ValueError:
            pass

    print(f"Exec directory: {exec_dir}")
    print(f"Inputs template folder: {args.template_folder}")
    print(
        "Notebook template folder: "
        f"{args.notebook_template_folder or args.template_folder}"
    )
    print(f"Nominal parameters: {nominal}")
    print(f"Fixed parameters used for duplicate checking: {fixed_values}")
    print(f"Duplicate-check thickness: {t_fe_m:.6e} m")

    existing_records = load_existing_parameter_sets(
        dataset_path,
        args.dataset_sheet,
    )
    print(f"Loaded {len(existing_records)} existing experiment rows")

    candidates = generate_candidates(
        count=args.count,
        nominal=nominal,
        existing_records=existing_records,
        t_fe_m=t_fe_m,
        seed=args.seed,
        design=args.design,
        fixed_values=fixed_values,
        ps_ratio=ratio_argument(args.ps_ratio),
        ec_ratio=ratio_argument(args.ec_ratio),
    )

    final_index = args.start_index + len(candidates) - 1
    safe_prefix = re.sub(r"[^A-Za-z0-9_.-]+", "_", args.prefix)
    plan_csv = exec_dir / f"batch_plan_{safe_prefix}_{args.start_index}_{final_index}.csv"
    status_csv = exec_dir / f"batch_status_{safe_prefix}_{args.start_index}_{final_index}.csv"

    clean_prefix = args.prefix.rstrip("_")
    planned_cases = [
        (exec_dir / f"{clean_prefix}_{args.start_index + offset}", candidate)
        for offset, candidate in enumerate(candidates)
    ]
    write_plan_csv(plan_csv, planned_cases, nominal)
    print(f"Plan file: {plan_csv}")
    for folder, candidate in planned_cases:
        print(
            f"  {folder.name}: alpha={candidate.alpha:.6e}, "
            f"beta={candidate.beta:.6e}, gamma={candidate.gamma:.6e}, "
            f"design={candidate.design}"
        )

    if args.dry_run:
        print("Dry run complete: no folders were created and no jobs were launched.")
        return 0

    prepared = prepare_cases(
        exec_dir=exec_dir,
        template_folder=args.template_folder,
        notebook_template_folder=args.notebook_template_folder,
        prefix=args.prefix,
        start_index=args.start_index,
        candidates=candidates,
        notebook_name=args.notebook_name,
        source_notebook_name=args.source_notebook_name,
        allow_existing_empty=args.allow_existing_empty,
    )
    print(f"Prepared {len(prepared)} cases")

    if not args.launch:
        print("Preparation complete; simulations were not launched.")
        return 0

    run_batch(
        exec_dir=exec_dir,
        prepared_cases=prepared,
        executable_name=args.executable,
        notebook_name=args.notebook_name,
        execute_notebook_after=args.execute_notebook,
        poll_seconds=args.poll_seconds,
        idle_memory_mb=args.gpu_idle_memory_mb,
        idle_utilization=args.gpu_idle_utilization,
        max_parallel=args.max_parallel,
        status_csv=status_csv,
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(f"ERROR: {error}", file=sys.stderr)
        raise

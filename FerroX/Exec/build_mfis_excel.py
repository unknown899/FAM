#!/usr/bin/env python3
"""
Build an Excel dataset index for FerroX / MFIS simulations.

Output design:
  1. summary          : generation metadata
  2. experiments      : one row per MFIS simulation folder
  3. pv_curve         : long table, one row per voltage point
  4. pz_stack_index   : one row per steady-state Pz slice; actual 2D arrays are saved in compressed .npz files
  5. warnings         : missing files / mismatch messages

Default folder layout expected:
  ./MFIS*/inputs
  ./MFIS*/MFIS_PV_curve.csv or somewhere below MFIS*/
  ./MFIS*/plts/plt########
  ./MFIS*/run.log or ./MFIS*/plts/run.log

Dependencies:
  pip install numpy pandas openpyxl yt
"""

from __future__ import annotations

from pathlib import Path
from datetime import datetime, timedelta
import argparse
import re
import sys

import numpy as np
import pandas as pd


# ==========================
# User settings / defaults
# ==========================

WANTED_PARAMS = [
    "alpha",
    "beta",
    "gamma",
    "BigGamma",
    "g11",
    "g44",
    "FE_lo",
    "FE_hi",
]

P_FIELD = ("boxlib", "Pz")
PHI_FIELD = ("boxlib", "Phi")

CHARGE_FIELD = ("boxlib", "charge")
EPSILON_FIELD = ("boxlib", "epsilon")
MASK_FIELD = ("boxlib", "mask")

# Keep the same convention as your plotting code: P *= -1
P_SIGN = -1.0

# y slice: None means y center
Y_INDEX = None

# If True, only keep FE x range based on FE_lo[0], FE_hi[0].
# If False, keep full x range.
USE_FE_X_RANGE = False

# Plotfile filtering, copied from your Pz stack code
SKIP_INITIAL = True
SKIP_FIRST_N = 9

# Applied gate sweep used if V_applied is not already in MFIS_PV_curve.csv
DEFAULT_VMIN = -4.5
DEFAULT_VMAX = 4.5
DEFAULT_DV = 0.5


# ==========================
# Basic helpers
# ==========================

def parse_value(values: str):
    """Parse FerroX input values like '-8e9' or '0 0 8e-9'."""
    values = values.split("#", 1)[0].strip()
    vals = [float(v) for v in values.split()]
    return vals[0] if len(vals) == 1 else vals


def sci_or_blank(v):
    if v is None or (isinstance(v, float) and not np.isfinite(v)):
        return ""
    return f"{float(v):.6e}"


def get_component(v, idx: int):
    if isinstance(v, (list, tuple)) and len(v) > idx:
        return v[idx]
    if idx == 0 and isinstance(v, (int, float)):
        return v
    return None


def get_step_from_name(name) -> int | None:
    m = re.search(r"plt(\d+)$", Path(name).name)
    return int(m.group(1)) if m else None


def find_plot_names(plot_dir: Path) -> list[str]:
    if not plot_dir.exists():
        return []
    names = [p.name for p in plot_dir.iterdir() if p.is_dir() and re.match(r"plt\d+$", p.name)]
    return sorted(names, key=lambda n: get_step_from_name(n) if get_step_from_name(n) is not None else 10**99)


def find_first_file(folder: Path, filename: str) -> Path | None:
    direct = folder / filename
    if direct.exists():
        return direct
    matches = sorted(folder.rglob(filename))
    return matches[0] if matches else None


def build_default_sweep(vmin=DEFAULT_VMIN, vmax=DEFAULT_VMAX, dv=DEFAULT_DV):
    # Example: -4.5, -4.0, ..., 4.5, 4.0, ..., -4.5 => 37 points
    up = np.arange(vmin, vmax + 0.5 * dv, dv)
    down = np.arange(vmax - dv, vmin - 0.5 * dv, -dv)
    values = np.round(np.concatenate([up, down]), 12)
    branches = ["positive_sweep"] * len(up) + ["negative_sweep"] * len(down)
    return values, branches


def normalize_col(c) -> str:
    return re.sub(r"[^a-z0-9]", "", str(c).strip().lower())


def find_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    """Robustly find a column by normalized names."""
    norm_map = {normalize_col(c): c for c in df.columns}
    norm_candidates = [normalize_col(c) for c in candidates]

    # exact match first
    for cand in norm_candidates:
        if cand in norm_map:
            return norm_map[cand]

    # then partial match, but avoid too-short ambiguous candidates first
    for cand in norm_candidates:
        if len(cand) < 2:
            continue
        for norm, raw in norm_map.items():
            if cand in norm:
                return raw

    return None


# ==========================
# Read original metadata
# ==========================

def read_inputs(path: Path) -> dict:
    params = {}
    with open(path, errors="ignore") as f:
        for line in f:
            line = line.strip()
            if not line or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            if key not in WANTED_PARAMS:
                continue
            try:
                params[key] = parse_value(value)
            except Exception:
                params[key] = value.split("#", 1)[0].strip()
    return params


def parse_date_line(line: str) -> datetime | None:
    """Parse date output like 'Fri Jun 19 00:36:36 CST 2026'.

    Python's %Z parsing for CST is system-dependent, so remove the timezone token.
    """
    line = line.strip()
    m = re.match(
        r"^(?P<dow>\w{3})\s+(?P<mon>\w{3})\s+(?P<day>\d{1,2})\s+"
        r"(?P<hms>\d{2}:\d{2}:\d{2})\s+(?P<tz>\S+)\s+(?P<year>\d{4})$",
        line,
    )
    if not m:
        return None
    s = f"{m.group('dow')} {m.group('mon')} {m.group('day')} {m.group('hms')} {m.group('year')}"
    try:
        return datetime.strptime(s, "%a %b %d %H:%M:%S %Y")
    except Exception:
        return None


def read_run_log(folder: Path):
    logfile = folder / "run.log"
    if not logfile.exists():
        logfile = folder / "plts" / "run.log"
    if not logfile.exists():
        return None, None, None, None

    with open(logfile, errors="ignore") as f:
        lines = [line.strip() for line in f if line.strip()]

    # Keep behavior close to your original script
    lines_tail = lines[-200:]

    end_time = None
    start_time = None
    elapsed_hms = None
    elapsed_sec = None

    if lines_tail:
        end_time = parse_date_line(lines_tail[-1])

    for line in lines_tail:
        if line.startswith("Total run time"):
            m = re.search(r"([\d.]+)", line)
            if m:
                elapsed_sec = float(m.group(1))
            break

    if elapsed_sec is not None:
        if end_time is not None:
            start_time = end_time - timedelta(seconds=elapsed_sec)

        h = int(elapsed_sec // 3600)
        m = int((elapsed_sec % 3600) // 60)
        s = int(elapsed_sec % 60)
        elapsed_hms = f"{h:02d}:{m:02d}:{s:02d}"

    return start_time, end_time, elapsed_hms, elapsed_sec


# ==========================
# Read P-V CSV
# ==========================

def read_pv_curve(folder: Path, run_id: str, sweep_values: np.ndarray, sweep_branches: list[str]):
    csv_path = find_first_file(folder, "MFIS_PV_curve.csv")
    if csv_path is None:
        return None, None, f"{run_id}: MFIS_PV_curve.csv not found"

    try:
        df_raw = pd.read_csv(csv_path)
    except Exception as e:
        return None, str(csv_path), f"{run_id}: failed to read {csv_path}: {e}"

    p_col = find_column(df_raw, ["P_mean", "Pavg", "P_average", "P"])
    vg_col = find_column(df_raw, ["Vg_mean", "Vtop_mean", "V_mean", "Vg", "V"])
    vapplied_col = find_column(df_raw, ["V_applied", "Vapp", "V_app", "gate_voltage", "Vg_applied"])
    step_col = find_column(df_raw, ["step_id", "step", "index"])

    if p_col is None or vg_col is None:
        return None, str(csv_path), (
            f"{run_id}: cannot identify P_mean/Vg_mean columns in {csv_path}. "
            f"Columns = {list(df_raw.columns)}"
        )

    n = len(df_raw)
    out = pd.DataFrame(index=np.arange(n))

    #out["run_id"] = run_id
    out["folder"] = folder.name
    out["point_id"] = np.arange(n, dtype=int)
    #out["step_id"] = df_raw[step_col].values if step_col else np.arange(n, dtype=int)

    if vapplied_col is not None:
        out["V_applied"] = pd.to_numeric(df_raw[vapplied_col], errors="coerce")
        out["branch"] = infer_branch_from_voltage(out["V_applied"].to_numpy())
    elif n == len(sweep_values):
        out["V_applied"] = sweep_values
        out["branch"] = sweep_branches
    else:
        out["V_applied"] = np.nan
        out["branch"] = "unknown"

    #out["Vg_mean"] = pd.to_numeric(df_raw[vg_col], errors="coerce")
    out["P_mean"] = pd.to_numeric(df_raw[p_col], errors="coerce")

    # Keep optional statistics if your CSV already contains them
    optional = {
        "P_min": ["P_min", "Pmin"],
        "P_max": ["P_max", "Pmax"],
        "P_std": ["P_std", "Pstd"],
        "Vg_min": ["Vg_min", "Vmin"],
        "Vg_max": ["Vg_max", "Vmax"],
        "Vg_std": ["Vg_std", "Vstd"],
    }
    
    '''
    
    for out_col, candidates in optional.items():
        c = find_column(df_raw, candidates)
        if c is not None:
            out[out_col] = pd.to_numeric(df_raw[c], errors="coerce")
            
    '''

    #out["pv_csv_path"] = str(csv_path)
    return out, str(csv_path), None


def infer_branch_from_voltage(v: np.ndarray):
    labels = []
    for i, x in enumerate(v):
        if i == 0 or not np.isfinite(x) or not np.isfinite(v[i - 1]):
            labels.append("start")
            continue
        dv = x - v[i - 1]
        if dv > 1e-10:
            labels.append("positive_sweep")
        elif dv < -1e-10:
            labels.append("negative_sweep")
        else:
            labels.append("same_V")
    return labels


# ==========================
# Read Pz steady-state slices from plotfiles
# ==========================

def to_numpy(field_data):
    try:
        return field_data.to_ndarray()
    except AttributeError:
        try:
            return field_data.v
        except AttributeError:
            return np.asarray(field_data)


def read_full_field_from_grids(ds, field):
    """Read a full-domain field.

    First try yt.covering_grid. If that fails, merge individual grids using their
    global start indices. This is intended for level-0 uniform AMReX/BoxLib data.
    """
    dims = tuple(int(x) for x in ds.domain_dimensions)

    try:
        cg = ds.covering_grid(level=0, left_edge=ds.domain_left_edge, dims=dims)
        return to_numpy(cg[field])
    except Exception:
        arr = np.full(dims, np.nan, dtype=np.float64)
        for g in ds.index.grids:
            data = to_numpy(g[field])
            start = np.array(g.get_global_startindex(), dtype=int)
            end = start + np.array(data.shape, dtype=int)
            arr[start[0]:end[0], start[1]:end[1], start[2]:end[2]] = data
        return arr


def read_one_plot(
    plotfile: Path,
    params: dict,
    yt_module,
    y_index=None,
    read_static_fields=False,
):
    ds = yt_module.load(str(plotfile))

    # --------------------------------------------------------
    # Required fields
    # --------------------------------------------------------
    for field in [
        P_FIELD,
        PHI_FIELD,
        CHARGE_FIELD,
    ]:
        if field not in ds.field_list:
            raise RuntimeError(
                f"{plotfile}: field {field} not found"
            )

    # --------------------------------------------------------
    # Read dynamic fields
    # --------------------------------------------------------
    P = read_full_field_from_grids(
        ds,
        P_FIELD,
    ).astype(np.float32)

    Phi = read_full_field_from_grids(
        ds,
        PHI_FIELD,
    ).astype(np.float32)

    charge = read_full_field_from_grids(
        ds,
        CHARGE_FIELD,
    ).astype(np.float32)

    # 保留原本 sign convention
    P *= P_SIGN

    if not (
        P.shape
        == Phi.shape
        == charge.shape
    ):
        raise RuntimeError(
            f"{plotfile}: field shapes do not match\n"
            f"Pz     : {P.shape}\n"
            f"Phi    : {Phi.shape}\n"
            f"charge : {charge.shape}"
        )

    Nx, Ny, Nz = P.shape

    # --------------------------------------------------------
    # Static fields
    # 只需要第一個 bias 讀一次
    # --------------------------------------------------------
    epsilon = None
    mask = None

    if read_static_fields:
        for field in [
            EPSILON_FIELD,
            MASK_FIELD,
        ]:
            if field not in ds.field_list:
                raise RuntimeError(
                    f"{plotfile}: field {field} not found"
                )

        epsilon = read_full_field_from_grids(
            ds,
            EPSILON_FIELD,
        ).astype(np.float32)

        mask = read_full_field_from_grids(
            ds,
            MASK_FIELD,
        ).astype(np.float32)

    # --------------------------------------------------------
    # Coordinates
    # --------------------------------------------------------
    lo = np.asarray(
        ds.domain_left_edge.to_value(),
        dtype=float,
    )

    hi = np.asarray(
        ds.domain_right_edge.to_value(),
        dtype=float,
    )

    dx = (
        (hi - lo)
        / np.array(
            [Nx, Ny, Nz],
            dtype=float,
        )
    )

    x_nm = (
        lo[0]
        + (np.arange(Nx) + 0.5) * dx[0]
    ) * 1e9

    y_nm = (
        lo[1]
        + (np.arange(Ny) + 0.5) * dx[1]
    ) * 1e9

    z_nm_full = (
        lo[2]
        + (np.arange(Nz) + 0.5) * dx[2]
    ) * 1e9

    # --------------------------------------------------------
    # Select y slice
    # --------------------------------------------------------
    if y_index is None:
        y_index_use = Ny // 2
    else:
        y_index_use = int(y_index)

    if not 0 <= y_index_use < Ny:
        raise IndexError(
            f"y_index={y_index_use}, "
            f"but valid range is 0 to {Ny - 1}"
        )

    # --------------------------------------------------------
    # FE region
    # --------------------------------------------------------
    fe_lo = params.get("FE_lo")
    fe_hi = params.get("FE_hi")

    fe_z_lo = get_component(
        fe_lo,
        2,
    )

    fe_z_hi = get_component(
        fe_hi,
        2,
    )

    fe_x_lo = get_component(
        fe_lo,
        0,
    )

    fe_x_hi = get_component(
        fe_hi,
        0,
    )

    # FE z range
    if (
        fe_z_lo is not None
        and fe_z_hi is not None
    ):
        z_lo_nm = (
            min(fe_z_lo, fe_z_hi)
            * 1e9
        )

        z_hi_nm = (
            max(fe_z_lo, fe_z_hi)
            * 1e9
        )

        z_sel_fe = (
            (z_nm_full >= z_lo_nm)
            & (z_nm_full <= z_hi_nm)
        )

    else:
        z_sel_fe = np.ones_like(
            z_nm_full,
            dtype=bool,
        )

    # x range
    if (
        USE_FE_X_RANGE
        and fe_x_lo is not None
        and fe_x_hi is not None
    ):
        x_lo_nm = (
            min(fe_x_lo, fe_x_hi)
            * 1e9
        )

        x_hi_nm = (
            max(fe_x_lo, fe_x_hi)
            * 1e9
        )

        x_sel = (
            (x_nm >= x_lo_nm)
            & (x_nm <= x_hi_nm)
        )

    else:
        x_sel = np.ones_like(
            x_nm,
            dtype=bool,
        )

    if not np.any(x_sel):
        raise RuntimeError(
            f"{plotfile}: no cells selected in x direction"
        )

    if not np.any(z_sel_fe):
        raise RuntimeError(
            f"{plotfile}: no cells selected in FE z direction"
        )

    # --------------------------------------------------------
    # Pz -> FE only
    # --------------------------------------------------------
    P_slice = P[
        np.ix_(
            x_sel,
            [y_index_use],
            z_sel_fe,
        )
    ][:, 0, :]

    # --------------------------------------------------------
    # Phi -> full device
    # --------------------------------------------------------
    Phi_slice = Phi[
        x_sel,
        y_index_use,
        :
    ]

    # --------------------------------------------------------
    # charge -> full device
    # --------------------------------------------------------
    charge_slice = charge[
        x_sel,
        y_index_use,
        :
    ]

    # --------------------------------------------------------
    # Static fields -> full device
    # --------------------------------------------------------
    epsilon_slice = None
    mask_slice = None

    if read_static_fields:
        epsilon_slice = epsilon[
            x_sel,
            y_index_use,
            :
        ]

        mask_slice = mask[
            x_sel,
            y_index_use,
            :
        ]

    return {
        # FE only
        "P_slice": P_slice,

        # full device
        "Phi_slice": Phi_slice,
        "charge_slice": charge_slice,

        # static, full device
        "epsilon_slice": epsilon_slice,
        "mask_slice": mask_slice,

        # coordinates
        "x_nm": x_nm[x_sel],
        "z_nm": z_nm_full[z_sel_fe],
        "z_nm_full": z_nm_full,

        "y_index": y_index_use,
        "y_nm": float(
            y_nm[y_index_use]
        ),

        "full_shape": tuple(
            int(v)
            for v in P.shape
        ),

        # Pz FE
        "slice_shape": tuple(
            int(v)
            for v in P_slice.shape
        ),

        # Phi / charge full
        "full_slice_shape": tuple(
            int(v)
            for v in Phi_slice.shape
        ),
    }
    
def load_existing_pz_npz(
    npz_path: Path,
    folder: Path,
    run_id: str,
    pv_df: pd.DataFrame | None,
):
    """
    原始 plotfiles 已不存在，但 NPZ 還存在時，
    直接利用 NPZ 重建 pz_stack_index。
    """

    with np.load(
        npz_path,
        allow_pickle=False,
    ) as data:

        if "Pz_stack" not in data.files:
            raise RuntimeError(
                f"{run_id}: "
                f"'Pz_stack' not found in {npz_path}"
            )

        Pz_stack = data["Pz_stack"]

        if Pz_stack.ndim != 3:
            raise RuntimeError(
                f"{run_id}: unexpected "
                f"Pz_stack shape {Pz_stack.shape}"
            )

        n_stacks = Pz_stack.shape[0]

        pz_slice_shape = tuple(
            int(v)
            for v in Pz_stack.shape[1:]
        )

        has_phi = (
            "Phi_stack"
            in data.files
        )

        y_index = (
            int(
                np.asarray(
                    data["y_index"]
                ).item()
            )
            if "y_index" in data.files
            else -1
        )

        y_nm = (
            float(
                np.asarray(
                    data["y_nm"]
                ).item()
            )
            if "y_nm" in data.files
            else np.nan
        )

        # ----------------------------------------------------
        # NPZ metadata fallback
        # ----------------------------------------------------
        V_npz = (
            np.asarray(
                data["V_applied"]
            )
            if "V_applied" in data.files
            else None
        )

        P_mean_npz = (
            np.asarray(
                data["P_mean"]
            )
            if "P_mean" in data.files
            else None
        )

        branch_npz = (
            np.asarray(
                data["branch"]
            ).astype(str)
            if "branch" in data.files
            else None
        )

        index_rows = []

        for i in range(n_stacks):

            # 優先使用目前 CSV 的 metadata
            if (
                pv_df is not None
                and i < len(pv_df)
            ):
                pv_row = pv_df.iloc[i]

                point_id = int(
                    pv_row.get(
                        "point_id",
                        i,
                    )
                )

                V_applied = pv_row.get(
                    "V_applied",
                    np.nan,
                )

                P_mean = pv_row.get(
                    "P_mean",
                    np.nan,
                )

                branch = pv_row.get(
                    "branch",
                    "",
                )

            else:
                point_id = i

                V_applied = (
                    V_npz[i]
                    if (
                        V_npz is not None
                        and i < len(V_npz)
                    )
                    else np.nan
                )

                P_mean = (
                    P_mean_npz[i]
                    if (
                        P_mean_npz is not None
                        and i < len(P_mean_npz)
                    )
                    else np.nan
                )

                branch = (
                    branch_npz[i]
                    if (
                        branch_npz is not None
                        and i < len(branch_npz)
                    )
                    else "unknown"
                )

            index_rows.append({
                "folder": folder.name,
                "point_id": point_id,

                "branch": branch,
                "V_applied": V_applied,
                "P_mean": P_mean,

                "phi_available": has_phi,

                "y_index": y_index,
                "y_nm": y_nm,

                # NPZ 沒有保存原始 3D full_shape，
                # 所以不要亂填錯誤資訊。
                "full_shape": "",

                "slice_shape_Nx_Nz": str(
                    pz_slice_shape
                ),

                "npz_path": str(
                    npz_path
                ),

                "pz_array_key": "Pz_stack",

                "phi_array_key": (
                    "Phi_stack"
                    if has_phi
                    else ""
                ),

                "x_nm_key": (
                    "x_nm"
                    if "x_nm" in data.files
                    else ""
                ),

                # z_nm 永遠代表 Pz 的 FE coordinates
                "z_nm_key": (
                    "z_nm"
                    if "z_nm" in data.files
                    else ""
                ),
            })

    warning = None

    if (
        pv_df is not None
        and len(pv_df) != n_stacks
    ):
        warning = (
            f"{run_id}: existing NPZ stacks "
            f"({n_stacks}) != "
            f"PV rows ({len(pv_df)})"
        )

    print(
        f"[REUSE NPZ] {run_id}: "
        f"{n_stacks} stacks from "
        f"{npz_path}"
    )

    return (
        index_rows,
        str(npz_path),
        warning,
    )

def extract_pz_stacks(
    folder: Path,
    run_id: str,
    params: dict,
    pv_df: pd.DataFrame | None,
    pz_dir: Path,
    y_index=None,
    skip_initial=True,
    skip_first_n=9,
):
    # --------------------------------------------------------
    # Output NPZ path
    # --------------------------------------------------------
    out_folder = (
        pz_dir
        / run_id
    )

    out_folder.mkdir(
        parents=True,
        exist_ok=True,
    )

    out_npz = (
        out_folder
        / "Pz_Phi_FE_all_voltage.npz"
    )

    # --------------------------------------------------------
    # Find plotfiles
    # --------------------------------------------------------
    plot_dir = (
        folder
        / "plts"
    )

    if not plot_dir.exists():
        plot_dir = folder

    plot_names = find_plot_names(
        plot_dir
    )

    if skip_initial:
        plot_names = [
            name
            for name in plot_names
            if get_step_from_name(name) != 0
        ]

    if skip_first_n > 0:
        plot_names = (
            plot_names[
                skip_first_n:
            ]
        )

    # --------------------------------------------------------
    # No plotfiles:
    # reuse existing NPZ if available
    # --------------------------------------------------------
    if not plot_names:

        if out_npz.exists():
            return load_existing_pz_npz(
                npz_path=out_npz,
                folder=folder,
                run_id=run_id,
                pv_df=pv_df,
            )

        return (
            [],
            None,
            (
                f"{run_id}: no steady plotfiles "
                "found after filtering and "
                "no existing NPZ found"
            ),
        )

    # --------------------------------------------------------
    # Only import yt when plotfiles actually need reading
    # --------------------------------------------------------
    try:
        import yt

    except Exception as e:
        raise RuntimeError(
            "yt is required for field extraction. "
            "Install with: pip install yt"
        ) from e

    # --------------------------------------------------------
    # Storage
    # --------------------------------------------------------
    index_rows = []

    pz_slices = []
    phi_slices = []
    charge_slices = []

    # Static fields: only one copy
    epsilon_static = None
    mask_static = None

    plot_steps = []

    # References
    x_nm_ref = None
    z_nm_ref = None
    z_nm_full_ref = None

    pz_shape_ref = None
    full_slice_shape_ref = None

    full_shape_ref = None
    y_index_ref = None
    y_nm_ref = None

    # --------------------------------------------------------
    # Read plotfiles
    # --------------------------------------------------------
    for i, name in enumerate(
        plot_names
    ):
        plotfile = (
            plot_dir
            / name
        )

        info = read_one_plot(
            plotfile,
            params=params,
            yt_module=yt,
            y_index=y_index,

            # epsilon/mask only first bias
            read_static_fields=(
                i == 0
            ),
        )

        # ----------------------------------------------------
        # Dynamic fields
        # ----------------------------------------------------
        P_slice = np.asarray(
            info["P_slice"],
            dtype=np.float32,
        )

        Phi_slice = np.asarray(
            info["Phi_slice"],
            dtype=np.float32,
        )

        charge_slice = np.asarray(
            info["charge_slice"],
            dtype=np.float32,
        )

        if (
            Phi_slice.shape
            != charge_slice.shape
        ):
            raise RuntimeError(
                f"{plotfile}: "
                f"Phi shape {Phi_slice.shape} "
                f"!= charge shape "
                f"{charge_slice.shape}"
            )

        # ----------------------------------------------------
        # First plotfile defines references
        # ----------------------------------------------------
        if x_nm_ref is None:

            x_nm_ref = np.asarray(
                info["x_nm"],
                dtype=np.float32,
            )

            # FE z coordinates
            z_nm_ref = np.asarray(
                info["z_nm"],
                dtype=np.float32,
            )

            # Full-device z coordinates
            z_nm_full_ref = np.asarray(
                info["z_nm_full"],
                dtype=np.float32,
            )

            pz_shape_ref = (
                P_slice.shape
            )

            full_slice_shape_ref = (
                Phi_slice.shape
            )

            full_shape_ref = (
                info["full_shape"]
            )

            y_index_ref = int(
                info["y_index"]
            )

            y_nm_ref = float(
                info["y_nm"]
            )

            epsilon_static = np.asarray(
                info["epsilon_slice"],
                dtype=np.float32,
            )

            mask_static = np.asarray(
                info["mask_slice"],
                dtype=np.float32,
            )

        else:
            # ------------------------------------------------
            # Consistency checks
            # ------------------------------------------------
            if (
                P_slice.shape
                != pz_shape_ref
            ):
                raise RuntimeError(
                    f"{plotfile}: "
                    f"Pz FE slice shape "
                    f"{P_slice.shape} "
                    f"does not match reference "
                    f"{pz_shape_ref}"
                )

            if (
                Phi_slice.shape
                != full_slice_shape_ref
            ):
                raise RuntimeError(
                    f"{plotfile}: "
                    f"full-device slice shape "
                    f"{Phi_slice.shape} "
                    f"does not match reference "
                    f"{full_slice_shape_ref}"
                )

            if not np.allclose(
                info["x_nm"],
                x_nm_ref,
                rtol=0.0,
                atol=1e-10,
            ):
                raise RuntimeError(
                    f"{plotfile}: "
                    "x coordinates do not match "
                    "the first selected plotfile"
                )

            if not np.allclose(
                info["z_nm"],
                z_nm_ref,
                rtol=0.0,
                atol=1e-10,
            ):
                raise RuntimeError(
                    f"{plotfile}: "
                    "FE z coordinates do not match "
                    "the first selected plotfile"
                )

            if not np.allclose(
                info["z_nm_full"],
                z_nm_full_ref,
                rtol=0.0,
                atol=1e-10,
            ):
                raise RuntimeError(
                    f"{plotfile}: "
                    "full z coordinates do not match "
                    "the first selected plotfile"
                )

        # ----------------------------------------------------
        # Append fields
        # ----------------------------------------------------
        pz_slices.append(
            P_slice
        )

        phi_slices.append(
            Phi_slice
        )

        charge_slices.append(
            charge_slice
        )

        plot_step = get_step_from_name(
            name
        )

        plot_steps.append(
            plot_step
        )

        # ----------------------------------------------------
        # P-V metadata
        # ----------------------------------------------------
        if (
            pv_df is not None
            and i < len(pv_df)
        ):
            pv_row = pv_df.iloc[i]

            V_applied = pv_row.get(
                "V_applied",
                np.nan,
            )

            P_mean_csv = pv_row.get(
                "P_mean",
                np.nan,
            )

            branch = pv_row.get(
                "branch",
                "",
            )

            point_id = int(
                pv_row.get(
                    "point_id",
                    i,
                )
            )

        else:
            V_applied = np.nan
            P_mean_csv = np.nan
            branch = "unknown"
            point_id = i

        # ----------------------------------------------------
        # Excel index row
        # 保持原本 schema 不變
        # ----------------------------------------------------
        index_rows.append({
            "folder": folder.name,
            "point_id": point_id,

            "branch": branch,
            "V_applied": V_applied,
            "P_mean": P_mean_csv,

            "phi_available": True,

            "y_index": info[
                "y_index"
            ],

            "y_nm": info[
                "y_nm"
            ],

            "full_shape": str(
                info["full_shape"]
            ),

            # 這裡仍代表 Pz FE shape
            "slice_shape_Nx_Nz": str(
                info["slice_shape"]
            ),

            "npz_path": str(
                out_npz
            ),

            "pz_array_key": (
                "Pz_stack"
            ),

            "phi_array_key": (
                "Phi_stack"
            ),

            "x_nm_key": (
                "x_nm"
            ),

            # FE z coordinate
            "z_nm_key": (
                "z_nm"
            ),
        })

    # --------------------------------------------------------
    # Stack all bias points
    # --------------------------------------------------------

    # (N_voltage, Nx, Nz_FE)
    Pz_stack = np.stack(
        pz_slices,
        axis=0,
    ).astype(
        np.float32
    )

    # (N_voltage, Nx, Nz_full)
    Phi_stack = np.stack(
        phi_slices,
        axis=0,
    ).astype(
        np.float32
    )

    # (N_voltage, Nx, Nz_full)
    charge_stack = np.stack(
        charge_slices,
        axis=0,
    ).astype(
        np.float32
    )

    # --------------------------------------------------------
    # Metadata arrays
    # --------------------------------------------------------
    V_applied_array = np.asarray(
        [
            row["V_applied"]
            for row in index_rows
        ],
        dtype=np.float32,
    )

    P_mean_array = np.asarray(
        [
            row["P_mean"]
            for row in index_rows
        ],
        dtype=np.float32,
    )

    branch_array = np.asarray(
        [
            row["branch"]
            for row in index_rows
        ],
        dtype=str,
    )

    plot_step_array = np.asarray(
        [
            (
                step
                if step is not None
                else -1
            )
            for step in plot_steps
        ],
        dtype=np.int64,
    )

    # --------------------------------------------------------
    # Save exactly the same NPZ structure
    # as the updated notebook version
    # --------------------------------------------------------
    np.savez_compressed(
        out_npz,

        # FE only
        Pz_stack=Pz_stack,

        # full device
        Phi_stack=Phi_stack,
        charge_stack=charge_stack,

        # static, one copy only
        epsilon_static=epsilon_static,
        mask_static=mask_static,

        # coordinates
        x_nm=x_nm_ref,

        # FE z
        z_nm=z_nm_ref,

        # full-device z
        z_nm_full=z_nm_full_ref,

        # sweep metadata
        V_applied=V_applied_array,
        P_mean=P_mean_array,
        branch=branch_array,
        plot_step=plot_step_array,

        y_index=np.asarray(
            y_index_ref,
            dtype=np.int64,
        ),

        y_nm=np.asarray(
            y_nm_ref,
            dtype=np.float32,
        ),
    )

    # --------------------------------------------------------
    # Report NPZ size
    # --------------------------------------------------------
    size_bytes = (
        out_npz.stat().st_size
    )

    size_kib = (
        size_bytes
        / 1024
    )

    print(
        f"Saved NPZ: {out_npz}"
    )

    print(
        "Pz_stack shape       :",
        Pz_stack.shape,
    )

    print(
        "Phi_stack shape      :",
        Phi_stack.shape,
    )

    print(
        "charge_stack shape   :",
        charge_stack.shape,
    )

    print(
        "epsilon_static shape :",
        epsilon_static.shape,
    )

    print(
        "mask_static shape    :",
        mask_static.shape,
    )

    print(
        f"NPZ size             : "
        f"{size_kib:.2f} KiB"
    )

    # --------------------------------------------------------
    # Warning
    # --------------------------------------------------------
    warning = None

    if (
        pv_df is not None
        and len(pv_df) != len(plot_names)
    ):
        warning = (
            f"{run_id}: "
            f"PV rows ({len(pv_df)}) "
            f"!= selected plotfiles "
            f"({len(plot_names)})"
        )

    return (
        index_rows,
        str(out_npz),
        warning,
    )

# ==========================
# Excel writing
# ==========================

def format_excel(out_xlsx: Path):
    """Light formatting using openpyxl after pandas writes the workbook."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception:
        return

    wb = load_workbook(out_xlsx)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for ws in wb.worksheets:
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        # reasonable width cap
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            max_len = 0
            for cell in col_cells[:200]:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 42)

        # number formats
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.000E+00"
                elif isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"

    wb.save(out_xlsx)


def write_workbook(
    out_xlsx: Path,
    summary_rows: list[dict],
    exp_rows: list[dict],
    pv_rows_all: list[pd.DataFrame],
    pz_index_rows: list[dict],
    warnings: list[str],
):
    summary_df = pd.DataFrame(summary_rows)
    exp_df = pd.DataFrame(exp_rows)
    pv_df = pd.concat(pv_rows_all, ignore_index=True) if pv_rows_all else pd.DataFrame()
    pz_df = pd.DataFrame(pz_index_rows)
    warn_df = pd.DataFrame({"warning": warnings}) if warnings else pd.DataFrame({"warning": []})

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        exp_df.to_excel(writer, sheet_name="experiments", index=False)
        pv_df.to_excel(writer, sheet_name="pv_curve", index=False)
        pz_df.to_excel(writer, sheet_name="pz_stack_index", index=False)
        warn_df.to_excel(writer, sheet_name="warnings", index=False)

    format_excel(out_xlsx)


# ==========================
# Main build
# ==========================

def build_dataset(args):
    root = Path(args.root).resolve()
    out_xlsx = Path(args.out).resolve()
    pz_dir = Path(args.pz_dir).resolve()
    pz_dir.mkdir(parents=True, exist_ok=True)

    sweep_values, sweep_branches = build_default_sweep(args.vmin, args.vmax, args.dv)

    exp_rows = []
    pv_rows_all = []
    pz_index_rows = []
    warnings = []

    folders = sorted([p for p in root.glob("MFIS*") if p.is_dir()])

    for folder in folders:
        run_id = folder.name
        inp = folder / "inputs"
        if not inp.exists():
            warnings.append(f"{run_id}: inputs not found; skipped")
            continue

        print(f"Processing {run_id}")
        params = read_inputs(inp)

        fe_lo_z = get_component(params.get("FE_lo"), 2)
        fe_hi_z = get_component(params.get("FE_hi"), 2)
        t_fe = fe_hi_z - fe_lo_z if fe_lo_z is not None and fe_hi_z is not None else np.nan

        start_time, end_time, elapsed_hms, elapsed_sec = read_run_log(folder)

        pv_df, pv_csv_path, pv_warning = read_pv_curve(folder, run_id, sweep_values, sweep_branches)
        if pv_warning:
            warnings.append(pv_warning)
        if pv_df is not None:
            pv_rows_all.append(pv_df)

        pz_npz_path = None
        n_pz_stacks = 0
        if args.extract_pz:
            try:
                rows, pz_npz_path, pz_warning = extract_pz_stacks(
                    folder=folder,
                    run_id=run_id,
                    params=params,
                    pv_df=pv_df,
                    pz_dir=pz_dir,
                    y_index=args.y_index,
                    skip_initial=not args.keep_initial,
                    skip_first_n=args.skip_first_n,
                )
                pz_index_rows.extend(rows)
                n_pz_stacks = len(rows)
                if pz_warning:
                    warnings.append(pz_warning)
            except Exception as e:
                warnings.append(f"{run_id}: Pz extraction failed: {e}")

        exp_row = {
            "run_id": run_id,
            "folder": folder.name,
            "inputs_path": str(inp),
            "pv_csv_path": pv_csv_path,
            "pz_npz_path": pz_npz_path,
            "n_pv_points": 0 if pv_df is None else len(pv_df),
            "n_pz_stacks": n_pz_stacks,
            "T_FE": t_fe,
            "T_FE_nm": t_fe * 1e9 if np.isfinite(t_fe) else np.nan,
            "Start Time": start_time,
            "End Time": end_time,
            "Elapsed": elapsed_hms,
            "Elapsed_sec": elapsed_sec,
        }

        for p in WANTED_PARAMS:
            if p in ["FE_lo", "FE_hi"]:
                v = params.get(p)
                exp_row[p] = " ".join(sci_or_blank(x) for x in v) if isinstance(v, list) else sci_or_blank(v)
            else:
                exp_row[p] = params.get(p, np.nan)

        exp_rows.append(exp_row)

    summary_rows = [
        {"item": "Last Update", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"item": "Root", "value": str(root)},
        {"item": "Total MFIS folders", "value": len(folders)},
        {"item": "Experiments written", "value": len(exp_rows)},
        {"item": "P-V points written", "value": sum(len(df) for df in pv_rows_all)},
        {"item": "Pz stacks indexed", "value": len(pz_index_rows)},
        {"item": "Default sweep", "value": f"{args.vmin} -> {args.vmax} -> {args.vmin}, dV={args.dv}"},
        {"item": "Pz storage", "value": "2D Pz arrays are saved in compressed .npz files; Excel stores npz_path + array_key."},
        {"item": "P sign convention", "value": f"Pz multiplied by {P_SIGN}, same as plotting code."},
    ]

    write_workbook(
        out_xlsx=out_xlsx,
        summary_rows=summary_rows,
        exp_rows=exp_rows,
        pv_rows_all=pv_rows_all,
        pz_index_rows=pz_index_rows,
        warnings=warnings,
    )

    print("Done")
    print(f"Excel saved: {out_xlsx}")
    print(f"Pz npz folder: {pz_dir}")
    if warnings:
        print(f"Warnings: {len(warnings)}; see warnings sheet")


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="Build FerroX MFIS Excel dataset index")
    parser.add_argument("--root", default=".", help="Root folder containing MFIS* simulation folders")
    parser.add_argument("--out", default="MFIS_dataset.xlsx", help="Output Excel filename")
    parser.add_argument("--pz-dir", default="extracted_pz", help="Folder for compressed Pz .npz outputs")
    parser.add_argument("--no-pz", dest="extract_pz", action="store_false", help="Do not read plotfiles / do not extract Pz stacks")
    parser.set_defaults(extract_pz=True)
    parser.set_defaults(extract_phi=True)
    parser.add_argument("--vmin", type=float, default=DEFAULT_VMIN)
    parser.add_argument("--vmax", type=float, default=DEFAULT_VMAX)
    parser.add_argument("--dv", type=float, default=DEFAULT_DV)
    parser.add_argument("--skip-first-n", type=int, default=SKIP_FIRST_N, help="Skip first N steady-state plotfiles after initial plt00000000 removal")
    parser.add_argument("--keep-initial", action="store_true", help="Do not remove plt00000000")
    parser.add_argument("--y-index", type=int, default=Y_INDEX, help="y index for Pz slice; default is center")
    return parser.parse_args(argv)


if __name__ == "__main__":
    build_dataset(parse_args())

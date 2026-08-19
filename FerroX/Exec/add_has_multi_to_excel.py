#!/usr/bin/env python3
"""Calculate ``has_multi`` for any FerroX structure and update Excel.

Use ``python add_has_multi_to_excel.py --help`` for the complete workflow and
copyable examples.

By default ``--structure MFIS`` selects ``MFIS_dataset.xlsx`` and cases matching
``MFIS*`` below ``extracted_pz``.  MFIM, MFM, and custom structures can use the
same script.  Workbook/sheet/columns, NPZ directory/name/key, case extraction,
and z selection are configurable.

For each selected NPZ, inspect ``Pz_stack[voltage_index, :, z_index]``. A case
has multiple domains when at least one voltage slice satisfies both conditions:

    abs(P_max - P_min) > var_threshold
    P_max > 0 and P_min < 0

Every nonblank experiment row is updated. If its case has no corresponding
NPZ, ``has_multi`` is set to ``no_data`` and processing continues.

With ``--dry-run``, the real workbook is not changed.  Use ``--no-preview`` for
a dependency-free dry run, or leave preview enabled when ``build_dash.py`` is
available.
"""

from __future__ import annotations

import argparse
import fnmatch
import math
import os
import tempfile
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_STRUCTURE = "MFIS"
DEFAULT_EXTRACTED_PZ_DIR = "extracted_pz"
DEFAULT_SHEET_NAME = "experiments"
DEFAULT_HAS_MULTI_COLUMN = "has_multi"
DEFAULT_NPZ_NAME = "Pz_Phi_FE_all_voltage.npz"
DEFAULT_PREVIEW_HTML = "index.html"
DEFAULT_MISSING_LABEL = "no_data"
DEFAULT_NPZ_KEY = "Pz_stack"
DEFAULT_Z_INDEX = 5

# Checked in this order when --case-column is omitted.
CASE_COLUMN_CANDIDATES = (
    "file_name",
    "folder",
    "folder_name",
    "case_key",
    "case",
)

HELP_EPILOG = r"""
判別方式:
  逐一檢查 Pz_stack[voltage_index, :, z_index]。只要任一切片同時滿足
  abs(P_max - P_min) > var_threshold、P_max > 0、P_min < 0，就設為 1。

使用流程:
  1. 先用 --dry-run 產生 HTML，確認 has_multi 與對應圖片。
  2. 視需要調整 --var-threshold 或 --z-index，再次預覽。
  3. 確認後移除 --dry-run，才會寫入指定 Excel。

結果值:
  1       至少一個 voltage slice 通過判別條件
  0       有 NPZ 資料，但沒有 slice 通過條件
  no_data 該 Excel case 找不到對應 NPZ（可用 --missing-label 修改）

Excel 欄位:
  新欄預設優先繼承 gamma 欄格式；沒有 gamma 時改用 case-name 欄。
  也可用 --style-source-column 明確指定。

範例 1：預覽 MFIM，不改 Excel、不產生 HTML
  python add_has_multi_to_excel.py \
      --structure MFIM \
      --var-threshold 0.1 \
      --dry-run --no-preview

範例 2：正式寫入 MFIM_dataset.xlsx
  python add_has_multi_to_excel.py \
      --structure MFIM \
      --var-threshold 0.1

範例 3：自訂 workbook、NPZ 檔名/key，並掃描所有 z
  python add_has_multi_to_excel.py \
      --var-threshold 0.1 \
      --excel custom.xlsx \
      --sheet runs --case-column case_id \
      --case-glob 'run_*' \
      --extracted-pz-dir arrays \
      --npz-name fields.npz \
      --npz-key polarization_stack \
      --z-index all

在其他 Python 程式中重用核心判別:
  from add_has_multi_to_excel import MultiAnalyzer

  analyzer = MultiAnalyzer(var_threshold=0.1, z_index=None)
  result = analyzer.analyze_npz("Pz_Phi_FE_all_voltage.npz")
  print(result.has_multi)

提示:
  --dry_run 等同 --dry-run
  --excel-name、--excel-path 等同 --excel
  --preview-html 等同 --html
  --sheet-name 等同 --sheet
"""


def nonnegative_float(text: str) -> float:
    """argparse converter for a finite, non-negative threshold."""

    try:
        value = float(text)
    except ValueError as exc:
        raise argparse.ArgumentTypeError(f"not a number: {text!r}") from exc
    if not math.isfinite(value) or value < 0:
        raise argparse.ArgumentTypeError(
            f"threshold must be finite and >= 0, got {text!r}"
        )
    return value


def nonnegative_int(text: str) -> int:
    """argparse converter for a non-negative integer."""

    try:
        value = int(text)
    except ValueError as exc:
        raise argparse.ArgumentTypeError(f"not an integer: {text!r}") from exc
    if value < 0:
        raise argparse.ArgumentTypeError(f"must be >= 0, got {text!r}")
    return value


def positive_int(text: str) -> int:
    """argparse converter for a positive integer."""

    value = nonnegative_int(text)
    if value == 0:
        raise argparse.ArgumentTypeError(f"must be >= 1, got {text!r}")
    return value


def z_index_value(text: str) -> int | None:
    """Accept a non-negative index or ``all``/``none`` for every z slice."""

    if text.strip().casefold() in {"all", "none"}:
        return None
    return nonnegative_int(text)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Set a multi-domain flag using a configurable NPZ polarization stack."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=HELP_EPILOG,
    )
    parser.add_argument(
        "--structure",
        default=DEFAULT_STRUCTURE,
        help=(
            "Structure label used by defaults: <structure>_dataset.xlsx and "
            "case glob <structure>* (default: MFIS)."
        ),
    )
    parser.add_argument(
        "--var-threshold",
        type=nonnegative_float,
        required=True,
        help="Required lower bound for abs(P_max - P_min); comparison is strict (>).",
    )
    parser.add_argument(
        "--root",
        type=Path,
        default=Path.cwd(),
        help="Execution directory containing Excel/data paths (default: cwd).",
    )
    parser.add_argument(
        "--excel",
        "--excel-name",
        "--excel-path",
        dest="excel",
        type=Path,
        default=None,
        help="Workbook path; default: <structure>_dataset.xlsx under --root.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Optional output workbook. By default, update --excel in place.",
    )
    parser.add_argument(
        "--case-glob",
        default=None,
        help=(
            "Select case names in both NPZ data and Excel; default: "
            "<structure>*. Use '*' for every case."
        ),
    )
    parser.add_argument(
        "--extracted-pz-dir",
        type=Path,
        default=None,
        help=(
            "Directory searched recursively for NPZ files, relative to "
            "--root unless absolute "
            "(default: extracted_pz/<structure>)."
        ),
    )
    parser.add_argument(
        "--sheet",
        "--sheet-name",
        dest="sheet",
        default=DEFAULT_SHEET_NAME,
        help=f"Worksheet name (default: {DEFAULT_SHEET_NAME}).",
    )
    parser.add_argument(
        "--header-row",
        type=positive_int,
        default=1,
        help="One-based worksheet header row (default: 1).",
    )
    parser.add_argument(
        "--case-column",
        default=None,
        help=(
            "Column containing case names. If omitted, detect file_name, folder, "
            "folder_name, case_key, or case automatically."
        ),
    )
    parser.add_argument(
        "--has-multi-column",
        default=DEFAULT_HAS_MULTI_COLUMN,
        help=f"Column to create/update (default: {DEFAULT_HAS_MULTI_COLUMN}).",
    )
    parser.add_argument(
        "--style-source-column",
        default=None,
        help=(
            "Column whose style is copied to the output column. Default: "
            "gamma when present, otherwise the case-name column."
        ),
    )
    parser.add_argument(
        "--npz-name",
        default=DEFAULT_NPZ_NAME,
        help=f"NPZ basename to discover recursively (default: {DEFAULT_NPZ_NAME}).",
    )
    parser.add_argument(
        "--npz-key",
        default=DEFAULT_NPZ_KEY,
        help=f"Array key inside each NPZ (default: {DEFAULT_NPZ_KEY}).",
    )
    parser.add_argument(
        "--case-parent-level",
        type=positive_int,
        default=1,
        help=(
            "Which NPZ ancestor supplies the case name: 1=parent, 2=grandparent, "
            "... (default: 1)."
        ),
    )
    parser.add_argument(
        "--z-index",
        type=z_index_value,
        default=DEFAULT_Z_INDEX,
        help=(
            f"Third-dimension index (default: {DEFAULT_Z_INDEX}); use 'all' "
            "to inspect every z slice."
        ),
    )
    parser.add_argument(
        "--missing-label",
        default=DEFAULT_MISSING_LABEL,
        help=f"Value for Excel cases without NPZ data (default: {DEFAULT_MISSING_LABEL}).",
    )
    parser.add_argument(
        "--no-preview",
        action="store_true",
        help="With --dry-run, skip build_dash.py and HTML generation.",
    )
    parser.add_argument(
        "--dry-run",
        "--dry_run",
        dest="dry_run",
        action="store_true",
        help=(
            "Write proposed values to a temporary workbook and generate a "
            "preview dashboard without changing the real workbook."
        ),
    )
    parser.add_argument(
        "--preview-html",
        "--html",
        dest="preview_html",
        type=Path,
        default=Path(DEFAULT_PREVIEW_HTML),
        help=(
            "HTML generated during --dry-run, relative to --root unless "
            f"absolute (default: {DEFAULT_PREVIEW_HTML})."
        ),
    )
    args = parser.parse_args()
    if not args.structure.strip():
        parser.error("--structure cannot be empty.")
    if not args.npz_name.strip():
        parser.error("--npz-name cannot be empty.")
    if not args.npz_key.strip():
        parser.error("--npz-key cannot be empty.")
    if not args.has_multi_column.strip():
        parser.error("--has-multi-column cannot be empty.")
    return args


def resolve_under_root(path: Path, root: Path) -> Path:
    return path if path.is_absolute() else root / path


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return "".join(str(value).strip().casefold().split())


def _find_header_column(
    worksheet: Any,
    header_row: int,
    aliases: set[str],
) -> int:
    normalized_aliases = {_normalize_header(alias) for alias in aliases}
    matches = [
        column
        for column in range(1, worksheet.max_column + 1)
        if _normalize_header(worksheet.cell(header_row, column).value)
        in normalized_aliases
    ]
    if not matches:
        raise KeyError(
            f"Could not find any of {sorted(aliases)!r} in header row {header_row}"
        )
    if len(matches) > 1:
        raise KeyError(f"Multiple matching columns for {sorted(aliases)!r}: {matches}")
    return matches[0]


def normalized_case_name(value: object) -> str:
    return "" if value is None else str(value).strip()


@dataclass(frozen=True, slots=True)
class MultiResult:
    has_multi: int
    voltage_index: int | None = None
    z_index: int | None = None
    p_min: float | None = None
    p_max: float | None = None

    @property
    def variation(self) -> float | None:
        if self.p_min is None or self.p_max is None:
            return None
        return abs(self.p_max - self.p_min)


@dataclass(frozen=True, slots=True)
class MultiAnalyzer:
    var_threshold: float
    z_index: int | None = DEFAULT_Z_INDEX
    npz_key: str = DEFAULT_NPZ_KEY

    def analyze_npz(self, npz_path: str | Path) -> MultiResult:
        path = Path(npz_path)
        try:
            with np.load(path, allow_pickle=False) as data:
                if self.npz_key not in data.files:
                    raise RuntimeError(
                        f"array key {self.npz_key!r} not found in {path}; "
                        f"available: {data.files}"
                    )
                stack = np.asarray(data[self.npz_key])
        except (OSError, ValueError) as exc:
            raise RuntimeError(f"failed to read {path}: {exc}") from exc

        if stack.ndim != 3:
            raise RuntimeError(
                f"{path}: expected a 3D array [voltage, x, z], got {stack.shape}"
            )
        if stack.shape[0] == 0 or stack.shape[1] == 0 or stack.shape[2] == 0:
            raise RuntimeError(f"{path}: polarization stack has an empty axis")

        if self.z_index is None:
            z_indices = range(stack.shape[2])
        else:
            if not 0 <= self.z_index < stack.shape[2]:
                raise RuntimeError(
                    f"{path}: z_index={self.z_index}, valid range is "
                    f"0..{stack.shape[2] - 1}"
                )
            z_indices = (self.z_index,)

        for voltage_index in range(stack.shape[0]):
            for z_index in z_indices:
                line = np.asarray(stack[voltage_index, :, z_index], dtype=float)
                if not np.all(np.isfinite(line)):
                    raise RuntimeError(
                        f"{path}: non-finite value at voltage_index="
                        f"{voltage_index}, z_index={z_index}"
                    )
                p_min = float(np.min(line))
                p_max = float(np.max(line))
                if (
                    abs(p_max - p_min) > self.var_threshold
                    and p_max > 0
                    and p_min < 0
                ):
                    return MultiResult(
                        has_multi=1,
                        voltage_index=voltage_index,
                        z_index=z_index,
                        p_min=p_min,
                        p_max=p_max,
                    )
        return MultiResult(has_multi=0)


def discover_npz_files(
    extracted_pz_dir: Path,
    npz_name: str,
    *,
    case_glob: str,
    case_parent_level: int,
) -> dict[str, Path]:
    """Map selected case names to unique NPZ paths."""

    if not extracted_pz_dir.is_dir():
        raise FileNotFoundError(
            f"Extracted-Pz directory does not exist: {extracted_pz_dir}"
        )

    result: dict[str, Path] = {}
    for npz_path in sorted(extracted_pz_dir.rglob(npz_name)):
        try:
            case_name = npz_path.parents[case_parent_level - 1].name.strip()
        except IndexError as exc:
            raise ValueError(
                f"Cannot get ancestor level {case_parent_level} for {npz_path}"
            ) from exc
        if not case_name:
            raise ValueError(f"Cannot determine the case name for {npz_path}.")
        if not fnmatch.fnmatchcase(case_name, case_glob):
            continue
        if case_name in result:
            raise ValueError(
                f"Duplicate NPZ files found for case {case_name!r}:\n"
                f"  {result[case_name]}\n"
                f"  {npz_path}"
            )
        result[case_name] = npz_path
    return result


def find_case_column(
    worksheet: Worksheet,
    header_row: int,
    requested_name: str | None,
) -> tuple[int, str]:
    """Find the experiment case-name column."""

    header_to_columns: dict[str, list[int]] = {}
    for column_index in range(1, worksheet.max_column + 1):
        header = _normalize_header(
            worksheet.cell(row=header_row, column=column_index).value
        )
        if header:
            header_to_columns.setdefault(header, []).append(column_index)

    names_to_try = (
        (_normalize_header(requested_name),)
        if requested_name is not None
        else CASE_COLUMN_CANDIDATES
    )
    for name in names_to_try:
        matches = header_to_columns.get(name, [])
        if len(matches) == 1:
            original_name = worksheet.cell(
                row=header_row,
                column=matches[0],
            ).value
            return matches[0], str(original_name)
        if len(matches) > 1:
            raise ValueError(
                f"Worksheet contains more than one column named {name!r}."
            )

    available = [
        str(worksheet.cell(row=header_row, column=index).value)
        for index in range(1, worksheet.max_column + 1)
        if worksheet.cell(row=header_row, column=index).value is not None
    ]
    wanted = (
        repr(requested_name)
        if requested_name is not None
        else "one of " + ", ".join(repr(name) for name in CASE_COLUMN_CANDIDATES)
    )
    raise ValueError(
        f"Could not find case-name column {wanted} in row {header_row}. "
        f"Available columns: {available}"
    )


def find_style_source_column(
    worksheet: Worksheet,
    *,
    header_row: int,
    requested_name: str | None,
    fallback_column: int,
) -> int:
    if requested_name is not None:
        return _find_header_column(worksheet, header_row, {requested_name})
    try:
        return _find_header_column(worksheet, header_row, {"gamma", "γ"})
    except KeyError:
        return fallback_column


def _ensure_output_column(
    worksheet: Any,
    header_row: int,
    name: str,
    style_source_column: int,
) -> tuple[int, bool]:
    normalized_name = _normalize_header(name)
    for column in range(1, worksheet.max_column + 1):
        if _normalize_header(worksheet.cell(header_row, column).value) == normalized_name:
            return column, False

    column = worksheet.max_column + 1
    source = worksheet.cell(header_row, style_source_column)
    target = worksheet.cell(header_row, column, name)
    if source.has_style:
        target._style = copy(source._style)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    return column, True


def _copy_style(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def _extend_filter_and_tables(
    worksheet: Any,
    *,
    header_row: int,
    old_max_column: int,
    new_max_column: int,
) -> None:
    if new_max_column <= old_max_column:
        return

    from openpyxl.utils import get_column_letter, range_boundaries
    from openpyxl.worksheet.table import TableColumn

    def extend_ref(reference: str) -> str:
        min_col, min_row, max_col, max_row = range_boundaries(reference)
        if min_row == header_row and max_col == old_max_column:
            max_col = new_max_column
        return (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )

    if worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = extend_ref(worksheet.auto_filter.ref)

    for table in worksheet.tables.values():
        _, min_row, max_col, _ = range_boundaries(table.ref)
        if min_row != header_row or max_col != old_max_column:
            continue

        next_id = max((column.id for column in table.tableColumns), default=0) + 1
        for column in range(old_max_column + 1, new_max_column + 1):
            table.tableColumns.append(
                TableColumn(
                    id=next_id,
                    name=str(worksheet.cell(header_row, column).value),
                )
            )
            next_id += 1
        table.ref = extend_ref(table.ref)
        if table.autoFilter is not None and table.autoFilter.ref:
            table.autoFilter.ref = extend_ref(table.autoFilter.ref)


def _repair_existing_output_column(
    worksheet: Any,
    *,
    header_row: int,
    style_source_column: int,
    output_column: int,
    width: float,
) -> None:
    """Repair a column created by an older run but left outside the filter."""

    from openpyxl.utils import get_column_letter, range_boundaries
    from openpyxl.worksheet.table import TableColumn

    _copy_style(
        worksheet.cell(header_row, style_source_column),
        worksheet.cell(header_row, output_column),
    )
    worksheet.column_dimensions[get_column_letter(output_column)].width = width

    def extend_ref(reference: str) -> str:
        min_col, min_row, max_col, max_row = range_boundaries(reference)
        if min_row == header_row and max_col < output_column:
            max_col = output_column
        return (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )

    if worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = extend_ref(worksheet.auto_filter.ref)

    for table in worksheet.tables.values():
        _, min_row, max_col, _ = range_boundaries(table.ref)
        if min_row != header_row or max_col >= output_column:
            continue

        next_id = max((column.id for column in table.tableColumns), default=0) + 1
        for column in range(max_col + 1, output_column + 1):
            table.tableColumns.append(
                TableColumn(
                    id=next_id,
                    name=str(worksheet.cell(header_row, column).value),
                )
            )
            next_id += 1
        table.ref = extend_ref(table.ref)
        if table.autoFilter is not None and table.autoFilter.ref:
            table.autoFilter.ref = extend_ref(table.autoFilter.ref)


def save_workbook_atomically(workbook: object, output_path: Path) -> None:
    """Write beside the destination, then replace it after save succeeds."""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output_path.stem}.",
        suffix=output_path.suffix,
        dir=output_path.parent,
    )
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
    finally:
        temporary_path.unlink(missing_ok=True)


def format_result(case_name: str, result: MultiResult) -> str:
    if not result.has_multi:
        return f"{case_name}: has_multi=0"

    assert result.voltage_index is not None
    assert result.p_min is not None
    assert result.p_max is not None
    assert result.variation is not None
    return (
        f"{case_name}: has_multi={result.has_multi} | "
        f"voltage_index={result.voltage_index}, z_index={result.z_index}, "
        f"P_min={result.p_min:.8g}, P_max={result.p_max:.8g}, "
        f"variation={result.variation:.8g}" 
    )


def run(args: argparse.Namespace) -> int:
    root = args.root.expanduser().resolve()
    if not root.is_dir():
        raise NotADirectoryError(f"Execution directory does not exist: {root}")

    excel_argument = (
        args.excel
        if args.excel is not None
        else Path(f"{args.structure}_dataset.xlsx")
    )
    excel_path = resolve_under_root(excel_argument.expanduser(), root).resolve()
    output_path = (
        excel_path
        if args.output is None
        else resolve_under_root(args.output.expanduser(), root).resolve()
    )
    extracted_pz_argument = (
        args.extracted_pz_dir
        if args.extracted_pz_dir is not None
        else Path(DEFAULT_EXTRACTED_PZ_DIR) / args.structure
    )

    extracted_pz_dir = resolve_under_root(
        extracted_pz_argument.expanduser(),
        root,
    ).resolve()
    preview_html_path = resolve_under_root(
        args.preview_html.expanduser(), root
    ).resolve()
    if not excel_path.is_file():
        raise FileNotFoundError(f"Workbook does not exist: {excel_path}")
    if excel_path.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise ValueError("--excel must be an .xlsx or .xlsm workbook")
    if output_path.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise ValueError("--output must use the .xlsx or .xlsm extension")

    case_glob = args.case_glob or f"{args.structure}*"
    if not case_glob.strip():
        raise ValueError("--case-glob cannot be empty")
    npz_paths = discover_npz_files(
        extracted_pz_dir,
        args.npz_name,
        case_glob=case_glob,
        case_parent_level=args.case_parent_level,
    )
    print(f"Discovered NPZ files: {len(npz_paths)}")

    analyzer = MultiAnalyzer(
        var_threshold=args.var_threshold,
        z_index=args.z_index,
        npz_key=args.npz_key,
    )
    results: dict[str, MultiResult] = {}
    errors: list[str] = []
    for case_name, npz_path in npz_paths.items():
        try:
            result = analyzer.analyze_npz(npz_path)
        except RuntimeError as exc:
            errors.append(f"{case_name}: {exc}")
            continue
        results[case_name] = result
        print(format_result(case_name, result))

    if errors:
        details = "\n".join(f"  - {error}" for error in errors)
        raise RuntimeError(
            "No workbook changes were saved because one or more NPZ files "
            f"could not be evaluated:\n{details}"
        )

    keep_vba = excel_path.suffix.casefold() == ".xlsm"
    workbook = load_workbook(excel_path, keep_vba=keep_vba, keep_links=True)
    try:
        if args.sheet not in workbook.sheetnames:
            raise KeyError(
                f"Worksheet {args.sheet!r} was not found. "
                f"Available worksheets: {workbook.sheetnames}"
            )
        worksheet = workbook[args.sheet]
        case_column, detected_case_column = find_case_column(
            worksheet,
            args.header_row,
            args.case_column,
        )
        style_column = find_style_source_column(
            worksheet,
            header_row=args.header_row,
            requested_name=args.style_source_column,
            fallback_column=case_column,
        )

        old_max_column = worksheet.max_column
        has_multi_column, _ = _ensure_output_column(
            worksheet,
            args.header_row,
            args.has_multi_column,
            style_column,
        )

        _extend_filter_and_tables(
            worksheet,
            header_row=args.header_row,
            old_max_column=old_max_column,
            new_max_column=worksheet.max_column,
        )
        # Also repair a column left by an older run: it may exist but still be
        # outside the AutoFilter or have the default header style.
        _repair_existing_output_column(
            worksheet,
            header_row=args.header_row,
            style_source_column=style_column,
            output_column=has_multi_column,
            width=14.0,
        )

        matched_cases: set[str] = set()
        positive_rows = 0
        negative_rows = 0
        missing_rows = 0
        updated_rows = 0

        for row_number in range(args.header_row + 1, worksheet.max_row + 1):
            case_name = normalized_case_name(
                worksheet.cell(row=row_number, column=case_column).value
            )
            if not case_name:
                continue
            if not fnmatch.fnmatchcase(case_name, case_glob):
                continue

            result = results.get(case_name)
            if result is None:
                value: int | str = args.missing_label
                missing_rows += 1
            else:
                value = result.has_multi
                matched_cases.add(case_name)
                if result.has_multi:
                    positive_rows += 1
                else:
                    negative_rows += 1

            target_cell = worksheet.cell(
                row=row_number,
                column=has_multi_column,
            )
            _copy_style(
                worksheet.cell(row_number, style_column),
                target_cell,
            )
            target_cell.value = value
            target_cell.number_format = "General"
            updated_rows += 1

        if updated_rows == 0:
            raise RuntimeError(
                f"No nonblank case matching {case_glob!r} was found in "
                f"worksheet column {detected_case_column!r}; no changes were saved."
            )

        unmatched_npz_cases = sorted(set(results) - matched_cases)
        for case_name in unmatched_npz_cases:
            print(
                f"[WARNING] {case_name}: NPZ was analyzed but no matching "
                f"Excel row was found in {detected_case_column!r}; skipped."
            )

        print(f"Case-name column: {detected_case_column}")
        if args.dry_run:
            if not args.no_preview:
                descriptor, temporary_name = tempfile.mkstemp(
                    prefix=f".{excel_path.stem}.has_multi_preview.",
                    suffix=excel_path.suffix,
                    dir=root,
                )
                os.close(descriptor)
                temporary_excel_path = Path(temporary_name)
                try:
                    try:
                        import build_dash
                    except ImportError as exc:
                        raise RuntimeError(
                            "HTML preview requires build_dash.py; rerun with "
                            "--no-preview for a dependency-free dry run"
                        ) from exc
                    workbook.save(temporary_excel_path)
                    build_dash.build_dashboard(
                        temporary_excel_path,
                        data_root=root,
                        output_path=preview_html_path,
                        sheet_name=args.sheet,
                    )
                finally:
                    temporary_excel_path.unlink(missing_ok=True)
                print(f"Preview dashboard: {preview_html_path}")

            print(
                f"[DRY RUN] Would update {updated_rows} worksheet row(s): "
                f"{positive_rows} has_multi=1, {negative_rows} has_multi=0, "
                f"{missing_rows} {args.missing_label!r}."
            )
        else:
            save_workbook_atomically(workbook, output_path)
            print(
                f"Saved {output_path} | updated {updated_rows} worksheet row(s): "
                f"{positive_rows} has_multi=1, {negative_rows} has_multi=0, "
                f"{missing_rows} {args.missing_label!r}."
            )
    finally:
        workbook.close()

    return 0


def main() -> int:
    args = parse_args()
    try:
        return run(args)
    except (
        FileNotFoundError,
        NotADirectoryError,
        KeyError,
        ValueError,
        RuntimeError,
    ) as exc:
        raise SystemExit(f"ERROR: {exc}") from exc


if __name__ == "__main__":
    raise SystemExit(main())

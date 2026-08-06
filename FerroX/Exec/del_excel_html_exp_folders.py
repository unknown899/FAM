#!/usr/bin/env python3
"""Delete selected experiment folders from an Excel workbook and an HTML table."""

from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass
import html
import os
from pathlib import Path
import re
import tempfile
from typing import Pattern, Sequence

from openpyxl import load_workbook


DEFAULT_EXCEL_PATH = Path("MFIS_dataset.xlsx")
DEFAULT_HTML_PATH = Path("index.html")
DEFAULT_SHEET_NAMES = (
    "experiments",
    "pv_curve",
    "pz_stack_index",
)
DEFAULT_HEADER_ROW = 1


@dataclass(frozen=True)
class FolderMatcher:
    """All folder-selection methods are combined with logical OR."""

    exact_folders: frozenset[str]
    prefixes: tuple[str, ...]
    regexes: tuple[Pattern[str], ...]

    def matches(self, folder: str) -> bool:
        return (
            folder in self.exact_folders
            or any(folder.startswith(prefix) for prefix in self.prefixes)
            or any(pattern.search(folder) for pattern in self.regexes)
        )


def build_parser() -> argparse.ArgumentParser:
    examples = r"""
指定要刪除的 folder（可混合使用，條件之間為 OR）：
  # 1. 位置參數：一次指定一個或多個完整 folder 名稱
  python del_excel_html_exp_folders.py MFIS_t_6_nomi_1_var_5_8
  python del_excel_html_exp_folders.py case_1 case_2 case_3

  # 2. --folder / -f：選項可重複
  python del_excel_html_exp_folders.py -f case_1 -f case_2

  # 3. 從文字檔讀取：每行一個 folder；空行及 # 註解會忽略
  python del_excel_html_exp_folders.py --folder-file folders_to_delete.txt

  # 4. 依 prefix 或 regular expression 批次指定
  python del_excel_html_exp_folders.py --prefix MFIS_t_7_nomi_16_gvar
  python del_excel_html_exp_folders.py --regex '^MFIS_t_7_nomi_(8|9|10)$'

指定 Excel / HTML 檔名：
  # 未指定時，預設就地修改 MFIS_dataset.xlsx 與 index.html
  python del_excel_html_exp_folders.py case_1

  # 指定輸入檔名，仍然就地修改
  python del_excel_html_exp_folders.py case_1 \
      --excel temp_dataset.xlsx --html temp_index.html

  # 保留原檔，將結果另存成新檔
  python del_excel_html_exp_folders.py case_1 \
      --excel MFIS_dataset.xlsx --excel-output cleaned_dataset.xlsx \
      --html index.html --html-output cleaned_index.html

其他：
  # 只預覽，不寫入檔案
  python del_excel_html_exp_folders.py case_1 --dry-run
"""

    parser = argparse.ArgumentParser(
        description=(
            "同時刪除 Excel 指定工作表與 HTML 表格中，"
            "folder 符合條件的整筆資料。任一檔案、工作表或資料不存在時會跳過。"
        ),
        epilog=examples,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    parser.add_argument(
        "folders",
        nargs="*",
        metavar="FOLDER",
        help="要刪除的完整 folder 名稱；可一次給多個",
    )
    parser.add_argument(
        "-f",
        "--folder",
        action="append",
        default=[],
        metavar="FOLDER",
        help="要刪除的完整 folder 名稱；可重複指定",
    )
    parser.add_argument(
        "--folder-file",
        action="append",
        default=[],
        type=Path,
        metavar="PATH",
        help="folder 清單文字檔；每行一個名稱，可重複指定此選項",
    )
    parser.add_argument(
        "--prefix",
        action="append",
        default=[],
        metavar="TEXT",
        help="刪除名稱以此字串開頭的 folder；可重複指定",
    )
    parser.add_argument(
        "--regex",
        action="append",
        default=[],
        metavar="PATTERN",
        help="刪除名稱符合此 regular expression 的 folder；可重複指定",
    )

    parser.add_argument(
        "--excel",
        type=Path,
        default=DEFAULT_EXCEL_PATH,
        metavar="PATH",
        help=f"Excel 輸入檔；預設：{DEFAULT_EXCEL_PATH}",
    )
    parser.add_argument(
        "--excel-output",
        type=Path,
        metavar="PATH",
        help="Excel 輸出檔；未指定時覆寫 --excel",
    )
    parser.add_argument(
        "--html",
        type=Path,
        default=DEFAULT_HTML_PATH,
        metavar="PATH",
        help=f"HTML 輸入檔；預設：{DEFAULT_HTML_PATH}",
    )
    parser.add_argument(
        "--html-output",
        type=Path,
        metavar="PATH",
        help="HTML 輸出檔；未指定時覆寫 --html",
    )
    parser.add_argument(
        "--sheet",
        action="append",
        dest="sheet_names",
        metavar="NAME",
        help=(
            "要處理的 Excel 工作表；可重複指定。未指定時處理："
            + ", ".join(DEFAULT_SHEET_NAMES)
        ),
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=DEFAULT_HEADER_ROW,
        metavar="N",
        help=f"Excel 標題列（從 1 起算）；預設：{DEFAULT_HEADER_ROW}",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="顯示會刪除的資料，但不寫入 Excel 或 HTML",
    )

    return parser


def read_exact_folders(
    positional_folders: Sequence[str],
    option_folders: Sequence[str],
    folder_files: Sequence[Path],
) -> set[str]:
    folders = {
        folder.strip()
        for folder in (*positional_folders, *option_folders)
        if folder.strip()
    }

    for path in folder_files:
        if not path.is_file():
            print(f"[SKIP] 找不到 folder 清單檔：{path.resolve()}")
            continue

        with path.open("r", encoding="utf-8-sig") as file:
            for raw_line in file:
                line = raw_line.strip()
                if line and not line.startswith("#"):
                    folders.add(line)

    return folders


def make_matcher(args: argparse.Namespace, parser: argparse.ArgumentParser) -> FolderMatcher:
    exact_folders = read_exact_folders(
        args.folders,
        args.folder,
        args.folder_file,
    )

    prefixes = tuple(prefix.strip() for prefix in args.prefix if prefix.strip())
    regexes: list[Pattern[str]] = []

    for pattern_text in args.regex:
        try:
            regexes.append(re.compile(pattern_text))
        except re.error as error:
            parser.error(f"無效的 --regex {pattern_text!r}：{error}")

    if not exact_folders and not prefixes and not regexes:
        parser.error(
            "至少要指定一個 FOLDER、--folder、--folder-file、"
            "--prefix 或 --regex。"
        )

    return FolderMatcher(
        exact_folders=frozenset(exact_folders),
        prefixes=prefixes,
        regexes=tuple(regexes),
    )


def should_delete(folder_value: object, matcher: FolderMatcher) -> bool:
    if folder_value is None:
        return False

    folder = str(folder_value).strip()
    return bool(folder) and matcher.matches(folder)


def find_folder_column(worksheet, header_row: int) -> int | None:
    for column_index in range(1, worksheet.max_column + 1):
        header = worksheet.cell(row=header_row, column=column_index).value
        if header is not None and str(header).strip().lower() == "folder":
            return column_index

    return None


def delete_matching_rows(
    worksheet,
    matcher: FolderMatcher,
    header_row: int = DEFAULT_HEADER_ROW,
) -> list[str] | None:
    """Return deleted folder names, or None when the folder column is absent."""
    folder_column = find_folder_column(worksheet, header_row)

    if folder_column is None:
        return None

    rows_to_delete: list[int] = []
    deleted_folders: list[str] = []

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        folder_value = worksheet.cell(row=row_index, column=folder_column).value

        if should_delete(folder_value, matcher):
            rows_to_delete.append(row_index)
            deleted_folders.append(str(folder_value).strip())

    # Rows must be deleted from bottom to top so earlier indices remain valid.
    for row_index in reversed(rows_to_delete):
        worksheet.delete_rows(row_index, 1)

    return deleted_folders


def atomic_save_workbook(workbook, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output_path.name}.",
        suffix=".tmp",
        dir=output_path.parent,
    )
    os.close(file_descriptor)
    temporary_path = Path(temporary_name)

    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
    finally:
        temporary_path.unlink(missing_ok=True)


def process_excel(
    input_path: Path,
    output_path: Path,
    sheet_names: Sequence[str],
    header_row: int,
    matcher: FolderMatcher,
    dry_run: bool,
) -> Counter[str]:
    deleted = Counter()

    if not input_path.is_file():
        print(f"[SKIP][Excel] 找不到檔案：{input_path.resolve()}")
        return deleted

    workbook = load_workbook(
        input_path,
        keep_vba=input_path.suffix.lower() == ".xlsm",
    )

    try:
        for sheet_name in sheet_names:
            if sheet_name not in workbook.sheetnames:
                print(f"[SKIP][Excel] 找不到工作表：{sheet_name}")
                continue

            result = delete_matching_rows(
                workbook[sheet_name],
                matcher,
                header_row=header_row,
            )

            if result is None:
                print(
                    f"[SKIP][Excel] 工作表 {sheet_name!r} "
                    f"第 {header_row} 列找不到 'folder' 欄位"
                )
                continue

            if not result:
                print(f"[SKIP][Excel] 工作表 {sheet_name!r} 找不到符合資料")
                continue

            deleted.update(result)
            print(f"[Excel] 工作表 {sheet_name!r}：刪除 {len(result)} 列")
            for folder in result:
                print(f"  - {folder}")

        if not deleted:
            print("[SKIP][Excel] 所有指定工作表皆無符合資料，不寫入檔案")
        elif dry_run:
            print(f"[DRY RUN][Excel] 不寫入：{output_path.resolve()}")
        else:
            atomic_save_workbook(workbook, output_path)
            print(f"[OK][Excel] 已寫入：{output_path.resolve()}")
    finally:
        workbook.close()

    return deleted


HTML_ROW_PATTERN = re.compile(
    # Prefer the normal </tr>.  The lookaheads also support generated HTML
    # whose row ends immediately before the next <tr> or </table>.
    r"<tr\b[^>]*>.*?(?:</tr\s*>|(?=<tr\b)|(?=</table\s*>))",
    flags=re.IGNORECASE | re.DOTALL,
)
FIRST_CELL_PATTERN = re.compile(
    r"\A\s*<tr\b[^>]*>\s*<td\b[^>]*>(.*?)</td\s*>",
    flags=re.IGNORECASE | re.DOTALL,
)
HTML_TAG_PATTERN = re.compile(r"<[^>]+>", flags=re.DOTALL)


def first_cell_text(row_html: str) -> str | None:
    match = FIRST_CELL_PATTERN.search(row_html)
    if match is None:
        return None

    cell_html = match.group(1)
    text = HTML_TAG_PATTERN.sub("", cell_html)
    return html.unescape(text).strip()


def delete_matching_html_rows(
    html_text: str,
    matcher: FolderMatcher,
) -> tuple[str, list[str]]:
    deleted_folders: list[str] = []

    def replace_row(match: re.Match[str]) -> str:
        row_html = match.group(0)
        folder = first_cell_text(row_html)

        if folder is not None and matcher.matches(folder):
            deleted_folders.append(folder)
            return ""

        return row_html

    new_html = HTML_ROW_PATTERN.sub(replace_row, html_text)
    return new_html, deleted_folders


def read_html(path: Path) -> tuple[str, str]:
    raw_bytes = path.read_bytes()
    encoding = "utf-8-sig" if raw_bytes.startswith(b"\xef\xbb\xbf") else "utf-8"
    return raw_bytes.decode(encoding), encoding


def atomic_write_html(path: Path, text: str, encoding: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    data = text.encode(encoding)

    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.",
        suffix=".tmp",
        dir=path.parent,
    )
    temporary_path = Path(temporary_name)

    try:
        with os.fdopen(file_descriptor, "wb") as file:
            file.write(data)
            file.flush()
            os.fsync(file.fileno())
        os.replace(temporary_path, path)
    finally:
        temporary_path.unlink(missing_ok=True)


def process_html(
    input_path: Path,
    output_path: Path,
    matcher: FolderMatcher,
    dry_run: bool,
) -> Counter[str]:
    deleted = Counter()

    if not input_path.is_file():
        print(f"[SKIP][HTML] 找不到檔案：{input_path.resolve()}")
        return deleted

    html_text, encoding = read_html(input_path)
    new_html, deleted_folders = delete_matching_html_rows(html_text, matcher)
    deleted.update(deleted_folders)

    if not deleted:
        print("[SKIP][HTML] 找不到符合的 <tr> 資料，不寫入檔案")
    else:
        print(f"[HTML] 刪除 {len(deleted_folders)} 筆 <tr> 資料")
        for folder in deleted_folders:
            print(f"  - {folder}")

        if dry_run:
            print(f"[DRY RUN][HTML] 不寫入：{output_path.resolve()}")
        else:
            atomic_write_html(output_path, new_html, encoding)
            print(f"[OK][HTML] 已寫入：{output_path.resolve()}")

    return deleted


def print_exact_folder_summary(
    exact_folders: Sequence[str],
    excel_deleted: Counter[str],
    html_deleted: Counter[str],
) -> None:
    if not exact_folders:
        return

    print("=" * 60)
    print("完整 folder 名稱處理摘要：")

    for folder in sorted(exact_folders):
        excel_count = excel_deleted[folder]
        html_count = html_deleted[folder]

        if excel_count == 0 and html_count == 0:
            print(f"[SKIP] {folder}：Excel 與 HTML 都找不到資料")
        else:
            print(
                f"[OK] {folder}：Excel {excel_count} 列，"
                f"HTML {html_count} 筆"
            )


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    if args.header_row < 1:
        parser.error("--header-row 必須大於或等於 1")

    matcher = make_matcher(args, parser)
    sheet_names = args.sheet_names or DEFAULT_SHEET_NAMES
    excel_output = args.excel_output or args.excel
    html_output = args.html_output or args.html

    excel_deleted = process_excel(
        input_path=args.excel,
        output_path=excel_output,
        sheet_names=sheet_names,
        header_row=args.header_row,
        matcher=matcher,
        dry_run=args.dry_run,
    )
    html_deleted = process_html(
        input_path=args.html,
        output_path=html_output,
        matcher=matcher,
        dry_run=args.dry_run,
    )

    print_exact_folder_summary(
        matcher.exact_folders,
        excel_deleted,
        html_deleted,
    )

    print("=" * 60)
    print(f"Excel 共刪除 {sum(excel_deleted.values())} 列")
    print(f"HTML 共刪除 {sum(html_deleted.values())} 筆 <tr> 資料")

    if args.dry_run:
        print("DRY RUN：未修改任何檔案")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

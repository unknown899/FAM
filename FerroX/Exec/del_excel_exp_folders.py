from pathlib import Path
import re

from openpyxl import load_workbook


# ==========================
# 使用者設定
# ==========================
INPUT_PATH = Path("MFIS_dataset.xlsx")
OUTPUT_PATH = Path("MFIS_dataset.xlsx")

SHEET_NAMES = [
    "experiments",
    "pv_curve",
    "pz_stack_index",
]

HEADER_ROW = 1

# 要刪除的 MFIS_t_7_nomi_(正整數)
DELETE_NOMI_NUMBERS = {
    8, 9, 10, 11, 12, 13, 14, 15, 16, 17
}


# ==========================
# 判斷 folder 是否需要刪除
# ==========================
def should_delete(folder_value) -> bool:
    if folder_value is None:
        return False

    folder = str(folder_value).strip()

    # 條件 1
    if folder == "MFIS_t_8_nomi_31":
        return True

    # 條件 2：
    # MFIS_t_7_nomi_(指定正整數)
    match = re.fullmatch(r"MFIS_t_7_nomi_(\d+)", folder)

    if match:
        number = int(match.group(1))

        if number in DELETE_NOMI_NUMBERS:
            return True

    # 條件 3：
    # MFIS_t_7_nomi_4_1e-6
    # MFIS_t_7_nomi_5_1e-6
    if re.fullmatch(r"MFIS_t_7_nomi_(4|5)_1e-6", folder):
        return True

    # 條件 4：
    # MFIS_t_7_nomi_4_bg=50
    # MFIS_t_7_nomi_4_bg=200
    if re.fullmatch(r"MFIS_t_7_nomi_4_bg=(50|200)", folder):
        return True

    return False


# ==========================
# 處理單一工作表
# ==========================
def delete_matching_rows(worksheet, header_row=1):
    # 找出 folder 欄位
    header_to_column = {}

    for column_index in range(1, worksheet.max_column + 1):
        header = worksheet.cell(
            row=header_row,
            column=column_index,
        ).value

        if header is not None:
            header_name = str(header).strip().lower()
            header_to_column[header_name] = column_index

    if "folder" not in header_to_column:
        raise ValueError(
            f"工作表 {worksheet.title!r} 的第 {header_row} 列"
            f"找不到 'folder' 欄位。\n"
            f"目前欄位：{list(header_to_column.keys())}"
        )

    folder_column = header_to_column["folder"]

    # 找出需要刪除的列
    rows_to_delete = []
    deleted_folders = []

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        folder = worksheet.cell(
            row=row_index,
            column=folder_column,
        ).value

        if should_delete(folder):
            rows_to_delete.append(row_index)
            deleted_folders.append(str(folder).strip())

    # 必須從下面往上刪除
    for row_index in reversed(rows_to_delete):
        worksheet.delete_rows(row_index, 1)

    return rows_to_delete, deleted_folders


# ==========================
# 讀取 Excel
# ==========================
if not INPUT_PATH.exists():
    raise FileNotFoundError(f"找不到檔案：{INPUT_PATH.resolve()}")

workbook = load_workbook(INPUT_PATH)


# ==========================
# 逐一處理工作表
# ==========================
total_deleted = 0

for sheet_name in SHEET_NAMES:
    if sheet_name not in workbook.sheetnames:
        print(f"[跳過] 找不到工作表：{sheet_name}")
        continue

    worksheet = workbook[sheet_name]

    rows_to_delete, deleted_folders = delete_matching_rows(
        worksheet,
        header_row=HEADER_ROW,
    )

    total_deleted += len(rows_to_delete)

    print("=" * 60)
    print(f"工作表：{sheet_name}")
    print(f"刪除列數：{len(rows_to_delete)}")

    if deleted_folders:
        print("刪除的 folder：")

        for folder in deleted_folders:
            print(f"  - {folder}")
    else:
        print("沒有找到符合條件的資料列。")


# ==========================
# 儲存結果
# ==========================
workbook.save(OUTPUT_PATH)

print("=" * 60)
print(f"輸入檔案：{INPUT_PATH.resolve()}")
print(f"輸出檔案：{OUTPUT_PATH.resolve()}")
print(f"三個工作表總共刪除 {total_deleted} 列")
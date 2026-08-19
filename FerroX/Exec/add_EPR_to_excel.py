#!/usr/bin/env python3
"""Add intrinsic Landau EPR columns to a FerroX dataset workbook.

By default, ``--structure MFIS`` selects ``MFIS_dataset.xlsx``.  Any other
structure (for example MFIM, MFM, or a custom name) can be selected with
``--structure`` or an explicit ``--excel`` path.  Worksheet and source/output
column names are configurable.

The default source columns are ``alpha``, ``beta`` and ``gamma``. The script
adds or updates these columns:

    Ec, Pr, Pc0, rp, landau_consistency_pass

The operation is idempotent: rerunning the script updates existing result
columns instead of appending duplicates.
"""

from __future__ import annotations

import argparse
import math
import os
import sys
import tempfile
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any

try:
    from utils.landau_EPR_transformer import check_landau_EPR, landau_to_EPR
except ModuleNotFoundError as error:
    check_landau_EPR = None
    landau_to_EPR = None
    _TRANSFORM_IMPORT_ERROR = error
else:
    _TRANSFORM_IMPORT_ERROR = None


OUTPUT_HEADERS = (
    "Ec",
    "Pr",
    "Pc0",
    "rp",
    "landau_consistency_pass",
)

DEFAULT_STRUCTURE = "MFIS"
DEFAULT_SHEET_NAME = "experiments"

HELP_EPILOG = r"""
Examples:
  # MFIM_dataset.xlsx -> MFIM_dataset_with_EPR.xlsx
  python add_EPR_to_excel.py --structure MFIM

  # Update MFIM_dataset.xlsx atomically in place
  python add_EPR_to_excel.py --structure MFIM --in-place

  # Completely custom workbook, worksheet, and column names
  python add_EPR_to_excel.py \
      --excel data/custom.xlsx \
      --sheet runs \
      --alpha-column landau_a \
      --beta-column landau_b \
      --gamma-column landau_c \
      --ec-column intrinsic_Ec \
      --in-place
"""


@dataclass(frozen=True, slots=True)
class UpdateSummary:
    output_path: Path
    calculated_rows: int
    passed_rows: int
    failed_rows: int
    skipped_blank_rows: int
    row_errors: tuple[str, ...]


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return "".join(str(value).strip().casefold().split())


def _find_header_column(
    worksheet: Any,
    header_row: int,
    aliases: set[str],
) -> int:
    normalized_aliases = {_normalize_header(alias) for alias in aliases}
    matches = [
        column
        for column in range(1, worksheet.max_column + 1)
        if _normalize_header(worksheet.cell(header_row, column).value)
        in normalized_aliases
    ]
    if not matches:
        raise KeyError(
            f"Could not find any of {sorted(aliases)!r} in header row {header_row}"
        )
    if len(matches) > 1:
        raise KeyError(f"Multiple matching columns for {sorted(aliases)!r}: {matches}")
    return matches[0]


def _source_aliases(requested_name: str, symbol: str) -> set[str]:
    """Keep Greek aliases for the conventional Landau column names."""

    aliases = {requested_name}
    conventional = {
        "α": "alpha",
        "β": "beta",
        "γ": "gamma",
    }
    if _normalize_header(requested_name) in {
        _normalize_header(symbol),
        conventional[symbol],
    }:
        aliases.update({symbol, conventional[symbol]})
    return aliases


def _resolve_under_root(path: Path, root: Path) -> Path:
    return path if path.is_absolute() else root / path


def _ensure_output_column(
    worksheet: Any,
    header_row: int,
    name: str,
    style_source_column: int,
) -> tuple[int, bool]:
    normalized_name = _normalize_header(name)
    for column in range(1, worksheet.max_column + 1):
        if _normalize_header(worksheet.cell(header_row, column).value) == normalized_name:
            return column, False

    column = worksheet.max_column + 1
    source = worksheet.cell(header_row, style_source_column)
    target = worksheet.cell(header_row, column, name)
    if source.has_style:
        target._style = copy(source._style)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    return column, True


def _copy_style(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def _to_float(value: Any, name: str) -> float:
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError(f"missing {name}")
    if isinstance(value, bool):
        raise ValueError(f"{name} is Boolean, not numeric")
    if isinstance(value, str) and value.lstrip().startswith("="):
        raise ValueError(f"{name} formula has no cached numeric value")
    try:
        result = float(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"invalid {name}: {value!r}") from error
    if not math.isfinite(result):
        raise ValueError(f"{name} must be finite; got {result!r}")
    return result


def _extend_filter_and_tables(
    worksheet: Any,
    *,
    header_row: int,
    old_max_column: int,
    new_max_column: int,
) -> None:
    if new_max_column <= old_max_column:
        return

    from openpyxl.utils import get_column_letter, range_boundaries
    from openpyxl.worksheet.table import TableColumn

    def extend_ref(reference: str) -> str:
        min_col, min_row, max_col, max_row = range_boundaries(reference)
        if min_row == header_row and max_col == old_max_column:
            max_col = new_max_column
        return (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )

    if worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = extend_ref(worksheet.auto_filter.ref)

    for table in worksheet.tables.values():
        _, min_row, max_col, _ = range_boundaries(table.ref)
        if min_row != header_row or max_col != old_max_column:
            continue

        next_id = max((column.id for column in table.tableColumns), default=0) + 1
        for column in range(old_max_column + 1, new_max_column + 1):
            table.tableColumns.append(
                TableColumn(
                    id=next_id,
                    name=str(worksheet.cell(header_row, column).value),
                )
            )
            next_id += 1
        table.ref = extend_ref(table.ref)
        if table.autoFilter is not None and table.autoFilter.ref:
            table.autoFilter.ref = extend_ref(table.autoFilter.ref)


def add_EPR_to_workbook(
    input_path: str | Path = "MFIS_dataset.xlsx",
    output_path: str | Path | None = None,
    *,
    sheet_name: str = "experiments",
    header_row: int = 1,
    rtol: float = 1.0e-9,
    atol: float = 0.0,
    overwrite: bool = False,
    alpha_column: str = "alpha",
    beta_column: str = "beta",
    gamma_column: str = "gamma",
    ec_column: str = "Ec",
    pr_column: str = "Pr",
    pc0_column: str = "Pc0",
    rp_column: str = "rp",
    consistency_column: str = "landau_consistency_pass",
    style_source_column: str | None = None,
) -> UpdateSummary:
    """Calculate and write EPR quantities for every valid experiments row."""

    if landau_to_EPR is None or check_landau_EPR is None:
        raise RuntimeError(
            "Cannot import utils.landau_EPR_transformer. Put this script "
            "beside the utils/ directory or install that module."
        ) from _TRANSFORM_IMPORT_ERROR

    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as error:
        raise RuntimeError(
            "openpyxl is required: python -m pip install openpyxl"
        ) from error

    source = Path(input_path).expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(source)
    if source.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise ValueError("Only .xlsx and .xlsm workbooks are supported")
    if header_row < 1:
        raise ValueError("header_row must be at least 1")

    if output_path is None:
        destination = source.with_name(
            f"{source.stem}_with_EPR{source.suffix}"
        )
    else:
        destination = Path(output_path).expanduser().resolve()
    if destination.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise ValueError("Output must use the .xlsx or .xlsm extension")

    same_file = destination == source
    if destination.exists() and not (overwrite or same_file):
        raise FileExistsError(
            f"Output already exists: {destination}; use --overwrite"
        )

    keep_vba = source.suffix.casefold() == ".xlsm"
    workbook = load_workbook(
        source,
        data_only=False,
        keep_vba=keep_vba,
        keep_links=True,
    )
    value_workbook = load_workbook(
        source,
        data_only=True,
        read_only=True,
        keep_vba=keep_vba,
        keep_links=True,
    )

    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(
                f"Worksheet {sheet_name!r} not found; available: {workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        value_worksheet = value_workbook[sheet_name]
        alpha_col = _find_header_column(
            worksheet,
            header_row,
            _source_aliases(alpha_column, "α"),
        )
        beta_col = _find_header_column(
            worksheet,
            header_row,
            _source_aliases(beta_column, "β"),
        )
        gamma_col = _find_header_column(
            worksheet,
            header_row,
            _source_aliases(gamma_column, "γ"),
        )
        style_col = (
            gamma_col
            if style_source_column is None
            else _find_header_column(
                worksheet,
                header_row,
                {style_source_column},
            )
        )

        output_names = {
            "Ec": ec_column,
            "Pr": pr_column,
            "Pc0": pc0_column,
            "rp": rp_column,
            "landau_consistency_pass": consistency_column,
        }
        normalized_output_names = [
            _normalize_header(name) for name in output_names.values()
        ]
        if any(not name for name in normalized_output_names):
            raise ValueError("Output column names cannot be empty")
        if len(set(normalized_output_names)) != len(normalized_output_names):
            raise ValueError("Output column names must be unique")
        source_columns = {alpha_col, beta_col, gamma_col}
        for logical_name, output_name in output_names.items():
            for source_column in source_columns:
                if _normalize_header(
                    worksheet.cell(header_row, source_column).value
                ) == _normalize_header(output_name):
                    raise ValueError(
                        f"Output column {logical_name}={output_name!r} collides "
                        "with a Landau source column"
                    )

        old_max_column = worksheet.max_column
        output_columns: dict[str, int] = {}
        new_columns: set[int] = set()
        for logical_name in OUTPUT_HEADERS:
            column, is_new = _ensure_output_column(
                worksheet,
                header_row,
                output_names[logical_name],
                style_col,
            )
            output_columns[logical_name] = column
            if is_new:
                new_columns.add(column)

        _extend_filter_and_tables(
            worksheet,
            header_row=header_row,
            old_max_column=old_max_column,
            new_max_column=worksheet.max_column,
        )

        widths = {
            "Ec": 18.0,
            "Pr": 18.0,
            "Pc0": 18.0,
            "rp": 13.0,
            "landau_consistency_pass": 26.0,
        }
        for header, column in output_columns.items():
            if column in new_columns:
                worksheet.column_dimensions[get_column_letter(column)].width = widths[header]

        calculated = passed = failed = skipped = 0
        errors: list[str] = []

        for row in range(header_row + 1, worksheet.max_row + 1):
            raw = (
                value_worksheet.cell(row, alpha_col).value,
                value_worksheet.cell(row, beta_col).value,
                value_worksheet.cell(row, gamma_col).value,
            )
            cells = {
                header: worksheet.cell(row, column)
                for header, column in output_columns.items()
            }

            if all(
                value is None or (isinstance(value, str) and not value.strip())
                for value in raw
            ):
                for cell in cells.values():
                    cell.value = None
                skipped += 1
                continue

            for column in new_columns:
                _copy_style(
                    worksheet.cell(row, style_col),
                    worksheet.cell(row, column),
                )

            try:
                alpha = _to_float(raw[0], alpha_column)
                beta = _to_float(raw[1], beta_column)
                gamma = _to_float(raw[2], gamma_column)
                epr = landau_to_EPR(alpha, beta, gamma)
                check = check_landau_EPR(
                    alpha,
                    beta,
                    gamma,
                    epr.Ec,
                    epr.Pr,
                    epr.Pc0,
                    epr.rp,
                    rtol=rtol,
                    atol=atol,
                )

                cells["Ec"].value = epr.Ec
                cells["Pr"].value = epr.Pr
                cells["Pc0"].value = epr.Pc0
                cells["rp"].value = epr.rp
                cells["landau_consistency_pass"].value = check.passed

                for header in ("Ec", "Pr", "Pc0"):
                    cells[header].number_format = "0.000000E+00"
                cells["rp"].number_format = "0.000000"
                cells["landau_consistency_pass"].number_format = "General"

                calculated += 1
                if check.passed:
                    passed += 1
                else:
                    failed += 1
                    errors.append(
                        f"row {row}: {check.transition_order} consistency check failed"
                    )
            except (ArithmeticError, OverflowError, ValueError) as error:
                for header in ("Ec", "Pr", "Pc0", "rp"):
                    cells[header].value = None
                cells["landau_consistency_pass"].value = False
                failed += 1
                errors.append(f"row {row}: {error}")

        destination.parent.mkdir(parents=True, exist_ok=True)
        descriptor, temporary_name = tempfile.mkstemp(
            prefix=f".{destination.stem}.",
            suffix=destination.suffix,
            dir=destination.parent,
        )
        os.close(descriptor)
        temporary = Path(temporary_name)
        try:
            workbook.save(temporary)
            os.replace(temporary, destination)
        finally:
            temporary.unlink(missing_ok=True)

        return UpdateSummary(
            output_path=destination,
            calculated_rows=calculated,
            passed_rows=passed,
            failed_rows=failed,
            skipped_blank_rows=skipped,
            row_errors=tuple(errors),
        )
    finally:
        value_workbook.close()
        workbook.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Add Ec, Pr, Pc0, rp and a Landau consistency flag to any "
            "FerroX dataset workbook."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=HELP_EPILOG,
    )
    parser.add_argument(
        "--structure",
        default=DEFAULT_STRUCTURE,
        help=(
            "Structure label used only when --excel is omitted; selects "
            "<structure>_dataset.xlsx (default: MFIS)."
        ),
    )
    parser.add_argument(
        "--root",
        type=Path,
        default=Path.cwd(),
        help="Base directory for relative paths (default: cwd).",
    )
    parser.add_argument(
        "--input",
        "--excel",
        dest="input",
        type=Path,
        default=None,
        help="Input workbook; default: <structure>_dataset.xlsx under --root.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Default: <input_stem>_with_EPR.xlsx",
    )
    parser.add_argument(
        "--in-place",
        action="store_true",
        help="Atomically replace the input workbook.",
    )
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME)
    parser.add_argument("--header-row", type=int, default=1)
    parser.add_argument("--alpha-column", default="alpha")
    parser.add_argument("--beta-column", default="beta")
    parser.add_argument("--gamma-column", default="gamma")
    parser.add_argument("--ec-column", default="Ec")
    parser.add_argument("--pr-column", default="Pr")
    parser.add_argument("--pc0-column", default="Pc0")
    parser.add_argument("--rp-column", default="rp")
    parser.add_argument(
        "--consistency-column",
        default="landau_consistency_pass",
    )
    parser.add_argument(
        "--style-source-column",
        default=None,
        help="Column whose formatting is copied to new columns; default: gamma column.",
    )
    parser.add_argument("--rtol", type=float, default=1.0e-9)
    parser.add_argument("--atol", type=float, default=0.0)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    if args.in_place and args.output:
        parser.error("--in-place and --output cannot be used together")
    if not args.structure.strip():
        parser.error("--structure cannot be empty")

    root = args.root.expanduser().resolve()
    if not root.is_dir():
        parser.error(f"--root is not a directory: {root}")

    input_argument = (
        args.input
        if args.input is not None
        else Path(f"{args.structure}_dataset.xlsx")
    )
    input_path = _resolve_under_root(input_argument.expanduser(), root).resolve()
    output_path = (
        None
        if args.output is None
        else _resolve_under_root(args.output.expanduser(), root).resolve()
    )

    output = input_path if args.in_place else output_path
    try:
        summary = add_EPR_to_workbook(
            input_path,
            output,
            sheet_name=args.sheet,
            header_row=args.header_row,
            rtol=args.rtol,
            atol=args.atol,
            overwrite=args.overwrite,
            alpha_column=args.alpha_column,
            beta_column=args.beta_column,
            gamma_column=args.gamma_column,
            ec_column=args.ec_column,
            pr_column=args.pr_column,
            pc0_column=args.pc0_column,
            rp_column=args.rp_column,
            consistency_column=args.consistency_column,
            style_source_column=args.style_source_column,
        )
    except Exception as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 1

    print(f"Saved: {summary.output_path}")
    print(
        f"Rows: calculated={summary.calculated_rows}, "
        f"passed={summary.passed_rows}, failed={summary.failed_rows}, "
        f"blank/skipped={summary.skipped_blank_rows}"
    )
    for message in summary.row_errors[:20]:
        print(f"WARNING: {message}", file=sys.stderr)
    if len(summary.row_errors) > 20:
        print(
            f"WARNING: {len(summary.row_errors) - 20} additional errors omitted",
            file=sys.stderr,
        )
    return 0 if summary.failed_rows == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())

#!/usr/bin/env python3
"""Calculate P-V loop validity for any FerroX structure and update Excel.

Use ``python add_valid_loop_to_excel.py --help`` for the complete command
workflow and copyable examples.

By default ``--structure MFIS`` selects ``MFIS_dataset.xlsx``, directories
matching ``MFIS*``, and recursively discovered ``MFIS_PV_curve.csv`` files.
Use ``--structure MFIM`` (or MFM, MIFM, ...) to change those defaults, or
override the workbook, folder glob, CSV filename/path, worksheet, and columns.

For each selected case, the default 37-point convention defines:

    P_start = -P_mean at zero-based index 0
    P_r-    = -P_mean at zero-based index 9
    P_r+    = -P_mean at zero-based index 27
    P_end   = -P_mean at index -1 (last row)

All four indices and the sign multiplier are configurable.

The loop is valid when all of the following are true:

    abs(P_r+ - P_r-) > open_threshold
    P_r+ > 0 and P_r- < 0
    abs(P_start - P_end) < close_threshold

Only worksheet rows whose ``folder`` matches a selected directory are
changed. Other rows are left untouched. If a matching case has no CSV, its
``valid_loop`` value is set to ``no_data`` and processing continues.

With ``--dry-run``, the real workbook is not changed.  A dashboard preview can
still be generated when ``build_dash.py`` is available; use ``--no-preview``
for a dependency-free dry run.
"""

from __future__ import annotations

import argparse
import csv
import math
import os
import tempfile
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_STRUCTURE = "MFIS"
DEFAULT_SHEET_NAME = "experiments"
DEFAULT_FILE_COLUMN = "folder"
DEFAULT_VALID_COLUMN = "valid_loop"
DEFAULT_PREVIEW_HTML = "index.html"
DEFAULT_P_COLUMN = "P_mean"
MISSING_DATA_VALUE = "no_data"

HELP_EPILOG = r"""
使用流程:
  1. 先用 --dry-run 產生 HTML，確認 valid_loop 與對應圖片。
  2. 視需要調整 threshold，重新執行 dry-run。
  3. 確認後移除 --dry-run，才會寫入指定 Excel。

結果值:
  1       三項 loop 條件皆通過
  0       至少一項 loop 條件未通過
  no_data 找不到該 case 的 P-V CSV

Excel 欄位:
  新欄預設優先繼承 gamma 欄格式；若沒有 gamma，改用 case-name 欄。
  也可用 --style-source-column 明確指定。

範例 1：預覽 MFIM cases，不寫入 Excel
  python add_valid_loop_to_excel.py \
      --structure MFIM \
      --open-threshold 0.1 \
      --close-threshold 0.02 \
      --dry-run --no-preview

範例 2：正式寫入 MFIM_dataset.xlsx
  python add_valid_loop_to_excel.py \
      --structure MFIM \
      --open-threshold 0.1 \
      --close-threshold 0.02

範例 3：自訂資料夾、CSV、工作表、欄位與取樣位置
  python add_valid_loop_to_excel.py \
      --excel custom.xlsx \
      --case-glob 'run_*' \
      --pv-filename hysteresis.csv \
      --sheet runs \
      --case-column case_id \
      --p-column polarization \
      --start-index 0 --pr-minus-index 20 \
      --pr-plus-index 60 --end-index -1 \
      --p-sign 1 \
      --open-threshold 0.1 \
      --close-threshold 0.02

提示:
  --dry_run 等同 --dry-run
  --excel-name 等同 --excel
  --preview-html 等同 --html
"""


def nonnegative_float(text: str) -> float:
    """argparse converter for a finite, non-negative threshold."""

    try:
        value = float(text)
    except ValueError as exc:
        raise argparse.ArgumentTypeError(f"not a number: {text!r}") from exc

    if not math.isfinite(value) or value < 0:
        raise argparse.ArgumentTypeError(
            f"threshold must be finite and >= 0, got {text!r}"
        )
    return value


def finite_float(text: str) -> float:
    try:
        value = float(text)
    except ValueError as exc:
        raise argparse.ArgumentTypeError(f"not a number: {text!r}") from exc
    if not math.isfinite(value):
        raise argparse.ArgumentTypeError(f"must be finite, got {text!r}")
    return value


def positive_int(text: str) -> int:
    try:
        value = int(text)
    except ValueError as exc:
        raise argparse.ArgumentTypeError(f"not an integer: {text!r}") from exc
    if value < 1:
        raise argparse.ArgumentTypeError(f"must be >= 1, got {text!r}")
    return value


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Set a loop-validity column using configurable case folders and "
            "P-V CSV files."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=HELP_EPILOG,
    )
    parser.add_argument(
        "--structure",
        default=DEFAULT_STRUCTURE,
        help=(
            "Structure label used by defaults: <structure>_dataset.xlsx, "
            "<structure>*, and <structure>_PV_curve.csv (default: MFIS)."
        ),
    )
    selection = parser.add_mutually_exclusive_group()
    selection.add_argument(
        "--case-glob",
        default=None,
        help="Glob for case directories under --root; default: <structure>*.",
    )
    selection.add_argument(
        "--prefix",
        default=None,
        help="Legacy shorthand for --case-glob '<prefix>*'.",
    )
    parser.add_argument(
        "--open-threshold",
        required=True,
        type=nonnegative_float,
        help="Required lower bound for abs(P_r_plus - P_r_minus).",
    )
    parser.add_argument(
        "--close-threshold",
        required=True,
        type=nonnegative_float,
        help="Required upper bound for abs(P_start - P_end).",
    )
    parser.add_argument(
        "--root",
        type=Path,
        default=Path.cwd(),
        help="Execution directory containing the Excel file and case folders (default: cwd).",
    )
    parser.add_argument(
        "--excel",
        "--excel-name",
        dest="excel",
        type=Path,
        default=None,
        help="Workbook path; default: <structure>_dataset.xlsx under --root.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Optional output workbook. By default, update --excel in place.",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET_NAME,
        help=f"Worksheet name (default: {DEFAULT_SHEET_NAME}).",
    )
    parser.add_argument(
        "--file-column",
        "--case-column",
        dest="file_column",
        default=DEFAULT_FILE_COLUMN,
        help=f"Column containing case-directory names (default: {DEFAULT_FILE_COLUMN}).",
    )
    parser.add_argument(
        "--valid-column",
        default=DEFAULT_VALID_COLUMN,
        help=f"Column to create/update (default: {DEFAULT_VALID_COLUMN}).",
    )
    parser.add_argument(
        "--header-row",
        type=positive_int,
        default=1,
        help="One-based worksheet header row (default: 1).",
    )
    parser.add_argument(
        "--style-source-column",
        default=None,
        help=(
            "Column whose style is copied to the output column. Default: "
            "gamma when present, otherwise the case-name column."
        ),
    )
    parser.add_argument(
        "--pv-filename",
        default=None,
        help=(
            "P-V CSV basename searched recursively in each case; default: "
            "<structure>_PV_curve.csv."
        ),
    )
    parser.add_argument(
        "--pv-relative-path",
        type=Path,
        default=None,
        help=(
            "Exact P-V CSV path relative to each case, for example "
            "figs/hysteresis.csv; overrides --pv-filename."
        ),
    )
    parser.add_argument(
        "--p-column",
        default=DEFAULT_P_COLUMN,
        help=f"Polarization column in each CSV (default: {DEFAULT_P_COLUMN}).",
    )
    parser.add_argument(
        "--p-sign",
        type=finite_float,
        default=-1.0,
        help="Multiply CSV polarization by this value (default: -1).",
    )
    parser.add_argument("--start-index", type=int, default=0)
    parser.add_argument("--pr-minus-index", type=int, default=9)
    parser.add_argument("--pr-plus-index", type=int, default=27)
    parser.add_argument(
        "--end-index",
        type=int,
        default=-1,
        help="Zero-based index; negative values count from the end (default: -1).",
    )
    parser.add_argument(
        "--missing-label",
        default=MISSING_DATA_VALUE,
        help=f"Value written when a selected case has no CSV (default: {MISSING_DATA_VALUE}).",
    )
    parser.add_argument(
        "--dry-run",
        "--dry_run",
        dest="dry_run",
        action="store_true",
        help=(
            "Write proposed values to a temporary workbook and generate "
            "a preview dashboard without changing the real workbook."
        ),
    )
    parser.add_argument(
        "--no-preview",
        action="store_true",
        help="With --dry-run, skip build_dash.py and HTML generation.",
    )
    parser.add_argument(
        "--preview-html",
        "--html",
        dest="preview_html",
        type=Path,
        default=Path(DEFAULT_PREVIEW_HTML),
        help=(
            "HTML generated during --dry-run, relative to --root unless "
            f"absolute (default: {DEFAULT_PREVIEW_HTML})."
        ),
    )
    args = parser.parse_args()
    if not args.structure.strip():
        parser.error("--structure cannot be empty")
    if args.pv_relative_path is not None and args.pv_relative_path.is_absolute():
        parser.error("--pv-relative-path must be relative to each case directory")
    for name in (
        "file_column",
        "valid_column",
        "p_column",
        "missing_label",
    ):
        if not str(getattr(args, name)).strip():
            parser.error(f"--{name.replace('_', '-')} cannot be empty")
    return args


def resolve_under_root(path: Path, root: Path) -> Path:
    return path if path.is_absolute() else root / path


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return "".join(str(value).strip().casefold().split())


def _find_header_column(
    worksheet: Any,
    header_row: int,
    aliases: set[str],
) -> int:
    normalized_aliases = {_normalize_header(alias) for alias in aliases}
    matches = [
        column
        for column in range(1, worksheet.max_column + 1)
        if _normalize_header(worksheet.cell(header_row, column).value)
        in normalized_aliases
    ]
    if not matches:
        raise KeyError(
            f"Could not find any of {sorted(aliases)!r} in header row {header_row}"
        )
    if len(matches) > 1:
        raise KeyError(f"Multiple matching columns for {sorted(aliases)!r}: {matches}")
    return matches[0]


def _ensure_output_column(
    worksheet: Any,
    header_row: int,
    name: str,
    style_source_column: int,
) -> tuple[int, bool]:
    normalized_name = _normalize_header(name)
    for column in range(1, worksheet.max_column + 1):
        if _normalize_header(worksheet.cell(header_row, column).value) == normalized_name:
            return column, False

    column = worksheet.max_column + 1
    source = worksheet.cell(header_row, style_source_column)
    target = worksheet.cell(header_row, column, name)
    if source.has_style:
        target._style = copy(source._style)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    return column, True


def _copy_style(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def _extend_filter_and_tables(
    worksheet: Any,
    *,
    header_row: int,
    old_max_column: int,
    new_max_column: int,
) -> None:
    if new_max_column <= old_max_column:
        return

    from openpyxl.utils import get_column_letter, range_boundaries
    from openpyxl.worksheet.table import TableColumn

    def extend_ref(reference: str) -> str:
        min_col, min_row, max_col, max_row = range_boundaries(reference)
        if min_row == header_row and max_col == old_max_column:
            max_col = new_max_column
        return (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )

    if worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = extend_ref(worksheet.auto_filter.ref)

    for table in worksheet.tables.values():
        _, min_row, max_col, _ = range_boundaries(table.ref)
        if min_row != header_row or max_col != old_max_column:
            continue

        next_id = max((column.id for column in table.tableColumns), default=0) + 1
        for column in range(old_max_column + 1, new_max_column + 1):
            table.tableColumns.append(
                TableColumn(
                    id=next_id,
                    name=str(worksheet.cell(header_row, column).value),
                )
            )
            next_id += 1
        table.ref = extend_ref(table.ref)
        if table.autoFilter is not None and table.autoFilter.ref:
            table.autoFilter.ref = extend_ref(table.autoFilter.ref)


def _repair_existing_output_column(
    worksheet: Any,
    *,
    header_row: int,
    style_source_column: int,
    output_column: int,
    width: float,
) -> None:
    """Repair a column created by an older run but left outside the filter."""

    from openpyxl.utils import get_column_letter, range_boundaries
    from openpyxl.worksheet.table import TableColumn

    _copy_style(
        worksheet.cell(header_row, style_source_column),
        worksheet.cell(header_row, output_column),
    )
    worksheet.column_dimensions[get_column_letter(output_column)].width = width

    def extend_ref(reference: str) -> str:
        min_col, min_row, max_col, max_row = range_boundaries(reference)
        if min_row == header_row and max_col < output_column:
            max_col = output_column
        return (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )

    if worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = extend_ref(worksheet.auto_filter.ref)

    for table in worksheet.tables.values():
        _, min_row, max_col, _ = range_boundaries(table.ref)
        if min_row != header_row or max_col >= output_column:
            continue

        next_id = max((column.id for column in table.tableColumns), default=0) + 1
        for column in range(max_col + 1, output_column + 1):
            table.tableColumns.append(
                TableColumn(
                    id=next_id,
                    name=str(worksheet.cell(header_row, column).value),
                )
            )
            next_id += 1
        table.ref = extend_ref(table.ref)
        if table.autoFilter is not None and table.autoFilter.ref:
            table.autoFilter.ref = extend_ref(table.autoFilter.ref)


def discover_case_directories(root: Path, case_glob: str) -> list[Path]:
    if not case_glob.strip():
        raise ValueError("Case-directory glob must not be empty")
    directories = sorted(
        (path for path in root.glob(case_glob) if path.is_dir()),
        key=lambda path: (path.name, str(path)),
    )
    by_name: dict[str, Path] = {}
    for directory in directories:
        previous = by_name.get(directory.name)
        if previous is not None:
            raise ValueError(
                f"Two selected case directories have the same basename "
                f"{directory.name!r}:\n  {previous}\n  {directory}"
            )
        by_name[directory.name] = directory
    return directories


def find_case_csv(
    case_directory: Path,
    *,
    pv_filename: str,
    pv_relative_path: Path | None,
) -> tuple[Path | None, Path]:
    """Return a unique CSV and the path used in missing-data messages."""

    if pv_relative_path is not None:
        expected = case_directory / pv_relative_path
        return (expected if expected.is_file() else None), expected

    matches = sorted(
        path for path in case_directory.rglob(pv_filename) if path.is_file()
    )
    if len(matches) > 1:
        details = "\n".join(f"  {path}" for path in matches)
        raise ValueError(
            f"{case_directory.name}: multiple {pv_filename!r} files found:\n"
            f"{details}\nUse --pv-relative-path to select one."
        )
    expected = case_directory / pv_filename
    return (matches[0] if matches else None), expected


def worksheet_rows_by_case(
    worksheet: Worksheet,
    file_column: int,
    header_row: int,
) -> dict[str, list[int]]:
    rows: dict[str, list[int]] = {}
    for row_number in range(header_row + 1, worksheet.max_row + 1):
        value = worksheet.cell(row=row_number, column=file_column).value
        if value is None:
            continue
        case_name = str(value).strip()
        if case_name:
            rows.setdefault(case_name, []).append(row_number)
    return rows


def find_style_source_column(
    worksheet: Worksheet,
    *,
    header_row: int,
    requested_name: str | None,
    fallback_column: int,
) -> int:
    if requested_name is not None:
        return _find_header_column(worksheet, header_row, {requested_name})
    try:
        return _find_header_column(worksheet, header_row, {"gamma", "γ"})
    except KeyError:
        return fallback_column


@dataclass(frozen=True, slots=True)
class LoopMetrics:
    valid_loop: int
    p_start: float
    p_r_minus: float
    p_r_plus: float
    p_end: float
    open_gap: float
    close_gap: float
    open_ok: bool
    sign_ok: bool
    close_ok: bool


class LoopAnalyzer:
    def __init__(
        self,
        *,
        open_threshold: float,
        close_threshold: float,
        p_column: str,
        p_sign: float,
        start_index: int,
        pr_minus_index: int,
        pr_plus_index: int,
        end_index: int,
    ) -> None:
        self.open_threshold = open_threshold
        self.close_threshold = close_threshold
        self.p_column = p_column
        self.p_sign = p_sign
        self.indices = {
            "P_start": start_index,
            "P_r_minus": pr_minus_index,
            "P_r_plus": pr_plus_index,
            "P_end": end_index,
        }

    @staticmethod
    def _resolve_index(index: int, size: int, label: str) -> int:
        resolved = index if index >= 0 else size + index
        if not 0 <= resolved < size:
            raise ValueError(
                f"{label} index {index} is outside a CSV with {size} data rows"
            )
        return resolved

    def analyze_csv(self, csv_path: Path) -> LoopMetrics:
        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames:
                raise ValueError("CSV has no header")
            normalized = {
                _normalize_header(name): name for name in reader.fieldnames
            }
            raw_column = normalized.get(_normalize_header(self.p_column))
            if raw_column is None:
                raise KeyError(
                    f"polarization column {self.p_column!r} not found; "
                    f"available: {reader.fieldnames}"
                )
            rows = list(reader)

        if not rows:
            raise ValueError("CSV contains no data rows")

        values: dict[str, float] = {}
        for label, index in self.indices.items():
            resolved = self._resolve_index(index, len(rows), label)
            raw_value = rows[resolved].get(raw_column)
            try:
                value = float(raw_value) * self.p_sign
            except (TypeError, ValueError) as error:
                raise ValueError(
                    f"{label} at data index {index} has invalid "
                    f"{raw_column}={raw_value!r}"
                ) from error
            if not math.isfinite(value):
                raise ValueError(
                    f"{label} at data index {index} is not finite: {value!r}"
                )
            values[label] = value

        p_start = values["P_start"]
        p_r_minus = values["P_r_minus"]
        p_r_plus = values["P_r_plus"]
        p_end = values["P_end"]
        open_gap = abs(p_r_plus - p_r_minus)
        close_gap = abs(p_start - p_end)
        open_ok = open_gap > self.open_threshold
        sign_ok = p_r_plus > 0 and p_r_minus < 0
        close_ok = close_gap < self.close_threshold
        return LoopMetrics(
            valid_loop=int(open_ok and sign_ok and close_ok),
            p_start=p_start,
            p_r_minus=p_r_minus,
            p_r_plus=p_r_plus,
            p_end=p_end,
            open_gap=open_gap,
            close_gap=close_gap,
            open_ok=open_ok,
            sign_ok=sign_ok,
            close_ok=close_ok,
        )


def save_workbook_atomically(workbook: object, output_path: Path) -> None:
    """Write beside the destination, then replace it only after save succeeds."""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output_path.stem}.",
        suffix=output_path.suffix,
        dir=output_path.parent,
    )
    os.close(descriptor)
    temporary_path = Path(temporary_name)

    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
    finally:
        temporary_path.unlink(missing_ok=True)


def format_result(case_name: str, metrics: LoopMetrics) -> str:
    return (
        f"{case_name}: valid_loop={metrics.valid_loop} | "
        f"P_r-={metrics.p_r_minus:.8g}, P_r+={metrics.p_r_plus:.8g}, "
        f"open_gap={metrics.open_gap:.8g} "
        f"({'PASS' if metrics.open_ok else 'FAIL'}) | "
        f"sign={'PASS' if metrics.sign_ok else 'FAIL'} | "
        f"P_start={metrics.p_start:.8g}, P_end={metrics.p_end:.8g}, "
        f"close_gap={metrics.close_gap:.8g} "
        f"({'PASS' if metrics.close_ok else 'FAIL'})"
    )


def run(args: argparse.Namespace) -> int:
    root = args.root.expanduser().resolve()
    if not root.is_dir():
        raise NotADirectoryError(f"Execution directory does not exist: {root}")

    excel_argument = (
        args.excel
        if args.excel is not None
        else Path(f"{args.structure}_dataset.xlsx")
    )
    excel_path = resolve_under_root(excel_argument.expanduser(), root).resolve()
    output_path = (
        excel_path
        if args.output is None
        else resolve_under_root(args.output.expanduser(), root).resolve()
    )
    preview_html_path = resolve_under_root(
        args.preview_html.expanduser(),
        root,
    ).resolve()
    if not excel_path.is_file():
        raise FileNotFoundError(f"Workbook does not exist: {excel_path}")
    if excel_path.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise ValueError("--excel must be an .xlsx or .xlsm workbook")
    if output_path.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise ValueError("--output must use the .xlsx or .xlsm extension")

    case_glob = (
        f"{args.prefix}*"
        if args.prefix is not None
        else (args.case_glob or f"{args.structure}*")
    )
    pv_filename = args.pv_filename or f"{args.structure}_PV_curve.csv"
    case_directories = discover_case_directories(root, case_glob)
    if not case_directories:
        raise FileNotFoundError(
            f"No case directories matched {case_glob!r} under {root}."
        )

    keep_vba = excel_path.suffix.casefold() == ".xlsm"
    workbook = load_workbook(excel_path, keep_vba=keep_vba, keep_links=True)
    try:
        if args.sheet not in workbook.sheetnames:
            raise KeyError(
                f"Worksheet {args.sheet!r} was not found. "
                f"Available worksheets: {workbook.sheetnames}"
            )
        worksheet = workbook[args.sheet]

        file_column = _find_header_column(
            worksheet,
            args.header_row,
            {args.file_column},
        )
        style_column = find_style_source_column(
            worksheet,
            header_row=args.header_row,
            requested_name=args.style_source_column,
            fallback_column=file_column,
        )

        old_max_column = worksheet.max_column
        valid_column, _ = _ensure_output_column(
            worksheet,
            args.header_row,
            args.valid_column,
            style_column,
        )
        _extend_filter_and_tables(
            worksheet,
            header_row=args.header_row,
            old_max_column=old_max_column,
            new_max_column=worksheet.max_column,
        )
        _repair_existing_output_column(
            worksheet,
            header_row=args.header_row,
            style_source_column=style_column,
            output_column=valid_column,
            width=14.0,
        )

        excel_rows = worksheet_rows_by_case(
            worksheet,
            file_column,
            args.header_row,
        )
        for row_numbers in excel_rows.values():
            for row_number in row_numbers:
                _copy_style(
                    worksheet.cell(row_number, style_column),
                    worksheet.cell(row_number, valid_column),
                )

        analyzer = LoopAnalyzer(
            open_threshold=args.open_threshold,
            close_threshold=args.close_threshold,
            p_column=args.p_column,
            p_sign=args.p_sign,
            start_index=args.start_index,
            pr_minus_index=args.pr_minus_index,
            pr_plus_index=args.pr_plus_index,
            end_index=args.end_index,
        )

        pending_updates: list[tuple[str, list[int], LoopMetrics]] = []
        missing_csv_updates: list[tuple[str, list[int], Path]] = []
        unmatched_directories: list[str] = []
        errors: list[str] = []

        for case_directory in case_directories:
            case_name = case_directory.name
            matching_rows = excel_rows.get(case_name)
            if not matching_rows:
                unmatched_directories.append(case_name)
                continue

            try:
                csv_path, expected_path = find_case_csv(
                    case_directory,
                    pv_filename=pv_filename,
                    pv_relative_path=args.pv_relative_path,
                )
                if csv_path is None:
                    missing_csv_updates.append(
                        (case_name, matching_rows, expected_path)
                    )
                    continue
                metrics = analyzer.analyze_csv(csv_path)
            except (OSError, KeyError, ValueError) as exc:
                errors.append(f"{case_name}: {exc}")
                continue

            pending_updates.append((case_name, matching_rows, metrics))

        if errors:
            details = "\n".join(f"  - {error}" for error in errors)
            raise RuntimeError(
                "No workbook changes were saved because one or more selected "
                f"cases could not be evaluated:\n{details}"
            )
        if not pending_updates and not missing_csv_updates:
            raise RuntimeError(
                "No selected case directory matched a value in worksheet "
                f"column {args.file_column!r}; no changes were saved."
            )

        valid_count = invalid_count = updated_row_count = 0
        for case_name, row_numbers, metrics in pending_updates:
            for row_number in row_numbers:
                target_cell = worksheet.cell(row=row_number, column=valid_column)
                target_cell.value = metrics.valid_loop
                target_cell.number_format = "General"
                updated_row_count += 1
            if metrics.valid_loop:
                valid_count += 1
            else:
                invalid_count += 1
            print(format_result(case_name, metrics))

        for case_name, row_numbers, expected_path in missing_csv_updates:
            for row_number in row_numbers:
                target_cell = worksheet.cell(row=row_number, column=valid_column)
                target_cell.value = args.missing_label
                target_cell.number_format = "General"
                updated_row_count += 1
            print(
                f"[NO DATA] {case_name}: CSV not found: {expected_path}; "
                f"{args.valid_column}={args.missing_label}"
            )

        for case_name in unmatched_directories:
            print(
                f"[WARNING] {case_name}: directory matched {case_glob!r} but "
                f"was not found in worksheet column {args.file_column!r}; skipped."
            )

        summary_text = (
            f"{updated_row_count} worksheet row(s): {valid_count} valid "
            f"case(s), {invalid_count} invalid case(s), "
            f"{len(missing_csv_updates)} case(s) without CSV."
        )
        if args.dry_run:
            if not args.no_preview:
                descriptor, temporary_name = tempfile.mkstemp(
                    prefix=f".{excel_path.stem}.valid_loop_preview.",
                    suffix=excel_path.suffix,
                    dir=root,
                )
                os.close(descriptor)
                temporary_excel_path = Path(temporary_name)
                try:
                    try:
                        import build_dash
                    except ImportError as exc:
                        raise RuntimeError(
                            "HTML preview requires build_dash.py; rerun with "
                            "--no-preview for a dependency-free dry run"
                        ) from exc
                    workbook.save(temporary_excel_path)
                    build_dash.build_dashboard(
                        temporary_excel_path,
                        data_root=root,
                        output_path=preview_html_path,
                        sheet_name=args.sheet,
                    )
                finally:
                    temporary_excel_path.unlink(missing_ok=True)
                print(f"Preview dashboard: {preview_html_path}")
            print(f"[DRY RUN] Would update {summary_text}")
        else:
            save_workbook_atomically(workbook, output_path)
            print(f"Saved {output_path} | updated {summary_text}")
        return 0
    finally:
        workbook.close()


def main() -> int:
    args = parse_args()
    try:
        return run(args)
    except (FileNotFoundError, NotADirectoryError, KeyError, ValueError, RuntimeError) as exc:
        raise SystemExit(f"ERROR: {exc}") from exc


if __name__ == "__main__":
    raise SystemExit(main())
